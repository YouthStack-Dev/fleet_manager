from fastapi import APIRouter, BackgroundTasks, Depends, HTTPException, status, Query, Request, UploadFile, File
from app.core.email_service import get_email_service, get_sms_service
from sqlalchemy.orm import Session
from typing import Optional, List, Dict, Any
from app.database.session import get_db
from app.models.employee import Employee
from app.schemas.employee import EmployeeCreate, EmployeeUpdate, EmployeeResponse, EmployeePaginationResponse
from app.utils.pagination import paginate_query
from common_utils.auth.permission_checker import PermissionChecker
from app.utils.response_utils import ResponseWrapper, handle_db_error, handle_http_error
from common_utils.auth.utils import hash_password
from app.crud.employee import employee_crud
from app.crud.team import team_crud
from app.crud.tenant import tenant_crud
from sqlalchemy.exc import SQLAlchemyError, IntegrityError
from app.core.logging_config import get_logger
from app.utils.audit_helper import log_audit
import openpyxl
from io import BytesIO
import re
from datetime import datetime

logger = get_logger(__name__)
router = APIRouter(prefix="/employees", tags=["employees"])

@router.post("/", status_code=status.HTTP_201_CREATED)
def create_employee(
    employee: EmployeeCreate,
    background_tasks: BackgroundTasks,
    request: Request,
    db: Session = Depends(get_db),
    user_data=Depends(PermissionChecker(["employee.create"], check_tenant=True)),
):
    """
    Create a new employee.

    Rules:
    - Vendors/Drivers → forbidden
    - Employees → tenant_id enforced from token
    - Admins → must provide tenant_id in payload

    Returns:
    - EmployeeResponse wrapped in ResponseWrapper
    """
    try:
        user_type = user_data.get("user_type")
        tenant_id = None

        # 🚫 Vendors/Drivers cannot create employees
        if user_type in {"vendor", "driver"}:
            raise HTTPException(
                status_code=status.HTTP_403_FORBIDDEN,
                detail=ResponseWrapper.error(
                    message="You don't have permission to create employees",
                    error_code="FORBIDDEN",
                ),
            )

        # 🔒 Tenant enforcement
        if user_type == "employee":
            tenant_id = user_data.get("tenant_id")
            if not tenant_id:
                raise HTTPException(
                    status_code=status.HTTP_403_FORBIDDEN,
                    detail=ResponseWrapper.error(
                        message="Tenant ID missing in token for employee",
                        error_code="TENANT_ID_REQUIRED",
                    ),
                )
        elif user_type == "admin":
            if not employee.tenant_id:
                raise HTTPException(
                    status_code=status.HTTP_400_BAD_REQUEST,
                    detail=ResponseWrapper.error(
                        message="Tenant ID is required for admin",
                        error_code="TENANT_ID_REQUIRED",
                    ),
                )
            tenant_id = employee.tenant_id
        logger.debug(f"Creating employee under tenant_id: {tenant_id}")
        tenant = tenant_crud.get_by_id(db=db, tenant_id=tenant_id)
        tenant_id = tenant.tenant_id if tenant else None
        if not tenant_id:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail=ResponseWrapper.error(
                    message=f"Tenant with ID {tenant_id} does not exist",
                    error_code="INVALID_TENANT_ID",
                ),
            )
        if not team_crud.is_team_in_tenant(db, team_id=employee.team_id, tenant_id=tenant_id):
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail=ResponseWrapper.error(
                    message=f"Team with ID {employee.team_id} does not belong to tenant {tenant_id}",
                    error_code="TEAM_TENANT_MISMATCH",
                ),
            )



        db_employee = employee_crud.create_with_tenant(db=db, obj_in=employee, tenant_id=tenant_id)
        logger.debug(f"Created employee object: {db_employee}")
        if not db_employee:
            raise HTTPException(
                status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
                detail=ResponseWrapper.error(
                    message="Failed to create employee",
                    error_code="EMPLOYEE_CREATION_FAILED",
                ),
            )
        # 🔥 Add background email task
        background_tasks.add_task(
            send_employee_created_email,
            employee_data={
                "name": db_employee.name,
                "email": db_employee.email,
                "phone": db_employee.phone,
                "employee_id": db_employee.employee_id,
                "tenant_id": tenant_id,
                "team_id": db_employee.team_id,
            },
        )
        db.commit()
        db.refresh(db_employee)

        # 🔍 Audit Log: Employee Creation
        try:
            employee_data_for_audit = {
                "employee_id": db_employee.employee_id,
                "name": db_employee.name,
                "email": db_employee.email,
                "phone": db_employee.phone,
                "employee_code": db_employee.employee_code,
                "team_id": db_employee.team_id,
                "is_active": db_employee.is_active
            }
            log_audit(
                db=db,
                tenant_id=tenant_id,
                module="employee",
                action="CREATE",
                user_data=user_data,
                description=f"Created employee '{db_employee.name}' ({db_employee.email})",
                new_values=employee_data_for_audit,
                request=request
            )
            logger.info(f"Audit log created for employee creation: employee_id={db_employee.employee_id} stored in audit_data with fields: {list(employee_data_for_audit.keys())}")
        except Exception as audit_error:
            logger.error(f"Failed to create audit log for employee creation: {str(audit_error)}")

        logger.info(
            f"Employee created successfully under tenant {tenant_id}: "
            f"employee_id={db_employee.employee_id}, name={db_employee.name}"
        )

        employee_response = EmployeeResponse.model_validate(db_employee, from_attributes=True).model_dump()
        # Add tenant location details
        if db_employee.tenant:
            employee_response["tenant_latitude"] = float(db_employee.tenant.latitude) if db_employee.tenant.latitude else None
            employee_response["tenant_longitude"] = float(db_employee.tenant.longitude) if db_employee.tenant.longitude else None
            employee_response["tenant_address"] = db_employee.tenant.address

        return ResponseWrapper.success(
            data={"employee": employee_response},
            message="Employee created successfully",
        )   

    except SQLAlchemyError as e:
        raise handle_db_error(e)
    except HTTPException:
        raise
    except Exception as e:
        logger.exception(f"Unexpected error while creating employee: {str(e)}")
        raise handle_http_error(e)

def send_employee_created_email(employee_data: dict):
    """Background task to send employee creation email and SMS."""
    try:
        email_service = get_email_service()
        sms_service = get_sms_service()

        # Send Email
        email_success = email_service.send_employee_created_email(
            user_email=employee_data["email"],
            user_name=employee_data["name"],
            details=employee_data,
        )

        if email_success:
            logger.info(f"Employee creation email sent: {employee_data['employee_id']}")
        else:
            logger.error(f"Employee creation email FAILED: {employee_data['employee_id']}")

        # Send SMS
        sms_message = (
            f"Welcome to {email_service.app_name}! "
            f"Your employee account has been created. "
            f"Employee ID: {employee_data['employee_id']}. "
            f"Login with your email: {employee_data['email']}"
        )
        
        sms_success = sms_service.send_sms(
            to_phone=employee_data["phone"],
            message=sms_message
        )

        if sms_success:
            logger.info(f"Employee creation SMS sent: {employee_data['employee_id']}")
        else:
            # Check if SMS is disabled vs actual failure
            if not sms_service.enabled:
                logger.info(f"Employee creation SMS skipped (service disabled): {employee_data['employee_id']}")
            else:
                logger.error(f"Employee creation SMS FAILED: {employee_data['employee_id']}")

    except Exception as e:
        logger.error(f"Error sending employee creation notifications: {str(e)}")


@router.get("/bulk-upload/template", status_code=status.HTTP_200_OK)
async def download_bulk_upload_template(
    user_data=Depends(PermissionChecker(["employee.create"], check_tenant=True)),
):
    """
    Download Excel template for bulk employee upload.
    
    Returns a pre-formatted Excel file with:
    - All required and optional column headers
    - Data type explanations in row 2
    - Sample data rows for reference
    
    This template can be filled and uploaded to /bulk-upload endpoint.
    """
    from fastapi.responses import StreamingResponse
    from openpyxl.styles import Font, PatternFill, Alignment
    
    try:
        # Create workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Employees"
        
        # Headers
        headers = ['name', 'email', 'phone', 'employee_code', 'team_id', 
                   'address', 'latitude', 'longitude', 'gender', 'password']
        
        # Style headers
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Add instructions row (row 2)
        instructions = [
            'Required',
            'Required (unique)',
            'Required (unique, 10+ digits)',
            'Optional (unique)',
            'Required (must exist)',
            'Optional',
            'Optional (-90 to 90)',
            'Optional (-180 to 180)',
            'Optional (Male/Female/Other)',
            'Optional (min 6 chars)'
        ]
        
        instruction_fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
        instruction_font = Font(italic=True, size=9)
        
        for col_num, instruction in enumerate(instructions, 1):
            cell = ws.cell(row=2, column=col_num)
            cell.value = instruction
            cell.fill = instruction_fill
            cell.font = instruction_font
            cell.alignment = Alignment(horizontal="left", vertical="center")
        
        # Add sample data (rows 3-5)
        sample_data = [
            ['John Doe', 'john.doe@company.com', '+1-234-567-8901', 'EMP001', '1', 
             '123 Main St, City, State', '37.7749', '-122.4194', 'Male', 'Welcome@123'],
            ['Jane Smith', 'jane.smith@company.com', '+1-234-567-8902', 'EMP002', '1',
             '456 Oak Ave, City, State', '34.0522', '-118.2437', 'Female', 'Secure@456'],
            ['Bob Johnson', 'bob.johnson@company.com', '+1-234-567-8903', 'EMP003', '2',
             '789 Pine Rd, City, State', '40.7128', '-74.0060', 'Other', 'Strong#789'],
        ]
        
        for row_num, row_data in enumerate(sample_data, 3):
            for col_num, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_num, column=col_num)
                cell.value = value
                cell.alignment = Alignment(horizontal="left", vertical="center")
        
        # Adjust column widths
        column_widths = [20, 30, 20, 15, 10, 30, 12, 12, 15]
        for col_num, width in enumerate(column_widths, 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(col_num)].width = width
        
        # Save to BytesIO
        excel_buffer = BytesIO()
        wb.save(excel_buffer)
        excel_buffer.seek(0)
        
        # Return as downloadable file
        filename = f"employee_bulk_upload_template_{datetime.now().strftime('%Y%m%d')}.xlsx"
        
        return StreamingResponse(
            excel_buffer,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={
                "Content-Disposition": f"attachment; filename={filename}"
            }
        )
        
    except Exception as e:
        logger.error(f"Error generating template: {str(e)}", exc_info=True)
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Failed to generate template: {str(e)}"
        )


@router.post("/bulk-upload", status_code=status.HTTP_200_OK)
async def bulk_create_employees(
    file: UploadFile = File(...),
    background_tasks: BackgroundTasks = BackgroundTasks(),
    request: Request = None,
    db: Session = Depends(get_db),
    user_data=Depends(PermissionChecker(["employee.create"], check_tenant=True)),
):
    """
    Bulk create employees from Excel file.
    
    Expected Excel columns:
    - name (required)
    - email (required)
    - phone (required)
    - employee_code (optional)
    - team_id (required)
    - address (optional)
    - latitude (optional)
    - longitude (optional)
    - gender (optional: Male/Female/Other)
    - password (optional, default: 'Welcome@123')
    
    Rules:
    - File must be .xlsx or .xls format
    - Maximum 500 rows per upload
    - Validates all data before creating any employee
    - Returns detailed success/failure report
    """
    try:
        user_type = user_data.get("user_type")
        tenant_id = None

        # 🚫 Vendors/Drivers cannot create employees
        if user_type in {"vendor", "driver"}:
            raise HTTPException(
                status_code=status.HTTP_403_FORBIDDEN,
                detail=ResponseWrapper.error(
                    message="You don't have permission to create employees",
                    error_code="FORBIDDEN",
                ),
            )

        # 🔒 Tenant enforcement
        if user_type == "employee":
            tenant_id = user_data.get("tenant_id")
            if not tenant_id:
                raise HTTPException(
                    status_code=status.HTTP_403_FORBIDDEN,
                    detail=ResponseWrapper.error(
                        message="Tenant ID missing in token for employee",
                        error_code="TENANT_ID_REQUIRED",
                    ),
                )
        elif user_type == "admin":
            # Admin must provide tenant_id in first data row or we'll validate later
            pass

        # Validate file type
        if not file.filename.endswith(('.xlsx', '.xls')):
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail=ResponseWrapper.error(
                    message="Invalid file format. Only .xlsx or .xls files are allowed",
                    error_code="INVALID_FILE_FORMAT",
                ),
            )

        # Read file content
        try:
            content = await file.read()
            
            # Check if file is empty
            if len(content) == 0:
                raise HTTPException(
                    status_code=status.HTTP_400_BAD_REQUEST,
                    detail=ResponseWrapper.error(
                        message="Uploaded file is empty",
                        error_code="EMPTY_FILE",
                    ),
                )
            
            # Log file details for debugging
            logger.info(f"Reading Excel file: {file.filename}, size: {len(content)} bytes")
            
            # Try to load the workbook
            workbook = openpyxl.load_workbook(BytesIO(content))
            sheet = workbook.active
            
        except HTTPException:
            raise
        except Exception as e:
            error_msg = str(e)
            logger.error(f"Failed to read Excel file '{file.filename}': {error_msg}")
            
            # Provide more specific error messages
            if "is not a zip file" in error_msg.lower() or "content_types" in error_msg.lower():
                message = "The uploaded file is not a valid Excel file. Please ensure you're uploading a .xlsx file created by Excel, Google Sheets, or similar spreadsheet software."
            elif "password" in error_msg.lower() or "encrypted" in error_msg.lower():
                message = "The Excel file is password-protected or encrypted. Please upload an unprotected file."
            elif "corrupted" in error_msg.lower() or "damaged" in error_msg.lower():
                message = "The Excel file appears to be corrupted. Please try re-downloading the template and filling it again."
            else:
                message = f"Failed to read Excel file: {error_msg}"
            
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail=ResponseWrapper.error(
                    message=message,
                    error_code="FILE_READ_ERROR",
                ),
            )

        # Parse header row
        headers = []
        for cell in sheet[1]:
            if cell.value:
                headers.append(str(cell.value).strip().lower())
            else:
                headers.append(None)

        # Validate required columns
        required_columns = ['name', 'email', 'phone', 'team_id']
        missing_columns = [col for col in required_columns if col not in headers]
        if missing_columns:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail=ResponseWrapper.error(
                    message=f"Missing required columns: {', '.join(missing_columns)}",
                    error_code="MISSING_COLUMNS",
                ),
            )

        # Parse data rows
        employees_data = []
        errors = []
        row_number = 2  # Excel row number (1-indexed, header is row 1, instructions row 2)
        
        # Start from row 3 to skip header (row 1) and instructions (row 2)
        for row in sheet.iter_rows(min_row=3, values_only=False):
            row_number += 1
            
            # Skip empty rows
            if all(cell.value is None or str(cell.value).strip() == '' for cell in row):
                logger.debug(f"Skipping empty row {row_number}")
                continue
            
            # Build employee data dict
            employee_dict = {}
            row_errors = []
            
            for idx, header in enumerate(headers):
                if header and idx < len(row):
                    cell_value = row[idx].value
                    if cell_value is not None:
                        employee_dict[header] = str(cell_value).strip()
            
            # Log the parsed row data for debugging
            logger.debug(f"Row {row_number} parsed data: {employee_dict}")
            
            # Validate this row
            row_validation = validate_employee_row(
                employee_dict, 
                row_number, 
                db, 
                tenant_id,
                user_type
            )
            
            if row_validation['valid']:
                logger.debug(f"Row {row_number} validation passed")
                employees_data.append({
                    'row': row_number,
                    'data': row_validation['data']
                })
            else:
                logger.warning(f"Row {row_number} validation failed: {row_validation['errors']}")
                errors.append({
                    'row': row_number,
                    'errors': row_validation['errors']
                })
        
        # Check if we have any data
        if not employees_data and not errors:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail=ResponseWrapper.error(
                    message="No valid data found in Excel file",
                    error_code="NO_DATA",
                ),
            )
        
        # Check row limit
        if len(employees_data) > 500:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail=ResponseWrapper.error(
                    message=f"Too many rows. Maximum 500 employees per upload (found {len(employees_data)})",
                    error_code="TOO_MANY_ROWS",
                ),
            )

        # If there are validation errors, return them without creating anything
        if errors:
            return ResponseWrapper.error(
                message=f"Validation failed for {len(errors)} row(s). No employees were created.",
                error_code="VALIDATION_FAILED",
                details={
                    "total_rows": len(employees_data) + len(errors),
                    "valid_rows": len(employees_data),
                    "invalid_rows": len(errors),
                    "errors": errors
                }
            )

        # Check for duplicates within the Excel file itself
        seen_emails = set()
        seen_phones = set()
        seen_codes = set()
        duplicate_errors = []
        
        for item in employees_data:
            row_num = item['row']
            emp_data = item['data']
            email = emp_data.get('email')
            phone = emp_data.get('phone')
            code = emp_data.get('employee_code')
            
            if email and email in seen_emails:
                duplicate_errors.append({
                    'row': row_num,
                    'errors': [f"Duplicate email '{email}' found in Excel file (already appears in another row)"]
                })
            elif email:
                seen_emails.add(email)
            
            if phone and phone in seen_phones:
                duplicate_errors.append({
                    'row': row_num,
                    'errors': [f"Duplicate phone '{phone}' found in Excel file (already appears in another row)"]
                })
            elif phone:
                seen_phones.add(phone)
            
            if code and code in seen_codes:
                duplicate_errors.append({
                    'row': row_num,
                    'errors': [f"Duplicate employee code '{code}' found in Excel file (already appears in another row)"]
                })
            elif code:
                seen_codes.add(code)
        
        if duplicate_errors:
            return ResponseWrapper.error(
                message=f"Duplicate entries found in Excel file. No employees were created.",
                error_code="DUPLICATE_IN_FILE",
                details={
                    "total_rows": len(employees_data),
                    "duplicate_rows": len(duplicate_errors),
                    "errors": duplicate_errors
                }
            )

        # All validation passed - now create employees
        logger.info(f"Starting bulk employee creation: {len(employees_data)} employees to process")
        created_employees = []
        failed_employees = []
        
        for item in employees_data:
            row_num = item['row']
            emp_data = item['data']
            
            try:
                # Determine tenant_id for this employee
                if user_type == "employee":
                    emp_tenant_id = tenant_id
                else:  # admin
                    emp_tenant_id = emp_data.get('tenant_id', tenant_id)
                
                # Double-check for duplicates before creating (race condition protection)
                email = emp_data['email']
                phone = emp_data['phone']
                employee_code = emp_data.get('employee_code')
                
                # Check if email already exists (including in this transaction)
                existing_email = db.query(Employee).filter(Employee.email == email).first()
                if existing_email:
                    raise ValueError(f"Email '{email}' already exists in database")
                
                # Check if phone already exists
                existing_phone = db.query(Employee).filter(Employee.phone == phone).first()
                if existing_phone:
                    raise ValueError(f"Phone '{phone}' already exists in database")
                
                # Check if employee code already exists (if provided)
                if employee_code:
                    existing_code = db.query(Employee).filter(Employee.employee_code == employee_code).first()
                    if existing_code:
                        raise ValueError(f"Employee code '{employee_code}' already exists in database")
                
                # Create employee object
                employee_create = EmployeeCreate(
                    name=emp_data['name'],
                    email=emp_data['email'],
                    phone=emp_data['phone'],
                    employee_code=emp_data.get('employee_code'),
                    team_id=emp_data['team_id'],
                    address=emp_data.get('address'),
                    latitude=emp_data.get('latitude'),
                    longitude=emp_data.get('longitude'),
                    gender=emp_data.get('gender'),
                    password=emp_data.get('password', 'Welcome@123'),
                    tenant_id=emp_tenant_id
                )
                
                # Create in database
                db_employee = employee_crud.create_with_tenant(
                    db=db, 
                    obj_in=employee_create, 
                    tenant_id=emp_tenant_id
                )
                
                db.flush()  # Get the ID without committing yet
                
                # Log successful creation
                logger.info(
                    f"✓ Row {row_num}: Employee created successfully - "
                    f"ID: {db_employee.employee_id}, Name: {db_employee.name}, Email: {db_employee.email}"
                )
                
                # Add to background email tasks
                background_tasks.add_task(
                    send_employee_created_email,
                    employee_data={
                        "name": db_employee.name,
                        "email": db_employee.email,
                        "phone": db_employee.phone,
                        "employee_id": db_employee.employee_id,
                        "tenant_id": emp_tenant_id,
                        "team_id": db_employee.team_id,
                    },
                )
                
                created_employees.append({
                    'row': row_num,
                    'employee_id': db_employee.employee_id,
                    'name': db_employee.name,
                    'email': db_employee.email
                })
                
            except IntegrityError as ie:
                db.rollback()
                error_msg = str(ie.orig) if hasattr(ie, 'orig') else str(ie)
                if 'unique constraint' in error_msg.lower() or 'duplicate' in error_msg.lower():
                    if 'email' in error_msg.lower():
                        error_msg = f"Email '{emp_data['email']}' already exists"
                    elif 'phone' in error_msg.lower():
                        error_msg = f"Phone '{emp_data['phone']}' already exists"
                    elif 'employee_code' in error_msg.lower():
                        error_msg = f"Employee code '{emp_data.get('employee_code')}' already exists"
                    else:
                        error_msg = "Duplicate entry found"
                
                failed_employees.append({
                    'row': row_num,
                    'name': emp_data.get('name', 'Unknown'),
                    'email': emp_data.get('email', 'Unknown'),
                    'error': error_msg
                })
                logger.error(
                    f"✗ Row {row_num}: FAILED - {emp_data.get('name', 'Unknown')} ({emp_data.get('email', 'Unknown')}) - "
                    f"Error: {error_msg}"
                )
                
            except Exception as e:
                db.rollback()
                failed_employees.append({
                    'row': row_num,
                    'name': emp_data.get('name', 'Unknown'),
                    'email': emp_data.get('email', 'Unknown'),
                    'error': str(e)
                })
                logger.error(
                    f"✗ Row {row_num}: FAILED - {emp_data.get('name', 'Unknown')} ({emp_data.get('email', 'Unknown')}) - "
                    f"Unexpected error: {str(e)}"
                )
        
        # Commit all successful creations
        if created_employees:
            try:
                db.commit()
                logger.info(
                    f"✅ Database transaction committed successfully: "
                    f"{len(created_employees)} employees stored in database"
                )
                
                # Log each created employee
                logger.info("=" * 80)
                logger.info("BULK EMPLOYEE CREATION SUMMARY - SUCCESSFULLY CREATED:")
                logger.info("=" * 80)
                for emp in created_employees:
                    logger.info(
                        f"  • Row {emp['row']}: {emp['name']} ({emp['email']}) - Employee ID: {emp['employee_id']}"
                    )
                logger.info("=" * 80)
                
                # Create audit log for bulk creation
                log_audit(
                    db=db,
                    tenant_id=tenant_id or emp_tenant_id,
                    module="employee",
                    action="BULK_CREATE",
                    user_data=user_data,
                    description=f"Bulk created {len(created_employees)} employees from Excel file",
                    new_values={
                        "total_created": len(created_employees),
                        "total_failed": len(failed_employees),
                        "file_name": file.filename,
                        "created_employee_ids": [emp['employee_id'] for emp in created_employees]
                    },
                    request=request
                )
                
            except Exception as commit_error:
                db.rollback()
                logger.error("=" * 80)
                logger.error("❌ BULK EMPLOYEE CREATION FAILED - TRANSACTION ROLLED BACK")
                logger.error("=" * 80)
                logger.error(f"Commit error: {str(commit_error)}")
                logger.error(f"NO EMPLOYEES WERE CREATED - All {len(created_employees)} employees rolled back")
                logger.error("=" * 80)
                raise HTTPException(
                    status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
                    detail=ResponseWrapper.error(
                        message=f"Failed to save employees to database. No employees were created. Error: {str(commit_error)}",
                        error_code="BULK_CREATION_COMMIT_FAILED",
                    ),
                )
        else:
            logger.warning("No employees were created - all rows had errors")
        
        # Log failures if any
        if failed_employees:
            logger.warning("=" * 80)
            logger.warning(f"BULK EMPLOYEE CREATION - FAILED ROWS ({len(failed_employees)}):")
            logger.warning("=" * 80)
            for emp in failed_employees:
                logger.warning(
                    f"  • Row {emp['row']}: {emp['name']} ({emp['email']}) - Error: {emp['error']}"
                )
            logger.warning("=" * 80)
        
        # Final summary log
        logger.info(
            f"BULK EMPLOYEE CREATION COMPLETED: "
            f"Total: {len(employees_data)}, Success: {len(created_employees)}, Failed: {len(failed_employees)}"
        )
        
        # Prepare response
        response_data = {
            "total_rows_processed": len(employees_data),
            "successful": len(created_employees),
            "failed": len(failed_employees),
            "created_employees": created_employees,
            "failed_employees": failed_employees
        }
        
        if failed_employees:
            message = f"Partially completed: {len(created_employees)} employees created, {len(failed_employees)} failed"
        else:
            message = f"Successfully created {len(created_employees)} employees"
        
        return ResponseWrapper.success(
            data=response_data,
            message=message
        )

    except HTTPException:
        raise
    except Exception as e:
        logger.exception(f"Unexpected error during bulk employee creation: {str(e)}")
        db.rollback()
        raise handle_http_error(e)


def validate_employee_row(
    row_data: Dict[str, Any], 
    row_number: int, 
    db: Session,
    token_tenant_id: Optional[str],
    user_type: str
) -> Dict[str, Any]:
    """
    Validate a single employee row from Excel.
    
    Returns:
        {
            'valid': bool,
            'data': dict (if valid),
            'errors': list (if invalid)
        }
    """
    errors = []
    validated_data = {}
    
    # 1. Validate required fields
    if not row_data.get('name') or not row_data['name'].strip():
        errors.append("Name is required")
    else:
        validated_data['name'] = row_data['name'].strip()
    
    # 2. Validate email
    email = row_data.get('email', '').strip()
    if not email:
        errors.append("Email is required")
    else:
        email_pattern = r'^[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}$'
        if not re.match(email_pattern, email):
            errors.append(f"Invalid email format: {email}")
        else:
            # Check if email already exists
            existing = db.query(Employee).filter(Employee.email == email).first()
            if existing:
                errors.append(f"Email '{email}' already exists in database")
            else:
                validated_data['email'] = email
    
    # 3. Validate phone
    phone = row_data.get('phone', '').strip()
    if not phone:
        errors.append("Phone is required")
    else:
        # Remove common separators
        phone_clean = re.sub(r'[\s\-\(\)\+]', '', phone)
        if not phone_clean.isdigit() or len(phone_clean) < 10:
            errors.append(f"Invalid phone format: {phone}")
        else:
            # Check if phone already exists
            existing = db.query(Employee).filter(Employee.phone == phone).first()
            if existing:
                errors.append(f"Phone '{phone}' already exists in database")
            else:
                validated_data['phone'] = phone
    
    # 4. Validate team_id
    team_id_str = row_data.get('team_id', '').strip()
    if not team_id_str:
        errors.append("Team ID is required")
    else:
        try:
            team_id = int(team_id_str)
            team = team_crud.get_by_id(db, team_id=team_id)
            if not team:
                errors.append(f"Team ID {team_id} does not exist")
            else:
                # Validate team belongs to tenant
                if token_tenant_id and team.tenant_id != token_tenant_id:
                    errors.append(f"Team ID {team_id} does not belong to your tenant")
                else:
                    validated_data['team_id'] = team_id
                    # Store tenant_id for admin users
                    if user_type == "admin" and not token_tenant_id:
                        validated_data['tenant_id'] = team.tenant_id
        except ValueError:
            errors.append(f"Team ID must be a number, got: {team_id_str}")
    
    # 5. Validate employee_code (optional but must be unique if provided)
    employee_code = row_data.get('employee_code', '').strip()
    if employee_code:
        existing = db.query(Employee).filter(Employee.employee_code == employee_code).first()
        if existing:
            errors.append(f"Employee code '{employee_code}' already exists")
        else:
            validated_data['employee_code'] = employee_code
    
    # 6. Validate optional fields
    if 'address' in row_data and row_data['address']:
        validated_data['address'] = row_data['address'].strip()
    
    # 7. Validate latitude/longitude
    if 'latitude' in row_data and row_data['latitude']:
        try:
            lat = float(row_data['latitude'])
            if -90 <= lat <= 90:
                validated_data['latitude'] = lat
            else:
                errors.append(f"Latitude must be between -90 and 90, got: {lat}")
        except ValueError:
            errors.append(f"Invalid latitude format: {row_data['latitude']}")
    
    if 'longitude' in row_data and row_data['longitude']:
        try:
            lon = float(row_data['longitude'])
            if -180 <= lon <= 180:
                validated_data['longitude'] = lon
            else:
                errors.append(f"Longitude must be between -180 and 180, got: {lon}")
        except ValueError:
            errors.append(f"Invalid longitude format: {row_data['longitude']}")
    
    # 8. Validate gender (optional)
    gender = row_data.get('gender', '').strip()
    if gender:
        valid_genders = ['Male', 'Female', 'Other', 'male', 'female', 'other']
        if gender in valid_genders:
            # Capitalize first letter to match enum
            validated_data['gender'] = gender.capitalize()
        else:
            errors.append(f"Invalid gender. Must be one of: Male, Female, Other. Got: {gender}")
    
    # 9. Validate password (optional)
    password = row_data.get('password', '').strip()
    if password:
        if len(password) < 6:
            errors.append("Password must be at least 6 characters long")
        else:
            validated_data['password'] = password
    else:
        validated_data['password'] = 'Welcome@123'  # Default password
    
    return {
        'valid': len(errors) == 0,
        'data': validated_data if len(errors) == 0 else None,
        'errors': errors
    }


@router.get("/", status_code=status.HTTP_200_OK)
def read_employees(
    skip: int = 0,
    limit: int = 100,
    name: Optional[str] = None,
    tenant_id: Optional[str] = None,
    team_id: Optional[int] = None,
    is_active: Optional[bool] = None,  # 👈 Added filter
    db: Session = Depends(get_db),
    user_data=Depends(PermissionChecker(["employee.read"], check_tenant=True)),
):
    """
    Fetch employees with role-based restrictions:
    - driver/vendor → forbidden
    - employee → only within their tenant
    - admin → must filter by tenant_id
    - team_id → optional filter, must belong to the same tenant
    - is_active → optional filter
    """
    try:
        user_type = user_data.get("user_type")

        # 🚫 Vendors/Drivers forbidden
        if user_type in {"vendor", "driver"}:
            raise HTTPException(
                status_code=status.HTTP_403_FORBIDDEN,
                detail=ResponseWrapper.error(
                    message="You don't have permission to view employees",
                    error_code="FORBIDDEN",
                ),
            )

        # Tenant enforcement
        if user_type == "employee":
            tenant_id = user_data.get("tenant_id")
            if not tenant_id:
                raise HTTPException(
                    status_code=status.HTTP_403_FORBIDDEN,
                    detail=ResponseWrapper.error(
                        message="Tenant ID missing in token for employee",
                        error_code="TENANT_ID_REQUIRED",
                    ),
                )

        elif user_type == "admin":
            if not tenant_id:
                raise HTTPException(
                    status_code=status.HTTP_400_BAD_REQUEST,
                    detail=ResponseWrapper.error(
                        message="Tenant ID is required for admin",
                        error_code="TENANT_ID_REQUIRED",
                    ),
                )

            # Ensure tenant exists
            if not tenant_crud.get_by_id(db, tenant_id=tenant_id):
                raise HTTPException(
                    status_code=status.HTTP_404_NOT_FOUND,
                    detail=ResponseWrapper.error(
                        message=f"Tenant {tenant_id} not found",
                        error_code="TENANT_NOT_FOUND",
                    ),
                )

        # --- Team filter check ---
        if team_id is not None:
            team = team_crud.get_by_id(db, team_id=team_id)
            if not team or team.tenant_id != tenant_id:  # 🔒 enforce tenant match
                raise HTTPException(
                    status_code=status.HTTP_403_FORBIDDEN,
                    detail=ResponseWrapper.error(
                        message=f"Team {team_id} does not belong to tenant {tenant_id}",
                        error_code="TEAM_NOT_IN_TENANT",
                    ),
                )

        # Query employees
        query = db.query(Employee).filter(Employee.tenant_id == tenant_id)

        if name:
            query = query.filter(Employee.name.ilike(f"%{name}%"))
        if team_id is not None:
            query = query.filter(Employee.team_id == team_id)
        if is_active is not None:  # 👈 Apply is_active filter
            query = query.filter(Employee.is_active == is_active)

        total, items = paginate_query(query, skip, limit)

        employees = []
        for emp in items:
            emp_dict = EmployeeResponse.model_validate(emp, from_attributes=True).model_dump()
            # Add tenant location details
            if emp.tenant:
                emp_dict["tenant_latitude"] = float(emp.tenant.latitude) if emp.tenant.latitude else None
                emp_dict["tenant_longitude"] = float(emp.tenant.longitude) if emp.tenant.longitude else None
                emp_dict["tenant_address"] = emp.tenant.address
            employees.append(emp_dict)

        return ResponseWrapper.success(
            data={"total": total, "items": employees},
            message="Employees fetched successfully"
        )

    except SQLAlchemyError as e:
        raise handle_db_error(e)
    except HTTPException:
        raise
    except Exception as e:
        logger.exception(f"Unexpected error while fetching employees: {str(e)}")
        raise handle_http_error(e)

@router.get("/{employee_id}", status_code=status.HTTP_200_OK)
def read_employee(
    employee_id: int,
    db: Session = Depends(get_db),
    user_data=Depends(PermissionChecker(["employee.read"], check_tenant=True)),
):
    """
    Fetch an employee by ID with role-based restrictions:
    - driver/vendor → forbidden
    - employee → only within their tenant
    - admin → unrestricted (but tenant_id should still match employee)
    """
    try:
        user_type = user_data.get("user_type")
        tenant_id = user_data.get("tenant_id")

        # 🚫 Vendors/Drivers forbidden
        if user_type in {"vendor", "driver"}:
            raise HTTPException(
                status_code=status.HTTP_403_FORBIDDEN,
                detail=ResponseWrapper.error(
                    message="You don't have permission to view employees",
                    error_code="FORBIDDEN",
                ),
            )

        # 🔒 Tenant enforcement
        query = db.query(Employee).filter(Employee.employee_id == employee_id)
        if user_type == "employee":
            query = query.filter(Employee.tenant_id == tenant_id)

        db_employee = query.first()
        if not db_employee:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail=ResponseWrapper.error(
                    message=f"Employee with ID {employee_id} not found",
                    error_code="EMPLOYEE_NOT_FOUND",
                ),
            )

        employee_data = EmployeeResponse.model_validate(db_employee, from_attributes=True).model_dump()
        # Add tenant location details
        if db_employee.tenant:
            employee_data["tenant_latitude"] = float(db_employee.tenant.latitude) if db_employee.tenant.latitude else None
            employee_data["tenant_longitude"] = float(db_employee.tenant.longitude) if db_employee.tenant.longitude else None
            employee_data["tenant_address"] = db_employee.tenant.address

        return ResponseWrapper.success(
            data={"employee": employee_data}, message="Employee fetched successfully"
        )

    except SQLAlchemyError as e:
        raise handle_db_error(e)
    except HTTPException:
        raise
    except Exception as e:
        logger.exception(f"Unexpected error while fetching employee {employee_id}: {str(e)}")
        raise handle_http_error(e)

@router.put("/{employee_id}", status_code=status.HTTP_200_OK)
def update_employee(
    employee_id: int,
    employee_update: EmployeeUpdate,
    request: Request,
    db: Session = Depends(get_db),
    user_data=Depends(PermissionChecker(["employee.update"], check_tenant=True)),
):
    """
    Update an employee with role-based restrictions:
    - driver/vendor → forbidden
    - employee → only within their tenant
    - admin → must provide valid tenant_id (employee must belong to that tenant)
    """
    try:
        user_type = user_data.get("user_type")
        token_tenant_id = user_data.get("tenant_id")

        # 🚫 Vendors/Drivers forbidden
        if user_type in {"vendor", "driver"}:
            raise HTTPException(
                status_code=status.HTTP_403_FORBIDDEN,
                detail=ResponseWrapper.error(
                    message="You don't have permission to update employees",
                    error_code="FORBIDDEN",
                ),
            )

        # Fetch employee
        db_employee = db.query(Employee).filter(Employee.employee_id == employee_id).first()
        if not db_employee:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail=ResponseWrapper.error(
                    message=f"Employee with ID {employee_id} not found",
                    error_code="EMPLOYEE_NOT_FOUND",
                ),
            )

        # 🔍 Capture old values before update
        old_values = {}
        update_data = employee_update.model_dump(exclude_unset=True)
        for key in update_data.keys():
            if key != "password":  # Don't log password
                old_val = getattr(db_employee, key, None)
                if old_val is not None:
                    old_values[key] = str(old_val) if not isinstance(old_val, (str, int, float, bool)) else old_val

        # 🔒 Tenant enforcement
        if user_type == "employee":
            if db_employee.tenant_id != token_tenant_id:
                raise HTTPException(
                    status_code=status.HTTP_403_FORBIDDEN,
                    detail=ResponseWrapper.error(
                        message="You cannot update employees outside your tenant",
                        error_code="TENANT_FORBIDDEN",
                    ),
                )
        elif user_type == "admin":
            if not db_employee.tenant_id:
                raise HTTPException(
                    status_code=status.HTTP_400_BAD_REQUEST,
                    detail=ResponseWrapper.error(
                        message="Employee tenant is missing in DB",
                        error_code="TENANT_MISSING",
                    ),
                )
            tenant = tenant_crud.get_by_id(db, tenant_id=db_employee.tenant_id)
            if not tenant:
                raise HTTPException(
                    status_code=status.HTTP_404_NOT_FOUND,
                    detail=ResponseWrapper.error(
                        message=f"Tenant {db_employee.tenant_id} not found",
                        error_code="TENANT_NOT_FOUND",
                    ),
                )

        # Apply updates (already captured above)
        if "password" in update_data:
            update_data["password"] = hash_password(update_data["password"])

        # 🚦 Team validation if updating team_id
        if "team_id" in update_data and update_data["team_id"] is not None:
            if not team_crud.is_team_in_tenant(
                db, team_id=update_data["team_id"], tenant_id=db_employee.tenant_id
            ):
                raise HTTPException(
                    status_code=status.HTTP_400_BAD_REQUEST,
                    detail=ResponseWrapper.error(
                        message=f"Team {update_data['team_id']} does not belong to tenant {db_employee.tenant_id}",
                        error_code="TEAM_TENANT_MISMATCH",
                    ),
                )

        # 🔐 Role validation if updating role_id
        if "role_id" in update_data and update_data["role_id"] is not None:
            from app.crud.iam.role import role_crud
            role = role_crud.get(db, id=update_data["role_id"])
            if not role:
                raise HTTPException(
                    status_code=status.HTTP_404_NOT_FOUND,
                    detail=ResponseWrapper.error(
                        message=f"Role with ID {update_data['role_id']} not found",
                        error_code="ROLE_NOT_FOUND",
                    ),
                )
            # Validate role belongs to the same tenant (or is a system role)
            if role.tenant_id and role.tenant_id != db_employee.tenant_id:
                raise HTTPException(
                    status_code=status.HTTP_400_BAD_REQUEST,
                    detail=ResponseWrapper.error(
                        message=f"Role {update_data['role_id']} does not belong to tenant {db_employee.tenant_id}",
                        error_code="ROLE_TENANT_MISMATCH",
                    ),
                )

        for key, value in update_data.items():
            setattr(db_employee, key, value)

        db.commit()
        db.refresh(db_employee)

        # 🔍 Capture new values after update
        new_values = {}
        for key in update_data.keys():
            if key != "password":  # Don't log password
                new_val = getattr(db_employee, key, None)
                if new_val is not None:
                    new_values[key] = str(new_val) if not isinstance(new_val, (str, int, float, bool)) else new_val

        # 🔍 Audit Log: Employee Update
        try:
            # Build description with changed fields
            changed_fields = list(update_data.keys())
            fields_str = ", ".join(changed_fields) if changed_fields else "details"
            
            log_audit(
                db=db,
                tenant_id=db_employee.tenant_id,
                module="employee",
                action="UPDATE",
                user_data=user_data,
                description=f"Updated employee '{db_employee.name}' - changed fields: {fields_str}",
                new_values={"employee_id": db_employee.employee_id, "old": old_values, "new": new_values},
                request=request
            )
            stored_fields = ["employee_id"] + list(old_values.keys()) + list(new_values.keys())
            logger.info(f"Audit log created for employee update: employee_id={db_employee.employee_id} stored in audit_data with fields: {stored_fields}")
        except Exception as audit_error:
            logger.error(f"Failed to create audit log for employee update: {str(audit_error)}", exc_info=True)

        logger.info(
            f"Employee updated successfully: employee_id={employee_id}, tenant_id={db_employee.tenant_id}"
        )

        employee_response = EmployeeResponse.model_validate(db_employee, from_attributes=True).model_dump()
        # Add tenant location details
        if db_employee.tenant:
            employee_response["tenant_latitude"] = float(db_employee.tenant.latitude) if db_employee.tenant.latitude else None
            employee_response["tenant_longitude"] = float(db_employee.tenant.longitude) if db_employee.tenant.longitude else None
            employee_response["tenant_address"] = db_employee.tenant.address

        return ResponseWrapper.success(
            data={"employee": employee_response},
            message="Employee updated successfully",
        )

    except SQLAlchemyError as e:
        raise handle_db_error(e)
    except HTTPException:
        raise
    except Exception as e:
        logger.exception(f"Unexpected error while updating employee {employee_id}: {str(e)}")
        raise handle_http_error(e)

@router.patch("/{employee_id}/toggle-status", status_code=status.HTTP_200_OK)
def toggle_employee_status(
    employee_id: int,
    request: Request,
    db: Session = Depends(get_db),
    user_data=Depends(PermissionChecker(["employee.update"], check_tenant=True)),
):
    """
    Toggle employee active/inactive status.
    - driver/vendor → forbidden
    - employee → only within their tenant
    - admin → must belong to tenant
    """
    try:
        user_type = user_data.get("user_type")
        token_tenant_id = user_data.get("tenant_id")

        # 🚫 Vendors/Drivers forbidden
        if user_type in {"vendor", "driver"}:
            raise HTTPException(
                status_code=status.HTTP_403_FORBIDDEN,
                detail=ResponseWrapper.error(
                    message="You don't have permission to modify employees",
                    error_code="FORBIDDEN",
                ),
            )

        # Fetch employee
        db_employee = db.query(Employee).filter(Employee.employee_id == employee_id).first()
        if not db_employee:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail=ResponseWrapper.error(
                    message=f"Employee with ID {employee_id} not found",
                    error_code="EMPLOYEE_NOT_FOUND",
                ),
            )

        # 🔍 Capture old status for audit
        old_status = db_employee.is_active

        # 🔒 Tenant enforcement
        if user_type == "employee":
            if db_employee.tenant_id != token_tenant_id:
                raise HTTPException(
                    status_code=status.HTTP_403_FORBIDDEN,
                    detail=ResponseWrapper.error(
                        message="You cannot modify employees outside your tenant",
                        error_code="TENANT_FORBIDDEN",
                    ),
                )
        elif user_type == "admin":
            tenant = tenant_crud.get_by_id(db, tenant_id=db_employee.tenant_id)
            if not tenant:
                raise HTTPException(
                    status_code=status.HTTP_404_NOT_FOUND,
                    detail=ResponseWrapper.error(
                        message=f"Tenant {db_employee.tenant_id} not found",
                        error_code="TENANT_NOT_FOUND",
                    ),
                )

        # 🚦 Toggle status
        old_status = db_employee.is_active
        db_employee.is_active = not db_employee.is_active
        db.commit()
        db.refresh(db_employee)

        # 🔍 Audit Log: Status Toggle
        try:
            status_text = 'active' if db_employee.is_active else 'inactive'
            log_audit(
                db=db,
                tenant_id=db_employee.tenant_id,
                module="employee",
                action="UPDATE",
                user_data=user_data,
                description=f"Toggled employee '{db_employee.name}' status to {status_text}",
                new_values={"employee_id": db_employee.employee_id, "old_status": old_status, "new_status": db_employee.is_active},
                request=request
            )
            stored_fields = ["employee_id", "old_status", "new_status"]
            logger.info(f"Audit log created for employee status toggle: employee_id={db_employee.employee_id} stored in audit_data with fields: {stored_fields}")
        except Exception as audit_error:
            logger.error(f"Failed to create audit log for status toggle: {str(audit_error)}")

        logger.info(
            f"Employee {employee_id} status toggled to "
            f"{'active' if db_employee.is_active else 'inactive'} "
            f"(tenant_id={db_employee.tenant_id})"
        )

        return ResponseWrapper.success(
            data={
                "employee_id": db_employee.employee_id,
                "is_active": db_employee.is_active,
            },
            message=f"Employee status updated to {'active' if db_employee.is_active else 'inactive'}",
        )

    except SQLAlchemyError as e:
        raise handle_db_error(e)
    except HTTPException:
        raise
    except Exception as e:
        logger.exception(f"Unexpected error while toggling employee {employee_id} status: {str(e)}")
        raise handle_http_error(e)

# @router.delete("/{employee_id}", status_code=status.HTTP_204_NO_CONTENT)
# def delete_employee(
#     employee_id: int, 
#     db: Session = Depends(get_db),
#     user_data=Depends(PermissionChecker(["employee.delete"], check_tenant=True))
# ):
#     db_employee = db.query(Employee).filter(Employee.employee_id == employee_id).first()
#     if not db_employee:
#         raise HTTPException(
#             status_code=status.HTTP_404_NOT_FOUND,
#             detail=f"Employee with ID {employee_id} not found"
#         )
    
#     db.delete(db_employee)
#     db.commit()
#     return None
