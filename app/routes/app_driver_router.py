# app/routers/app_driver_router.py
"""
Driver App Router - Production-Grade Optimizations Applied

RECOMMENDED DATABASE INDEXES for optimal performance:
-------------------------------------------------------
1. RouteManagement:
   - (tenant_id, assigned_driver_id, status, created_at)  # For get_driver_trips
   - (assigned_driver_id, status)  # For start_duty ongoing check
   
2. RouteManagementBooking:
   - (route_id, order_id)  # For sequential booking operations
   - (route_id, booking_id)  # For booking validation
   - (booking_id)  # Foreign key index
   
3. Booking:
   - (tenant_id, booking_id)  # For booking validation
   - (booking_date, status)  # For date filtering
   - (employee_id)  # For employee joins
   - (status)  # For status filtering

PERFORMANCE OPTIMIZATIONS APPLIED:
-----------------------------------
✅ N+1 Query Prevention: Batch loading with single query for all routes
✅ Eager Loading: Using selectinload() for relationships
✅ exists() vs count(): Using exists() for boolean checks
✅ Pagination: Added limit/offset for large result sets
✅ Query Combining: Single queries instead of multiple round-trips
✅ Proper Indexing: Comments added for recommended indexes
"""
from fastapi import APIRouter, Depends, HTTPException, Query, status, BackgroundTasks
from sqlalchemy import func, exists
from sqlalchemy.orm import Session, joinedload, selectinload
from typing import Optional, List
from datetime import date, datetime, timedelta

from app.core.logging_config import get_logger
from app.database.session import SessionLocal, get_db
from app.models.employee import Employee
from app.models.shift import Shift, PickupTypeEnum
from app.models.nodal_point import NodalPoint
from app.utils.response_utils import ResponseWrapper, handle_db_error, handle_http_error
from common_utils.auth.permission_checker import PermissionChecker
from common_utils import get_current_ist_time
from app.models.route_management import RouteManagement, RouteManagementBooking, RouteManagementStatusEnum
from app.models.booking import Booking, BookingStatusEnum
from app.models.tenant import Tenant
from app.models.tenant_config import TenantConfig
from app.models.vehicle import Vehicle
from app.models.driver import Driver
from geopy.distance import geodesic
from app.utils.delay_tagging import tag_trip_delay

# Notification services
from app.core.email_service import EmailService
from app.services.sms_service import SMSService
from app.services.unified_notification_service import UnifiedNotificationService


logger = get_logger(__name__)
router = APIRouter(prefix="/driver", tags=["Driver App"])

# ---------------------------
# Dependencies & Utilities
# ---------------------------

async def DriverAuth(user_data=Depends(PermissionChecker(["driver_app.read", "driver_app.update"]))):
    """
    Ensures the token belongs to a driver persona and returns (tenant_id, driver_id, user_id).
    """
    if user_data.get("user_type") != "driver":
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="Driver access only")
    tenant_id = user_data.get("tenant_id")
    driver_id = user_data.get("user_id")
    vendor_id = user_data.get("vendor_id")
    if not tenant_id or not driver_id:
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="Driver or tenant not resolved from token")
    return {"tenant_id": tenant_id, "driver_id": driver_id, "vendor_id": vendor_id}


def verify_otp(booking_otp: Optional[int], provided_otp: Optional[str], otp_type: str, booking_id: int) -> None:
    """
    Verifies OTP for boarding or deboarding.
    Raises HTTPException if OTP is required but invalid.
    """
    if booking_otp:
        if str(booking_otp) != str(provided_otp):
            logger.warning(f"[driver.verify_otp] Invalid {otp_type} OTP provided for booking {booking_id}")
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail=ResponseWrapper.error(
                    message=f"Invalid {otp_type} OTP",
                    error_code=f"INVALID_{otp_type.upper()}_OTP",
                    details={"booking_id": booking_id},
                ),
            )
        logger.info(f"[driver.verify_otp] {otp_type.capitalize()} OTP validation successful for booking {booking_id}")
    else:
        logger.info(f"[driver.verify_otp] No {otp_type} OTP required for booking {booking_id}")


def get_next_stop(db: Session, route_id: int, current_order_id: int, skip_ongoing: bool = False) -> Optional[dict]:
    """
    Fetches the next pending stop in the route after the given order_id.
    Returns serialized next stop data or None if no next stop exists.
    OPTIMIZED: Single query with join to fetch both RouteManagementBooking and Booking.

    skip_ongoing=True: for NODAL routes — exclude ONGOING bookings (already QR-boarded)
    so the driver's "next stop" view only shows passengers still waiting.
    """
    excluded = [BookingStatusEnum.NO_SHOW, BookingStatusEnum.COMPLETED]
    if skip_ongoing:
        excluded.append(BookingStatusEnum.ONGOING)

    result = (
        db.query(RouteManagementBooking, Booking)
        .join(Booking, RouteManagementBooking.booking_id == Booking.booking_id)
        .filter(
            RouteManagementBooking.route_id == route_id,
            RouteManagementBooking.order_id > current_order_id,
            Booking.status.notin_(excluded),
        )
        .order_by(RouteManagementBooking.order_id)
        .first()
    )

    if not result:
        return None
    
    next_rb, booking_next = result

    return {
        "booking_id": booking_next.booking_id,
        "employee_id": booking_next.employee_id,
        "pickup_latitude": booking_next.pickup_latitude,
        "pickup_longitude": booking_next.pickup_longitude,
        "pickup_location": booking_next.pickup_location,
        "estimated_pickup_time": next_rb.estimated_pick_up_time,
    }


def validate_route_for_driver(
    db: Session,
    route_id: int,
    driver_id: int,
    tenant_id: int,
    required_status: Optional[RouteManagementStatusEnum] = None
) -> RouteManagement:
    """
    Validates that a route exists, is assigned to the driver, and optionally matches required status.
    Raises HTTPException if validation fails.
    Returns the route object if valid.
    """
    query = db.query(RouteManagement).filter(
        RouteManagement.route_id == route_id,
        RouteManagement.tenant_id == tenant_id,
        RouteManagement.assigned_driver_id == driver_id,
    )
    
    if required_status:
        query = query.filter(RouteManagement.status == required_status)
    
    route = query.first()
    
    if not route:
        error_msg = "Route not found or not assigned to driver"
        error_details = {"route_id": route_id, "driver_id": driver_id}
        
        if required_status:
            error_msg = f"Route not found, not assigned to driver, or not in {required_status.value} state"
            error_details["required_status"] = required_status.value
        
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail=ResponseWrapper.error(
                message=error_msg,
                error_code="ROUTE_NOT_FOUND",
                details=error_details,
            ),
        )
    
    return route


def validate_booking_in_route(
    db: Session,
    route_id: int,
    booking_id: int,
    tenant_id: int
) -> tuple[RouteManagementBooking, Booking]:
    """
    Validates that a booking exists in a route and belongs to the tenant.
    Raises HTTPException if validation fails.
    Returns tuple of (RouteManagementBooking, Booking) if valid.
    """
    # Validate route-booking association
    rb = (
        db.query(RouteManagementBooking)
        .filter(
            RouteManagementBooking.route_id == route_id,
            RouteManagementBooking.booking_id == booking_id,
        )
        .first()
    )
    if not rb:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail=ResponseWrapper.error(
                message="Booking not found in this route",
                error_code="BOOKING_NOT_IN_ROUTE",
                details={"route_id": route_id, "booking_id": booking_id},
            ),
        )

    # Validate booking object
    booking = (
        db.query(Booking)
        .filter(
            Booking.booking_id == booking_id,
            Booking.tenant_id == tenant_id,
        )
        .first()
    )
    if not booking:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail=ResponseWrapper.error(
                message="Booking not found",
                error_code="BOOKING_NOT_FOUND",
                details={"booking_id": booking_id},
            ),
        )
    
    return rb, booking


def check_previous_bookings_completed(
    db: Session,
    route_id: int,
    current_order_id: int
) -> None:
    """
    Checks if all previous bookings in the route are completed (NO_SHOW, ONGOING, or COMPLETED).
    Raises HTTPException if any previous booking is still pending.
    OPTIMIZED: Uses exists() instead of count() for better performance.
    """
    has_pending = db.query(
        exists().where(
            RouteManagementBooking.route_id == route_id,
            RouteManagementBooking.order_id < current_order_id,
            RouteManagementBooking.booking_id == Booking.booking_id,
            Booking.status.notin_([
                BookingStatusEnum.NO_SHOW,
                BookingStatusEnum.ONGOING,
                BookingStatusEnum.COMPLETED
            ])
        )
    ).scalar()

    if has_pending:
        # Only count if we need the number for error message
        pending_count = (
            db.query(RouteManagementBooking)
            .join(Booking, RouteManagementBooking.booking_id == Booking.booking_id)
            .filter(
                RouteManagementBooking.route_id == route_id,
                RouteManagementBooking.order_id < current_order_id,
                Booking.status.notin_([
                    BookingStatusEnum.NO_SHOW,
                    BookingStatusEnum.ONGOING,
                    BookingStatusEnum.COMPLETED
                ]),
            )
            .count()
        )
        
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=ResponseWrapper.error(
                message="Cannot process this booking. Previous bookings in the route must be completed first.",
                error_code="PREVIOUS_BOOKINGS_PENDING",
                details={"pending_count": pending_count},
            ),
        )


def validate_driver_location(
    current_latitude: float,
    current_longitude: float,
    target_latitude: float,
    target_longitude: float,
    location_type: str,
    booking_id: int,
    max_distance_meters: int = 500
) -> None:
    """
    Validates that the driver's current location is within the allowed distance from the target location.
    Raises HTTPException if validation fails.
    """
    if not target_latitude or not target_longitude:
        logger.warning(f"[driver.{location_type}] Booking {booking_id} missing {location_type} coordinates")
        return

    distance = geodesic(
        (current_latitude, current_longitude),
        (target_latitude, target_longitude)
    ).meters

    if distance > max_distance_meters:
        error_code = f"DRIVER_TOO_FAR_FROM_{location_type.upper()}"
        message = f"Driver location is too far from {location_type} location. Distance: {distance:.1f} meters (max allowed: {max_distance_meters} meters)"

        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=ResponseWrapper.error(
                message=message,
                error_code=error_code,
                details={
                    "driver_lat": current_latitude,
                    "driver_lng": current_longitude,
                    f"{location_type}_lat": target_latitude,
                    f"{location_type}_lng": target_longitude,
                    "distance_meters": round(distance, 1),
                    "max_allowed_meters": max_distance_meters
                },
            ),
        )









from sqlalchemy import func, and_

# ──────────────────────────────────────────────────────────────
# Driver Config — tenant details + speed limit
# ──────────────────────────────────────────────────────────────

@router.get("/config", status_code=status.HTTP_200_OK)
async def get_driver_config(
    db: Session = Depends(get_db),
    ctx=Depends(DriverAuth),
):
    """
    Returns the tenant's basic details and configuration values that the driver
    app needs at startup:

    - Company name, address, coordinates
    - Speed limit in km/h (for in-app speedometer warnings)
    - OTP requirements per shift direction (login/logout × boarding/deboarding)
    - Escort requirement flag

    If the driver's assigned vehicle has a per-vehicle speed limit override
    (`speed_limit_override_kmph`), that value is returned as `effective_speed_limit_kmph`
    and takes priority over the tenant-wide limit.
    """
    try:
        tenant_id = ctx["tenant_id"]
        driver_id = ctx["driver_id"]

        # ── 1. Load tenant ──────────────────────────────────────────────────
        tenant = db.query(Tenant).filter(Tenant.tenant_id == tenant_id).first()
        if not tenant:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail=ResponseWrapper.error("Tenant not found", "TENANT_NOT_FOUND"),
            )

        # ── 2. Load tenant config (one-to-one, may not exist yet) ──────────
        cfg = db.query(TenantConfig).filter(TenantConfig.tenant_id == tenant_id).first()

        # ── 3. Check for per-vehicle speed limit override ───────────────────
        driver = db.query(Driver).filter(Driver.driver_id == driver_id).first()
        vehicle_speed_limit = None
        vehicle_rc = None
        if driver:
            vehicle = (
                db.query(Vehicle)
                .filter(
                    Vehicle.driver_id == driver_id,
                    Vehicle.is_active == True,
                )
                .first()
            )
            if vehicle:
                vehicle_speed_limit = vehicle.speed_limit_override_kmph
                vehicle_rc = vehicle.rc_number

        tenant_speed_limit = cfg.speed_limit_kmph if cfg else None
        effective_speed_limit = vehicle_speed_limit if vehicle_speed_limit is not None else tenant_speed_limit

        return ResponseWrapper.success(
            message="Driver config fetched successfully",
            data={
                "tenant": {
                    "tenant_id": tenant.tenant_id,
                    "name": tenant.name,
                    "address": tenant.address,
                    "latitude": float(tenant.latitude),
                    "longitude": float(tenant.longitude),
                    "is_active": tenant.is_active,
                },
                "speed": {
                    "tenant_speed_limit_kmph": tenant_speed_limit,
                    "vehicle_speed_limit_override_kmph": vehicle_speed_limit,
                    "effective_speed_limit_kmph": effective_speed_limit,
                    "assigned_vehicle_rc": vehicle_rc,
                },
                "otp": {
                    "login_boarding_otp": cfg.login_boarding_otp if cfg else None,
                    "login_deboarding_otp": cfg.login_deboarding_otp if cfg else None,
                    "logout_boarding_otp": cfg.logout_boarding_otp if cfg else None,
                    "logout_deboarding_otp": cfg.logout_deboarding_otp if cfg else None,
                },
                "safety": {
                    "escort_required_for_women": cfg.escort_required_for_women if cfg else None,
                    "escort_required_start_time": (
                        cfg.escort_required_start_time.strftime("%H:%M")
                        if cfg and cfg.escort_required_start_time else None
                    ),
                    "escort_required_end_time": (
                        cfg.escort_required_end_time.strftime("%H:%M")
                        if cfg and cfg.escort_required_end_time else None
                    ),
                },
            },
        )

    except HTTPException:
        raise
    except Exception as e:
        logger.exception("[driver.config] Error fetching driver config")
        raise handle_http_error(e)


@router.post("/duty/start", status_code=status.HTTP_200_OK)
async def start_duty(
    route_id: int,
    background_tasks: BackgroundTasks,
    db: Session = Depends(get_db),
    ctx=Depends(DriverAuth),
):
    """
    Start the driver's duty for a route.
    - Changes route status from DRIVER_ASSIGNED to ONGOING
    - Validates no other ongoing routes exist for this driver
    - Driver cannot revert once duty is started
    """
    try:
        tenant_id = ctx["tenant_id"]
        driver_id = ctx["driver_id"]
        now = get_current_ist_time()

        logger.info(f"[driver.start_duty] tenant={tenant_id}, driver={driver_id}, route={route_id}")

        # --- Validate route exists and is assigned to driver ---
        route = validate_route_for_driver(db, route_id, driver_id, tenant_id)

        # --- Check route is in DRIVER_ASSIGNED state ---
        if route.status == RouteManagementStatusEnum.ONGOING:
            # Idempotent: already started
            return ResponseWrapper.success(
                message="Duty already started for this route",
                data={
                    "route_id": route.route_id,
                    "route_status": route.status.value,
                },
            )

        if route.status != RouteManagementStatusEnum.DRIVER_ASSIGNED:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail=ResponseWrapper.error(
                    message="Route must be in DRIVER_ASSIGNED state to start duty",
                    error_code="INVALID_ROUTE_STATE",
                    details={"current_status": route.status.value},
                ),
            )

        # --- Check if driver has any other ongoing route ---
        ongoing_route = (
            db.query(RouteManagement)
            .filter(
                RouteManagement.tenant_id == tenant_id,
                RouteManagement.assigned_driver_id == driver_id,
                RouteManagement.status == RouteManagementStatusEnum.ONGOING,
                RouteManagement.route_id != route_id,
            )
            .first()
        )

        if ongoing_route:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail=ResponseWrapper.error(
                    message="Driver already has an ongoing route. Please complete it before starting a new duty.",
                    error_code="DRIVER_HAS_ONGOING_ROUTE",
                    details={
                        "ongoing_route_id": ongoing_route.route_id,
                        "ongoing_route_code": ongoing_route.route_code,
                    },
                ),
            )

        # --- Update route to ONGOING ---
        route.status = RouteManagementStatusEnum.ONGOING
        route.actual_start_time = now
        route.updated_at = now

        db.add(route)
        db.commit()
        db.refresh(route)

        logger.info(f"[driver.start_duty] Duty started for route {route_id} by driver {driver_id}")

        # --- Initialize Firebase node for real-time tracking ---
        driver_obj = db.query(Driver).filter(Driver.driver_id == route.assigned_driver_id).first() if route.assigned_driver_id else None

        vehicle_rc_number = None
        vehicle_type = None
        if route.assigned_vehicle_id:
            vehicle = db.query(Vehicle).filter(Vehicle.vehicle_id == route.assigned_vehicle_id).first()
            if vehicle:
                vehicle_rc_number = vehicle.rc_number
                vehicle_type = vehicle.vehicle_type_name

        background_tasks.add_task(
            _initialize_firebase_node_bg,
            tenant_id = tenant_id,
            vendor_id = ctx.get("vendor_id"),
            driver_id = driver_id,
            driver_name = driver_obj.name if driver_obj else "Unknown",
            driver_code = driver_obj.code if driver_obj else "N/A",
            route_id = route.route_id,
            vehicle_rc_number = vehicle_rc_number,
            vehicle_type = vehicle_type,
        )

        # Send notifications to all employees on this route
        background_tasks.add_task(
            send_duty_start_notifications,
            db=db,
            route_id=route_id,
            driver_id=driver_id
        )

        return ResponseWrapper.success(
            message="Duty started successfully",
            data={
                "route_id": route.route_id,
                "route_status": route.status.value,
            },
        )

    except HTTPException as e:
        logger.warning(f"[driver.start_duty] HTTP error: {e.detail}")
        raise handle_http_error(e)
    except Exception as e:
        logger.exception("[driver.start_duty] Unexpected error")
        db.rollback()
        raise handle_db_error(e)


@router.post("/escort/board", status_code=status.HTTP_200_OK)
async def board_escort(
    route_id: int,
    otp: int = Query(..., description="Escort OTP told verbally by the escort to the driver"),
    db: Session = Depends(get_db),
    ctx=Depends(DriverAuth),
):
    """
    Verify escort OTP and mark the escort as boarded.

    Flow:
    1. Driver starts duty: POST /driver/duty/start
    2. Driver arrives at escort pickup point; escort tells their OTP verbally.
    3. Driver enters that OTP here: POST /driver/escort/board
    4. After escort is confirmed on board, driver can start picking up employees.

    The escort OTP is generated during route dispatch and sent to the escort via SMS.
    """
    try:
        tenant_id = ctx["tenant_id"]
        driver_id = ctx["driver_id"]

        logger.info(f"[driver.escort_board] tenant={tenant_id}, driver={driver_id}, route={route_id}")

        route = validate_route_for_driver(db, route_id, driver_id, tenant_id)

        if route.status != RouteManagementStatusEnum.ONGOING:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail=ResponseWrapper.error(
                    message="Route must be in ONGOING state (duty started) before boarding the escort",
                    error_code="ROUTE_NOT_ONGOING",
                    details={"current_status": route.status.value},
                ),
            )

        if not route.assigned_escort_id:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail=ResponseWrapper.error(
                    message="No escort is assigned to this route",
                    error_code="NO_ESCORT_ASSIGNED",
                    details={"route_id": route_id},
                ),
            )

        # Idempotent
        if route.escort_boarded:
            return ResponseWrapper.success(
                message="Escort already boarded",
                data={"route_id": route_id, "escort_id": route.assigned_escort_id, "escort_boarded": True},
            )

        if not route.escort_otp:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail=ResponseWrapper.error(
                    message="Escort OTP has not been generated yet. Please dispatch the route first.",
                    error_code="ESCORT_OTP_NOT_GENERATED",
                    details={"route_id": route_id},
                ),
            )

        # Verify OTP
        if str(route.escort_otp).strip() != str(otp).strip():
            logger.warning(f"[driver.escort_board] Invalid escort OTP for route {route_id}")
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail=ResponseWrapper.error(
                    message="Invalid escort OTP",
                    error_code="INVALID_ESCORT_OTP",
                    details={"route_id": route_id},
                ),
            )

        # Mark escort as boarded
        route.escort_boarded = True
        db.add(route)
        db.commit()

        logger.info(f"[driver.escort_board] Escort boarded for route {route_id} by driver {driver_id}")

        return ResponseWrapper.success(
            message="Escort boarded successfully. You can now start picking up employees.",
            data={
                "route_id": route_id,
                "escort_id": route.assigned_escort_id,
                "escort_boarded": True,
            },
        )

    except HTTPException:
        raise
    except Exception as e:
        logger.exception("[driver.escort_board] Unexpected error")
        db.rollback()
        raise handle_db_error(e)


@router.get("/trips", status_code=status.HTTP_200_OK)
async def get_driver_trips(
    status_filter: str = Query(..., pattern="^(upcoming|ongoing|completed)$", description="Trip status filter"),
    booking_date: date = Query(default=date.today(), description="Filter trips by booking date (YYYY-MM-DD)"),
    limit: int = Query(default=50, ge=1, le=100, description="Maximum number of routes to return"),
    offset: int = Query(default=0, ge=0, description="Number of routes to skip"),
    db: Session = Depends(get_db),
    ctx=Depends(DriverAuth),
):
    """
    Fetch driver trips by status: upcoming | ongoing | completed.
    Filters by booking_date from the Booking table for `upcoming` and `completed`.
    For `ongoing`, the endpoint returns all currently ongoing routes assigned to the driver regardless
    of the booking_date (covers overnight / timezone edge cases where a route started on a different date).
    Unified structure for mobile driver app.
    Derives start time from the earliest actual/estimated pickup in RouteManagementBooking.
    OPTIMIZED: Uses eager loading to prevent N+1 queries, includes pagination.
    """
    try:
        tenant_id = ctx["tenant_id"]
        driver_id = ctx["driver_id"]

        logger.info(f"[driver.trips] tenant={tenant_id}, driver={driver_id}, status={status_filter}, date={booking_date}, limit={limit}, offset={offset}")

        # --- Map status_filter to RouteManagementStatusEnum ---
        if status_filter == "upcoming":
            status_enum = RouteManagementStatusEnum.DRIVER_ASSIGNED
        elif status_filter == "ongoing":
            status_enum = RouteManagementStatusEnum.ONGOING
        elif status_filter == "completed":
            status_enum = RouteManagementStatusEnum.COMPLETED

        # --- Fetch all routes with eager loading to prevent N+1 ---
        # Use subquery to get distinct route_ids first, then fetch with relationships
        route_ids_subquery = (
            db.query(RouteManagement.route_id)
            .join(RouteManagementBooking, RouteManagementBooking.route_id == RouteManagement.route_id)
            .join(Booking, Booking.booking_id == RouteManagementBooking.booking_id)
            .filter(
                RouteManagement.tenant_id == tenant_id,
                RouteManagement.assigned_driver_id == driver_id,
                RouteManagement.status == status_enum,
            )
        )
        
        if status_filter != "ongoing":
            route_ids_subquery = route_ids_subquery.filter(func.date(Booking.booking_date) == booking_date)
        else:
            logger.info(f"[driver.trips] fetching ongoing routes without date filter to cover edge cases (driver_id={driver_id})")
        
        route_ids_subquery = (
            route_ids_subquery
            .group_by(RouteManagement.route_id)
            .order_by(RouteManagement.created_at.desc())
            .limit(limit)
            .offset(offset)
            .subquery()
        )

        # Fetch routes with all needed data in optimized way
        routes = (
            db.query(RouteManagement)
            .join(route_ids_subquery, RouteManagement.route_id == route_ids_subquery.c.route_id)
            .order_by(RouteManagement.created_at.desc())
            .all()
        )

        if not routes:
            return ResponseWrapper.success(
                data={"routes": [], "count": 0, "limit": limit, "offset": offset},
                message=f"No {status_filter} trips found for {booking_date}",
            )

        # Get all route IDs for batch fetching
        route_ids = [r.route_id for r in routes]
        shift_ids = list(set([r.shift_id for r in routes if r.shift_id]))
        
        # Batch fetch all shifts in ONE query
        shifts_dict = {}
        if shift_ids:
            shifts = db.query(Shift).filter(Shift.shift_id.in_(shift_ids)).all()
            shifts_dict = {s.shift_id: s for s in shifts}
        
        # Batch fetch all bookings for all routes in ONE query
        all_bookings_query = (
            db.query(RouteManagementBooking, Booking, Employee)
            .join(Booking, RouteManagementBooking.booking_id == Booking.booking_id)
            .outerjoin(Employee, Booking.employee_id == Employee.employee_id)
            .filter(RouteManagementBooking.route_id.in_(route_ids))
        )
        
        if status_filter != "ongoing":
            all_bookings_query = all_bookings_query.filter(func.date(Booking.booking_date) == booking_date)
        
        all_bookings_rows = all_bookings_query.order_by(
            RouteManagementBooking.route_id,
            RouteManagementBooking.order_id
        ).all()
        
        # Group bookings by route_id for efficient lookup
        bookings_by_route = {}
        for rb, booking, employee in all_bookings_rows:
            if rb.route_id not in bookings_by_route:
                bookings_by_route[rb.route_id] = []
            bookings_by_route[rb.route_id].append((rb, booking, employee))

        response_routes = []

        for route in routes:
            rows = bookings_by_route.get(route.route_id, [])
            
            if not rows:
                continue

            stops, pickup_datetimes = [], []

            for rb, booking, employee in rows:
                booking_date_str = booking.booking_date.isoformat()
                est_pick = getattr(rb, "estimated_pick_up_time", None)
                act_pick = getattr(rb, "actual_pick_up_time", None)
                pick_time_str = act_pick or est_pick

                if pick_time_str:
                    try:
                        dt = datetime.combine(
                            booking.booking_date,
                            datetime.strptime(pick_time_str, "%H:%M").time(),
                        )
                        pickup_datetimes.append(dt)
                    except Exception:
                        pass

                stops.append({
                    "booking_id": booking.booking_id,
                    "tenant_id": booking.tenant_id,
                    "employee_id": booking.employee_id,
                    "employee_code": getattr(employee, "employee_code", None),
                    "employee_name": getattr(employee, "name", None),
                    "employee_phone": getattr(employee, "phone", None),
                    "shift_id": booking.shift_id,
                    "team_id": booking.team_id,
                    "booking_date": booking_date_str,
                    "pickup_latitude": booking.pickup_latitude,
                    "pickup_longitude": booking.pickup_longitude,
                    "pickup_location": booking.pickup_location,
                    "drop_latitude": booking.drop_latitude,
                    "drop_longitude": booking.drop_longitude,
                    "drop_location": booking.drop_location,
                    "status": booking.status.value if booking.status else None,
                    "booking_type": booking.booking_type.value if booking.booking_type else None,
                    "reason": booking.reason,
                    "is_active": getattr(booking, "is_active", True),
                     "is_boarding_otp_required": booking.boarding_otp is not None,
                     "is_deboarding_otp_required": booking.deboarding_otp is not None,
                     "nodal_point_id": booking.nodal_point_id,
                     "is_nodal_stop": booking.nodal_point_id is not None,
                    "created_at": booking.created_at.isoformat() if booking.created_at else None,
                    "updated_at": booking.updated_at.isoformat() if booking.updated_at else None,
                    "order_id": rb.order_id,
                    "estimated_pick_up_time": est_pick,
                    "estimated_drop_time": rb.estimated_drop_time,
                    "estimated_distance": rb.estimated_distance,
                    "actual_pick_up_time": act_pick,
                    "actual_drop_time": rb.actual_drop_time,
                    "actual_distance": rb.actual_distance,
                })

            # --- Derive first pickup ---
            if pickup_datetimes:
                first_pickup_dt = min(pickup_datetimes)
            else:
                first_pickup_dt = datetime.combine(booking_date, datetime.min.time())

            # Use batch-loaded shift (no additional query!)
            shift = shifts_dict.get(route.shift_id) if route.shift_id else None
            pickup_type_val = shift.pickup_type.value if shift and shift.pickup_type else None

            # ── Hub groups for NODAL routes ─────────────────────────────────
            hub_groups = None
            if shift and shift.pickup_type == PickupTypeEnum.NODAL:
                from collections import defaultdict
                grouped: dict = defaultdict(list)
                for s in stops:
                    key = s.get("nodal_point_id") or "unassigned"
                    grouped[key].append({
                        "booking_id": s["booking_id"],
                        "employee_id": s["employee_id"],
                        "employee_name": s.get("employee_name"),
                        "employee_phone": s.get("employee_phone"),
                        "status": s["status"],
                        "order_id": s["order_id"],
                        "is_boarding_otp_required": s.get("is_boarding_otp_required"),
                        "is_boarded": s["status"] == BookingStatusEnum.ONGOING.value,
                    })
                # Batch-load nodal point details
                nodal_ids = [k for k in grouped if isinstance(k, int)]
                np_map = {
                    np.nodal_point_id: np
                    for np in db.query(NodalPoint)
                    .filter(NodalPoint.nodal_point_id.in_(nodal_ids))
                    .all()
                } if nodal_ids else {}
                hub_groups = []
                for key in sorted(grouped.keys(), key=lambda x: x if isinstance(x, int) else 9999999):
                    np_obj = np_map.get(key)
                    hub_groups.append({
                        "nodal_point_id": key if isinstance(key, int) else None,
                        "hub_name": np_obj.name if np_obj else None,
                        "hub_address": np_obj.address if np_obj else None,
                        "hub_latitude": float(np_obj.latitude) if np_obj else None,
                        "hub_longitude": float(np_obj.longitude) if np_obj else None,
                        "passenger_count": len(grouped[key]),
                        "boarded_count": sum(1 for b in grouped[key] if b.get("is_boarded")),
                        "pending_count": sum(1 for b in grouped[key] if not b.get("is_boarded")),
                        "bookings": grouped[key],
                    })

            response_routes.append({
                "route_id": route.route_id,
                "tenant_id": route.tenant_id,
                "shift_id": route.shift_id,
                "route_code": route.route_code,
                "assigned_vendor_id": getattr(route, "assigned_vendor_id", None),
                "assigned_vehicle_id": getattr(route, "assigned_vehicle_id", None),
                "assigned_driver_id": getattr(route, "assigned_driver_id", None),
                "assigned_escort_id": getattr(route, "assigned_escort_id", None),
                "escort_required": route.escort_required,
                "escort_boarded": getattr(route, "escort_boarded", False),
                "shift_time": shift.shift_time.strftime("%H:%M:%S") if shift and shift.shift_time else None,
                "log_type": shift.log_type.value if shift and shift.log_type else None,
                "pickup_type": pickup_type_val,
                "status": route.status.value,
                "estimated_total_time": route.estimated_total_time,
                "estimated_total_distance": route.estimated_total_distance,
                "actual_total_time": route.actual_total_time,
                "actual_total_distance": route.actual_total_distance,
                "buffer_time": route.buffer_time,
                "start_time": first_pickup_dt.strftime("%Y-%m-%d %H:%M"),
                "stops": stops,
                "hub_groups": hub_groups,
                "summary": {
                    "total_stops": len(stops),
                    "total_distance_km": float(route.actual_total_distance or route.estimated_total_distance or 0),
                    "total_time_minutes": float(route.actual_total_time or route.estimated_total_time or 0),
                },
            })

        # --- Sorting logic ---
        if status_filter == "upcoming":
            response_routes.sort(key=lambda r: datetime.strptime(r["start_time"], "%Y-%m-%d %H:%M"))
        else:
            response_routes.sort(key=lambda r: r["start_time"], reverse=True)

        return ResponseWrapper.success(
            data={
                "routes": response_routes, 
                "count": len(response_routes),
                "limit": limit,
                "offset": offset,
                "has_more": len(response_routes) == limit
            },
            message=f"Fetched {len(response_routes)} {status_filter} routes for {booking_date}",
        )

    except HTTPException as e:
        logger.warning(f"[driver.trips] HTTP error: {e.detail}")
        raise handle_http_error(e)
    except Exception as e:
        logger.exception("[driver.trips] Unexpected error")
        raise handle_db_error(e)

@router.get("/history/report", status_code=status.HTTP_200_OK)
async def driver_history_report(
    # Date range — accept both naming conventions so driver app and web frontend
    # both work without a coordinated rename.
    start_date: Optional[date] = Query(None, description="Start date (YYYY-MM-DD)"),
    end_date:   Optional[date] = Query(None, description="End date (YYYY-MM-DD)"),
    from_date:  Optional[date] = Query(None, description="Start date alias used by web frontend (YYYY-MM-DD)"),
    to_date:    Optional[date] = Query(None, description="End date alias used by web frontend (YYYY-MM-DD)"),
    # Driver ID — required for admin / manager callers; auto-resolved from JWT for drivers
    driver_id: Optional[int] = Query(None, description="Driver ID (required for admin/manager; ignored for driver JWT)"),
    format: str = Query("json", pattern="^(json|excel)$", description="Response format: json or excel"),
    db: Session = Depends(get_db),
    # Accept EITHER driver_app.read (driver JWT) OR report.read (admin / transport manager)
    user_data=Depends(PermissionChecker(["driver_app.read", "report.read"], check_tenant=False)),
):
    """
    Driver trip history report for a date range.

    Accessible by:
    - **Driver** (JWT with driver_app.read) — sees only their own trips; driver_id
      is resolved from the JWT, the query param is ignored.
    - **Admin / Transport Manager / Employee** (JWT with report.read) — must supply
      an explicit `driver_id` query param.

    Date params accept two naming conventions:
    - `start_date` / `end_date`  (driver app)
    - `from_date`  / `to_date`   (web frontend)

    Use `format=excel` to download an XLSX file instead of JSON.
    """
    import io
    try:
        # ── Resolve dates (accept either naming convention) ──────────────────────
        resolved_start = start_date or from_date
        resolved_end   = end_date   or to_date

        if not resolved_start or not resolved_end:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail=ResponseWrapper.error(
                    message="start_date (or from_date) and end_date (or to_date) are required",
                    error_code="MISSING_DATE_RANGE",
                ),
            )

        if resolved_start > resolved_end:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail=ResponseWrapper.error(
                    message="start_date must be on or before end_date",
                    error_code="INVALID_DATE_RANGE",
                ),
            )

        # ── Resolve caller identity ───────────────────────────────────────────────
        user_type = user_data.get("user_type")
        tenant_id = user_data.get("tenant_id")

        if user_type == "driver":
            # Driver: always use their own ID from the JWT
            resolved_driver_id = user_data.get("user_id")
        else:
            # Admin / employee (transport manager) / vendor
            if not driver_id:
                raise HTTPException(
                    status_code=status.HTTP_400_BAD_REQUEST,
                    detail=ResponseWrapper.error(
                        message="driver_id is required for admin and manager users",
                        error_code="DRIVER_ID_REQUIRED",
                    ),
                )
            resolved_driver_id = driver_id

        if not tenant_id or not resolved_driver_id:
            raise HTTPException(
                status_code=status.HTTP_403_FORBIDDEN,
                detail=ResponseWrapper.error(
                    message="Could not resolve tenant or driver from token",
                    error_code="AUTH_CONTEXT_MISSING",
                ),
            )

        logger.info(
            "[driver.history_report] caller_type=%s driver=%s tenant=%s %s→%s format=%s",
            user_type, resolved_driver_id, tenant_id, resolved_start, resolved_end, format,
        )

        # ── Single optimised query: join Route ▸ RMBooking ▸ Booking ▸ Employee ─
        rows = (
            db.query(
                RouteManagement,
                RouteManagementBooking,
                Booking,
                Employee,
            )
            .join(RouteManagementBooking, RouteManagementBooking.route_id == RouteManagement.route_id)
            .join(Booking, Booking.booking_id == RouteManagementBooking.booking_id)
            .outerjoin(Employee, Employee.employee_id == Booking.employee_id)
            .filter(
                RouteManagement.tenant_id == tenant_id,
                RouteManagement.assigned_driver_id == resolved_driver_id,
                RouteManagement.status == RouteManagementStatusEnum.COMPLETED,
                Booking.booking_date >= resolved_start,
                Booking.booking_date <= resolved_end,
            )
            .order_by(Booking.booking_date.asc(), RouteManagementBooking.order_id.asc())
            .all()
        )

        # ── Build flat record list ───────────────────────────────────────────────
        records = []
        for route, rb, booking, employee in rows:
            records.append({
                "route_id": route.route_id,
                "route_code": route.route_code or "",
                "route_status": route.status.value if route.status else "",
                "estimated_total_distance_km": route.estimated_total_distance,
                "actual_total_distance_km": route.actual_total_distance,
                "booking_id": booking.booking_id,
                "booking_date": booking.booking_date.isoformat(),
                "booking_status": booking.status.value if booking.status else "",
                "booking_type": booking.booking_type.value if booking.booking_type else "",
                "order_in_route": rb.order_id,
                "employee_id": booking.employee_id,
                "employee_name": getattr(employee, "name", None),
                "employee_code": getattr(employee, "employee_code", None),
                "pickup_location": booking.pickup_location,
                "drop_location": booking.drop_location,
                "estimated_pickup_time": rb.estimated_pick_up_time,
                "actual_pickup_time": rb.actual_pick_up_time,
                "estimated_drop_time": rb.estimated_drop_time,
                "actual_drop_time": rb.actual_drop_time,
                "estimated_distance_km": rb.estimated_distance,
                "actual_distance_km": rb.actual_distance,
            })

        # ── Summary counts ───────────────────────────────────────────────────────
        total_bookings = len(records)
        completed = sum(1 for r in records if r["booking_status"] == "Completed")
        no_show = sum(1 for r in records if r["booking_status"] == "No-Show")
        cancelled = sum(1 for r in records if r["booking_status"] == "Cancelled")
        total_actual_km = sum(
            r["actual_distance_km"] or 0 for r in records
        )
        unique_routes = len(set(r["route_id"] for r in records))

        summary = {
            "start_date": resolved_start.isoformat(),
            "end_date": resolved_end.isoformat(),
            "total_routes": unique_routes,
            "total_bookings": total_bookings,
            "completed": completed,
            "no_show": no_show,
            "cancelled": cancelled,
            "total_actual_km": round(total_actual_km, 2),
        }

        # ── JSON response ────────────────────────────────────────────────────────
        if format == "json":
            return ResponseWrapper.success(
                data={"summary": summary, "bookings": records},
                message=f"Driver history report: {resolved_start} to {resolved_end}",
            )

        # ── Excel response ───────────────────────────────────────────────────────
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.utils import get_column_letter
            from fastapi.responses import StreamingResponse
        except ImportError:
            raise HTTPException(
                status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
                detail=ResponseWrapper.error(
                    message="openpyxl is required for Excel export",
                    error_code="MISSING_DEPENDENCY",
                ),
            )

        wb = Workbook()

        # ── Sheet 1: Summary ─────────────────────────────────────────────────────
        ws_sum = wb.active
        ws_sum.title = "Summary"

        header_fill = PatternFill("solid", fgColor="2C5AA0")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        label_font = Font(bold=True, size=10)
        thin = Side(style="thin", color="CCCCCC")
        cell_border = Border(left=thin, right=thin, top=thin, bottom=thin)

        ws_sum.append(["Driver History Report"])
        ws_sum["A1"].font = Font(bold=True, size=14, color="2C5AA0")
        ws_sum.append([f"Period: {start_date}  →  {end_date}"])
        ws_sum.append([f"Generated: {datetime.utcnow().strftime('%Y-%m-%d %H:%M UTC')}"])
        ws_sum.append([])

        for label, value in [
            ("Total Routes", summary["total_routes"]),
            ("Total Bookings", summary["total_bookings"]),
            ("Completed", summary["completed"]),
            ("No-Show", summary["no_show"]),
            ("Cancelled", summary["cancelled"]),
            ("Total Actual KM", summary["total_actual_km"]),
        ]:
            row = ws_sum.append([label, value])
            # style label cell
            cell_a = ws_sum.cell(ws_sum.max_row, 1)
            cell_a.font = label_font
            cell_a.border = cell_border
            ws_sum.cell(ws_sum.max_row, 2).border = cell_border

        ws_sum.column_dimensions["A"].width = 22
        ws_sum.column_dimensions["B"].width = 18

        # ── Sheet 2: Booking Detail ───────────────────────────────────────────────
        ws = wb.create_sheet("Booking Detail")

        columns = [
            ("Route ID",                    "route_id"),
            ("Route Code",                  "route_code"),
            ("Route Status",                "route_status"),
            ("Booking ID",                  "booking_id"),
            ("Booking Date",                "booking_date"),
            ("Booking Status",              "booking_status"),
            ("Booking Type",                "booking_type"),
            ("Order in Route",              "order_in_route"),
            ("Employee ID",                 "employee_id"),
            ("Employee Name",               "employee_name"),
            ("Employee Code",               "employee_code"),
            ("Pickup Location",             "pickup_location"),
            ("Drop Location",               "drop_location"),
            ("Est. Pickup Time",            "estimated_pickup_time"),
            ("Actual Pickup Time",          "actual_pickup_time"),
            ("Est. Drop Time",              "estimated_drop_time"),
            ("Actual Drop Time",            "actual_drop_time"),
            ("Est. Distance (km)",          "estimated_distance_km"),
            ("Actual Distance (km)",        "actual_distance_km"),
            ("Route Est. Total Dist (km)",  "estimated_total_distance_km"),
            ("Route Act. Total Dist (km)",  "actual_total_distance_km"),
        ]

        # Header row
        for col_idx, (header, _) in enumerate(columns, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", wrap_text=True)
            cell.border = cell_border

        # Data rows
        alt_fill = PatternFill("solid", fgColor="F0F4FF")
        for row_idx, record in enumerate(records, 2):
            fill = alt_fill if row_idx % 2 == 0 else None
            for col_idx, (_, key) in enumerate(columns, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=record.get(key))
                cell.alignment = Alignment(horizontal="center")
                cell.border = cell_border
                if fill:
                    cell.fill = fill

        # Auto-width columns
        for col_idx, (header, _) in enumerate(columns, 1):
            col_letter = get_column_letter(col_idx)
            max_len = len(header)
            for row_idx in range(2, len(records) + 2):
                val = ws.cell(row=row_idx, column=col_idx).value
                if val is not None:
                    max_len = max(max_len, len(str(val)))
            ws.column_dimensions[col_letter].width = min(max_len + 4, 40)

        ws.freeze_panes = "A2"

        # Write to in-memory buffer
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        filename = f"driver_{resolved_driver_id}_report_{resolved_start}_to_{resolved_end}.xlsx"
        from fastapi.responses import StreamingResponse as SR
        return SR(
            content=buffer,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f'attachment; filename="{filename}"'},
        )

    except HTTPException:
        raise
    except Exception as e:
        logger.exception("[driver.history_report] Unexpected error")
        raise handle_db_error(e)


@router.post("/trip/start", status_code=status.HTTP_200_OK)
async def start_trip(
    route_id: int,
    booking_id: int,
    current_latitude: float = Query(..., description="Driver's current latitude"),
    current_longitude: float = Query(..., description="Driver's current longitude"),
    background_tasks: BackgroundTasks = BackgroundTasks(),
    otp: Optional[str] = None,
    db: Session = Depends(get_db),
    ctx=Depends(DriverAuth),
):
    """
    Start pickup for a booking (employee boards the vehicle).
    - Route must already be in ONGOING state (duty started)
    - Verify OTP and location
    - Update booking status to ONGOING
    - Update actual_pick_up_time
    - Return next stop details
    """
    try:
        tenant_id = ctx["tenant_id"]
        driver_id = ctx["driver_id"]

        logger.info(f"[driver.start_trip] tenant={tenant_id}, driver={driver_id}, route={route_id}, booking={booking_id}")

        # --- Validate route exists and is assigned to driver ---
        route = validate_route_for_driver(db, route_id, driver_id, tenant_id)
        
        if not route:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail=ResponseWrapper.error(
                    message="Route not found or not assigned to driver",
                    error_code="ROUTE_NOT_FOUND",
                    details={"route_id": route_id, "driver_id": driver_id},
                ),
            )

        # Route must be ONGOING (duty already started)
        if route.status != RouteManagementStatusEnum.ONGOING:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail=ResponseWrapper.error(
                    message="Route must be in ONGOING state. Please start duty first.",
                    error_code="ROUTE_NOT_ONGOING",
                    details={"current_status": route.status.value},
                ),
            )

        # --- Escort must be boarded before picking up any employees ---
        if route.assigned_escort_id and not getattr(route, "escort_boarded", False):
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail=ResponseWrapper.error(
                    message="Escort must be boarded before picking up employees. Use POST /driver/escort/board first.",
                    error_code="ESCORT_NOT_BOARDED",
                    details={"route_id": route_id, "escort_id": route.assigned_escort_id},
                ),
            )

        # --- Validate booking in route ---
        rb, booking = validate_booking_in_route(db, route_id, booking_id, tenant_id)
        
        if not rb:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail=ResponseWrapper.error(
                    message="Booking not found in this route",
                    error_code="BOOKING_NOT_FOUND_IN_ROUTE",
                    details={"route_id": route_id, "booking_id": booking_id},
                ),
            )

        # ── Detect nodal shift ──────────────────────────────────────────────
        is_nodal_shift = False
        if booking.shift_id:
            _shift = db.query(Shift).filter(Shift.shift_id == booking.shift_id).first()
            if _shift and _shift.pickup_type == PickupTypeEnum.NODAL:
                is_nodal_shift = True

        # ── Sequential gate (skip for nodal — hub passengers board simultaneously) ──
        if not is_nodal_shift:
            check_previous_bookings_completed(db, route_id, rb.order_id)

        # Booking object already validated by validate_booking_in_route

        # --- Validate driver's location is near pickup location ---
        validate_driver_location(
            current_latitude, current_longitude,
            booking.pickup_latitude, booking.pickup_longitude,
            "pickup", booking_id
        )

        # ── F12: Female Employee Dark-Hour Boarding Block ────────────────────
        _dh_warnings: list = []
        try:
            from app.services.dark_hour_boarding_service import check_dark_hour_boarding
            from app.models.user_session import UserSession
            _cfg_dh = db.query(TenantConfig).filter(
                TenantConfig.tenant_id == tenant_id
            ).first()
            _emp_dh = db.query(Employee).filter(
                Employee.employee_id == booking.employee_id
            ).first()
            if _cfg_dh and _emp_dh:
                _escort_ok = bool(
                    route.assigned_escort_id
                    and getattr(route, "escort_boarded", False)
                )
                _gender_val = (
                    _emp_dh.gender.value
                    if _emp_dh.gender is not None
                    else None
                )
                _now_time = get_current_ist_time().time()
                _dh = check_dark_hour_boarding(
                    gender=_gender_val,
                    escort_present_and_boarded=_escort_ok,
                    cfg=_cfg_dh,
                    now_time=_now_time,
                )
                if not _dh["ok"]:
                    # Hard block — fire security notification before raising
                    background_tasks.add_task(
                        send_dark_hour_block_notification,
                        db=db,
                        tenant_id=tenant_id,
                        booking_id=booking_id,
                        employee_name=_emp_dh.name or "Female Employee",
                    )
                    raise HTTPException(
                        status_code=status.HTTP_423_LOCKED,
                        detail=ResponseWrapper.error(
                            message=(
                                "Boarding blocked: female employee in dark hours "
                                "without a boarded escort."
                            ),
                            error_code="DARK_HOUR_NO_ESCORT",
                            details={"booking_id": booking_id, "route_id": route_id},
                        ),
                    )
                _dh_warnings = _dh.get("warnings", [])
        except HTTPException:
            raise
        except Exception as _dh_exc:
            # Fail open — dark-hour check must never block a trip on error
            logger.warning(
                "[driver.start_trip] Dark-hour check failed (failing open): %s",
                _dh_exc,
            )

        # ── Boarding confirmation ────────────────────────────────────────────
        now = get_current_ist_time()
        if is_nodal_shift:
            # Nodal: QR scan already moves booking to ONGOING; driver call just
            # records pick-up time and fetches next stop.  If the employee hasn't
            # scanned yet (SCHEDULED), the driver can still confirm them manually.
            if booking.status not in [BookingStatusEnum.SCHEDULED, BookingStatusEnum.ONGOING]:
                raise HTTPException(
                    status_code=status.HTTP_400_BAD_REQUEST,
                    detail=ResponseWrapper.error(
                        message="Booking is not in a boardable state for nodal confirmation",
                        error_code="BOOKING_NOT_BOARDABLE",
                        details={"current_status": booking.status.value},
                    ),
                )
            if booking.status == BookingStatusEnum.SCHEDULED:
                booking.status = BookingStatusEnum.ONGOING
                logger.info(
                    f"[driver.start_trip] Nodal manual board: "
                    f"booking={booking_id} set to ONGOING"
                )
        else:
            # Normal pickup: verify boarding OTP and set status
            verify_otp(booking.boarding_otp, otp, "boarding", booking_id)
            booking.status = BookingStatusEnum.ONGOING

        # Record actual pick-up time if not already set by QR scan
        if not rb.actual_pick_up_time:
            rb.actual_pick_up_time = now.strftime("%H:%M")

        db.add_all([booking, rb])
        db.commit()

        logger.info(
            f"[driver.start_trip] Trip started: route={route_id}, booking={booking_id}, "
            f"actual_pick_up_time={rb.actual_pick_up_time}"
        )

        # Send onboard notification to employee
        background_tasks.add_task(
            send_onboard_notification,
            db=db,
            booking_id=booking_id,
            route_id=route_id
        )

        # --- Fetch next stop ---
        next_stop = get_next_stop(db, route_id, rb.order_id, skip_ongoing=is_nodal_shift)

        return ResponseWrapper.success(
            message="Trip started successfully",
            data={
                "route_id": route.route_id,
                "route_status": route.status.value,
                "started_at": rb.actual_pick_up_time,
                "current_booking_id": booking.booking_id,
                "current_status": booking.status.value,
                "actual_pick_up_time": rb.actual_pick_up_time,
                "next_stop": next_stop,
                "warnings": _dh_warnings,
            },
        )

    except HTTPException as e:
        logger.warning(f"[driver.start_trip] HTTP error: {e.detail}")
        raise handle_http_error(e)
    except Exception as e:
        logger.exception("[driver.start_trip] Unexpected error")
        db.rollback()
        return ResponseWrapper.error(message=str(e))

@router.put("/trip/no-show", status_code=status.HTTP_200_OK)
async def mark_no_show(
    route_id: int,
    booking_id: int,
    background_tasks: BackgroundTasks = BackgroundTasks(),
    reason: Optional[str] = Query(None, description="Reason for marking as no-show"),
    db: Session = Depends(get_db),
    ctx=Depends(DriverAuth),
):
    """
    Mark a booking as NO_SHOW when employee did not board.
    - Allowed only if route is assigned to the driver.
    - Updates booking.status = NO_SHOW.
    - Updates route.status = ONGOING if it wasn’t already.
    - If route has only one booking → auto-complete route.
    - Otherwise, only mark route COMPLETED if all DROPS are completed.
    """
    try:
        tenant_id = ctx["tenant_id"]
        driver_id = ctx["driver_id"]
        now = get_current_ist_time()

        logger.info(f"[driver.no_show] tenant={tenant_id}, driver={driver_id}, route={route_id}, booking={booking_id}")

        # --- Validate route ---
        route = validate_route_for_driver(db, route_id, driver_id, tenant_id)
        
        if not route:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail=ResponseWrapper.error(
                    message="Route not found or not assigned to driver",
                    error_code="ROUTE_NOT_FOUND",
                    details={"route_id": route_id, "driver_id": driver_id},
                ),
            )

        # --- Validate booking in route ---
        rb, booking = validate_booking_in_route(db, route_id, booking_id, tenant_id)
        
        # --- Prevent marking ongoing or completed bookings as no-show ---
        if booking.status in [BookingStatusEnum.ONGOING, BookingStatusEnum.COMPLETED]:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail=ResponseWrapper.error(
                    message="Cannot mark no-show for an active or completed booking",
                    error_code="INVALID_BOOKING_STATE",
                    details={"booking_status": booking.status.value},
                ),
            )

        # --- Allow marking current booking as no-show only if all previous stops are done ---
        check_previous_bookings_completed(db, route_id, rb.order_id)

        # --- Update booking as NO_SHOW (DO NOT change route status) ---
        booking.status = BookingStatusEnum.NO_SHOW
        booking.reason = reason or "Employee did not board"
        booking.updated_at = now
        rb.actual_pick_up_time = now.strftime("%H:%M")

        db.add_all([booking, rb])
        db.commit()

        logger.info(
            f"[driver.no_show] Booking {booking_id} marked NO_SHOW; route={route_id}, driver={driver_id}"
        )

        # Send no-show notification to employee
        background_tasks.add_task(
            send_no_show_notification,
            db=db,
            booking_id=booking_id,
            reason=reason or "Employee did not board"
        )

        # --- Get next stop ---
        next_stop = get_next_stop(db, route_id, rb.order_id)

        return ResponseWrapper.success(
            message="Booking marked as no-show successfully",
            data={
                "route_id": route.route_id,
                "route_status": route.status.value,
                "booking_id": booking.booking_id,
                "booking_status": booking.status.value,
                "actual_pick_up_time": rb.actual_pick_up_time,
                "next_stop": next_stop,
            },
        )

    except HTTPException as e:
        logger.warning(f"[driver.no_show] HTTP error: {e.detail}")
        raise handle_http_error(e)
    except Exception as e:
        logger.exception("[driver.no_show] Unexpected error")
        db.rollback()
        return ResponseWrapper.error(message=str(e))

@router.put("/trip/drop", status_code=status.HTTP_200_OK)
async def verify_drop_and_complete_route(
    route_id: int,
    booking_id: int,
    current_latitude: float = Query(..., description="Driver's current latitude"),
    current_longitude: float = Query(..., description="Driver's current longitude"),
    background_tasks: BackgroundTasks = BackgroundTasks(),
    otp: Optional[str] = None,
    db: Session = Depends(get_db),
    ctx=Depends(DriverAuth),
):
    """
    Driver confirms drop for a booking.
    - Marks booking as COMPLETED and sets actual_drop_time.
    - If all other bookings are NO_SHOW or COMPLETED, auto-completes the route.
    """
    try:
        tenant_id = ctx["tenant_id"]
        driver_id = ctx["driver_id"]
        now = get_current_ist_time()

        logger.info(f"[driver.drop] tenant={tenant_id}, driver={driver_id}, route={route_id}, booking={booking_id}")

        # --- Validate route ---
        route = validate_route_for_driver(
            db, route_id, driver_id, tenant_id, 
            required_status=RouteManagementStatusEnum.ONGOING
        )

        # --- Validate booking in route ---
        rb, booking = validate_booking_in_route(db, route_id, booking_id, tenant_id)
        
        # --- Validate driver's location is near drop location ---
        validate_driver_location(
            current_latitude, current_longitude,
            booking.drop_latitude, booking.drop_longitude,
            "drop", booking_id
        )

        # --- Verify deboarding OTP if present ---
        verify_otp(booking.deboarding_otp, otp, "deboarding", booking_id)

        # --- Prevent re-marking if already dropped ---
        if booking.status == BookingStatusEnum.COMPLETED:
            return ResponseWrapper.success(
                message="Booking already marked as completed",
                data={
                    "route_id": route_id,
                    "booking_id": booking_id,
                    "status": booking.status.value,
                    "actual_drop_time": rb.actual_drop_time,
                },
            )

        # --- Mark booking as completed (DO NOT complete route) ---
        booking.status = BookingStatusEnum.COMPLETED
        booking.updated_at = now
        rb.actual_drop_time = now.strftime("%H:%M")

        db.add_all([booking, rb])
        db.commit()

        logger.info(f"[driver.drop] Booking {booking_id} marked as completed by driver {driver_id}")

        # Send drop completion notification to employee
        background_tasks.add_task(
            send_drop_completion_notification,
            booking_id=booking_id,
            route_id=route_id
        )

        return ResponseWrapper.success(
            message="Drop verified successfully",
            data={
                "route_id": route.route_id,
                "booking_id": booking.booking_id,
                "booking_status": booking.status.value,
                "actual_drop_time": rb.actual_drop_time,
                "route_status": route.status.value,
            },
        )

    except HTTPException as e:
        db.rollback()
        logger.warning(f"[driver.drop] HTTP error: {e.detail}")
        raise handle_http_error(e)
    except Exception as e:
        db.rollback()
        logger.exception("[driver.drop] Unexpected error")
        raise handle_db_error(e)


@router.post("/location", status_code=status.HTTP_200_OK)
async def update_driver_location(
    route_id: int,
    latitude: float,
    longitude: float,
    background_tasks: BackgroundTasks,
    speed: Optional[float] = Query(None, description="Speed in km/h reported by the device"),
    db: Session = Depends(get_db),
    ctx=Depends(DriverAuth),
):
    """
    IMP-1 / IMP-2 / IMP-9 — GPS location ping endpoint.

    Called by the driver app periodically (e.g. every 5–10 s) while a route is ONGOING.

    Actions performed:
      1. Validates the route is ONGOING and belongs to this driver.
      2. Writes the coordinates to driver_location_history (PostgreSQL — full trail).
      3. Pushes the latest position to Firebase RTDB in a BackgroundTask
         (non-blocking — a Firebase failure never fails the HTTP response).
      4. IMP-7: Runs geofence check — if driver is within arrival radius of next
         stop, pushes "Driver arriving" FCM to the waiting employee (BackgroundTask).
      5. IMP-6: Recalculates ETAs for all remaining stops and pushes FCM to
         affected employees if ETA changed by more than the tenant threshold
         (BackgroundTask).

    Returns a minimal 200 so the mobile client can ACK and move on quickly.
    """
    try:
        tenant_id = ctx["tenant_id"]
        driver_id = ctx["driver_id"]
        vendor_id = ctx.get("vendor_id")
        now       = get_current_ist_time()

        logger.debug(
            "[driver.location] tenant=%s driver=%s route=%s lat=%.6f lng=%.6f",
            tenant_id, driver_id, route_id, latitude, longitude,
        )

        # --- Validate route is ONGOING and belongs to this driver ---
        route = validate_route_for_driver(db, route_id, driver_id, tenant_id)

        if route.status != RouteManagementStatusEnum.ONGOING:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail=ResponseWrapper.error(
                    message="Location pings are only accepted for ONGOING routes",
                    error_code="ROUTE_NOT_ONGOING",
                    details={"route_status": route.status.value},
                ),
            )

        # --- Persist GPS breadcrumb to PostgreSQL ---
        from app.models.driver_location_history import DriverLocationHistory

        ping = DriverLocationHistory(
            tenant_id   = tenant_id,
            route_id    = route_id,
            driver_id   = driver_id,
            vendor_id   = vendor_id,
            latitude    = latitude,
            longitude   = longitude,
            speed       = speed,
            recorded_at = now,
        )
        db.add(ping)
        db.commit()

        # --- Push latest position to Firebase RTDB (best-effort, non-blocking) ---
        driver_obj = db.query(Driver).filter(Driver.driver_id == route.assigned_driver_id).first() if route.assigned_driver_id else None
        background_tasks.add_task(
            _push_location_to_firebase_bg,
            tenant_id  = tenant_id,
            vendor_id  = vendor_id,
            driver_id  = driver_id,
            latitude   = latitude,
            longitude  = longitude,
            speed      = speed,
            driver_name = driver_obj.name if driver_obj else None,
            driver_code = driver_obj.code if driver_obj else None,
            route_id   = route.route_id,
        )

        # --- IMP-7: Geofence arrival check (non-blocking) ---
        background_tasks.add_task(
            _geofence_check_bg,
            tenant_id  = tenant_id,
            route_id   = route_id,
            driver_lat = latitude,
            driver_lng = longitude,
            db         = db,
        )

        # --- IMP-6: ETA recalculation for remaining stops (non-blocking) ---
        background_tasks.add_task(
            _eta_recalc_bg,
            tenant_id          = tenant_id,
            route_id           = route_id,
            driver_lat         = latitude,
            driver_lng         = longitude,
            driver_speed_kmph  = speed,
            now                = now,
            db                 = db,
        )

        # --- IMP-10: Server-side speed violation check (non-blocking, only when speed reported) ---
        if speed is not None:
            background_tasks.add_task(
                _speed_violation_check_bg,
                tenant_id   = tenant_id,
                route_id    = route_id,
                driver_id   = driver_id,
                vehicle_id  = route.assigned_vehicle_id,
                speed_kmph  = speed,
                latitude    = latitude,
                longitude   = longitude,
                recorded_at = now,
                db          = db,
            )

        return ResponseWrapper.success(
            message="Location updated",
            data={
                "route_id":  route_id,
                "latitude":  latitude,
                "longitude": longitude,
            },
        )

    except HTTPException as e:
        logger.warning("[driver.location] HTTP error: %s", e.detail)
        raise handle_http_error(e)
    except Exception as e:
        db.rollback()
        logger.exception("[driver.location] Unexpected error")
        raise handle_db_error(e)


def _push_location_to_firebase_bg(
    tenant_id: str,
    vendor_id: int,
    driver_id: int,
    latitude: float,
    longitude: float,
    speed: float = None,
    driver_name: str = None,
    driver_code: str = None,
    route_id: int = None,
) -> None:
    """
    Background task wrapper for Firebase location push.
    Swallows all exceptions so a Firebase failure never propagates to the HTTP layer.
    """
    try:
        from app.firebase.driver_location import push_driver_location_to_firebase
        push_driver_location_to_firebase(
            tenant_id = tenant_id,
            vendor_id = vendor_id,
            driver_id = driver_id,
            latitude  = latitude,
            longitude = longitude,
            speed     = speed,
            driver_name = driver_name,
            driver_code = driver_code,
            route_id  = route_id,
        )
    except Exception as exc:
        logger.exception(
            "[driver.location] Firebase background push failed for driver %s: %s",
            driver_id, exc,
        )


def _initialize_firebase_node_bg(
    tenant_id: str,
    vendor_id: int,
    driver_id: int,
    driver_name: str,
    driver_code: str,
    route_id: int,
    vehicle_rc_number: Optional[str] = None,
    vehicle_type: Optional[str] = None,
) -> None:
    """
    Background task wrapper for Firebase node initialization on duty start.
    Swallows all exceptions so a Firebase failure never propagates to the HTTP layer.
    """
    try:
        from app.firebase.driver_location import initialize_driver_node_on_duty_start
        initialize_driver_node_on_duty_start(
            tenant_id = tenant_id,
            vendor_id = vendor_id,
            driver_id = driver_id,
            driver_name = driver_name,
            driver_code = driver_code,
            route_id = route_id,
            vehicle_rc_number = vehicle_rc_number,
            vehicle_type = vehicle_type,
        )
        
    except Exception as exc:
        logger.exception(
            "[driver.start_duty] Firebase node initialization failed for driver %s: %s",
            driver_id, exc,
        )


def _geofence_check_bg(
    tenant_id: str,
    route_id: int,
    driver_lat: float,
    driver_lng: float,
    db: Session,
) -> None:
    """
    IMP-7 — Background task wrapper for geofence arrival check.
    Swallows all exceptions so a geofence failure never propagates to the HTTP layer.
    """
    try:
        from app.services.geofence_service import check_and_fire_arrival_geofence
        check_and_fire_arrival_geofence(
            db         = db,
            tenant_id  = tenant_id,
            route_id   = route_id,
            driver_lat = driver_lat,
            driver_lng = driver_lng,
        )
    except Exception as exc:
        logger.exception(
            "[driver.location] Geofence check failed for route %s: %s",
            route_id, exc,
        )


def _eta_recalc_bg(
    tenant_id: str,
    route_id: int,
    driver_lat: float,
    driver_lng: float,
    driver_speed_kmph: Optional[float],
    now: "datetime",
    db: Session,
) -> None:
    """
    IMP-6 — Background task wrapper for ETA recalculation.
    Swallows all exceptions so an ETA failure never propagates to the HTTP layer.
    """
    try:
        from app.services.eta_service import recalculate_eta_for_remaining_stops
        recalculate_eta_for_remaining_stops(
            db                = db,
            tenant_id         = tenant_id,
            route_id          = route_id,
            driver_lat        = driver_lat,
            driver_lng        = driver_lng,
            driver_speed_kmph = driver_speed_kmph,
            now               = now,
        )
    except Exception as exc:
        logger.exception(
            "[driver.location] ETA recalc failed for route %s: %s",
            route_id, exc,
        )


def _clear_firebase_location_bg(
    tenant_id: str,
    vendor_id: int,
    driver_id: int,
) -> None:
    """
    IMP-11 — Background task wrapper for Firebase location node cleanup.
    Marks the driver's RTDB node as offline (is_active=False).
    Swallows all exceptions so a Firebase failure never propagates to the HTTP layer.
    """
    try:
        from app.firebase.driver_location import clear_driver_location_from_firebase
        clear_driver_location_from_firebase(
            tenant_id = tenant_id,
            vendor_id = vendor_id,
            driver_id = driver_id,
        )
    except Exception as exc:
        logger.exception(
            "[driver.end_duty] Firebase cleanup failed for driver %s: %s",
            driver_id, exc,
        )


def _speed_violation_check_bg(
    tenant_id: str,
    route_id: int,
    driver_id: int,
    vehicle_id: Optional[int],
    speed_kmph: float,
    latitude: float,
    longitude: float,
    recorded_at: "datetime",
    db: Session,
) -> None:
    """
    IMP-10 — Background task wrapper for server-side speed violation detection.
    Swallows all exceptions so a violation check failure never propagates to
    the HTTP layer.
    """
    try:
        from app.services.speed_violation_service import detect_and_record_speed_violation
        detect_and_record_speed_violation(
            db          = db,
            tenant_id   = tenant_id,
            route_id    = route_id,
            driver_id   = driver_id,
            vehicle_id  = vehicle_id,
            speed_kmph  = speed_kmph,
            latitude    = latitude,
            longitude   = longitude,
            recorded_at = recorded_at,
        )
    except Exception as exc:
        logger.exception(
            "[driver.location] Speed violation check failed for route %s: %s",
            route_id, exc,
        )


@router.put("/duty/end", status_code=status.HTTP_200_OK)
async def end_duty(
    route_id: int,
    reason: Optional[str] = Query(None, description="Reason for ending duty"),
    background_tasks: BackgroundTasks = BackgroundTasks(),
    db: Session = Depends(get_db),
    ctx=Depends(DriverAuth),
):
    """
    End the driver's duty for a route.
    - Only allowed for routes assigned to the driver and in ONGOING state.
    - Only allowed if all bookings are COMPLETED, NO_SHOW, or CANCELLED.
    - Completes the route and sets actual end time.
    """
    try:
        tenant_id = ctx["tenant_id"]
        driver_id = ctx["driver_id"]
        vendor_id = ctx.get("vendor_id")
        now = get_current_ist_time().replace(tzinfo=None)

        logger.info(f"[driver.end_duty] tenant={tenant_id}, driver={driver_id}, route={route_id}")
        route = validate_route_for_driver(db, route_id, driver_id, tenant_id)
        
        # Only allow ending a route that is ongoing
        if route.status != RouteManagementStatusEnum.ONGOING:
            # Idempotent: if already completed, return success
            if route.status == RouteManagementStatusEnum.COMPLETED:
                return ResponseWrapper.success(message="Route already completed", data={"route_id": route.route_id, "route_status": route.status.value})
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail=ResponseWrapper.error(
                    message="Route must be ongoing to end duty",
                    error_code="INVALID_ROUTE_STATE",
                    details={"route_status": route.status.value},
                ),
            )

        # --- Check if all bookings are finalized (OPTIMIZED: use exists() first) ---
        has_pending = db.query(
            exists().where(
                RouteManagementBooking.route_id == route_id,
                RouteManagementBooking.booking_id == Booking.booking_id,
                Booking.status.notin_([BookingStatusEnum.COMPLETED, BookingStatusEnum.NO_SHOW, BookingStatusEnum.CANCELLED])
            )
        ).scalar()

        if has_pending:
            # Only count if we need the number for error message
            pending_bookings = (
                db.query(RouteManagementBooking)
                .join(Booking, RouteManagementBooking.booking_id == Booking.booking_id)
                .filter(
                    RouteManagementBooking.route_id == route_id,
                    Booking.status.notin_([BookingStatusEnum.COMPLETED, BookingStatusEnum.NO_SHOW, BookingStatusEnum.CANCELLED]),
                )
                .count()
            )
            
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail=ResponseWrapper.error(
                    message="Cannot end duty. All bookings must be completed or marked as no-show first.",
                    error_code="PENDING_BOOKINGS_EXIST",
                    details={"pending_count": pending_bookings},
                ),
            )

        # --- Complete the route ---
        route.status = RouteManagementStatusEnum.COMPLETED
        route.actual_end_time = now
        route.updated_at = now

        # --- Tag OTD delay (best-effort; never blocks duty completion) ---
        try:
            tag_trip_delay(db=db, route=route, now=now)
        except Exception as delay_err:
            logger.warning(
                "[driver.end_duty] Delay tagging failed for route %s (non-fatal): %s",
                route_id, delay_err,
            )

        # --- IMP-8: Compute actual GPS distance (best-effort; never blocks duty completion) ---
        try:
            from app.services.distance_service import compute_and_persist_actual_distance
            compute_and_persist_actual_distance(db=db, route=route)
        except Exception as dist_err:
            logger.warning(
                "[driver.end_duty] Distance computation failed for route %s (non-fatal): %s",
                route_id, dist_err,
            )

        db.add(route)
        db.commit()

        logger.info(f"[driver.end_duty] Route {route_id} completed by driver {driver_id}.")

        # IMP-11 — clear Firebase node (mark driver offline, best-effort)
        background_tasks.add_task(
            _clear_firebase_location_bg,
            tenant_id = tenant_id,
            vendor_id = vendor_id,
            driver_id = driver_id,
        )

        return ResponseWrapper.success(
            message="Duty ended and route closed",
            data={
                "route_id": route_id,
                "route_status": route.status.value,
            },
        )

    except HTTPException as e:
        db.rollback()
        logger.warning(f"[driver.end_duty] HTTP error: {e.detail}")
        raise handle_http_error(e)
    except Exception as e:
        db.rollback()
        logger.exception("[driver.end_duty] Unexpected error")
        raise handle_db_error(e)


# ---------------------------
# Background Notification Tasks
# ---------------------------

def send_duty_start_notifications(db: Session, route_id: int, driver_id: int):
    """
    Send notifications to all employees on the route when driver starts duty.
    """
    try:
        logger.info(f"[notify.duty_start] Sending notifications for route {route_id}")
        
        # Get driver details
        driver = db.query(Driver).filter(Driver.driver_id == driver_id).first()
        driver_name = driver.name if driver else f"Driver {driver_id}"
        
        # Get route details
        route = db.query(RouteManagement).filter(RouteManagement.route_id == route_id).first()
        if not route:
            logger.warning(f"[notify.duty_start] Route {route_id} not found")
            return
        
        # Get all bookings for this route with employee details
        bookings_data = (
            db.query(RouteManagementBooking, Booking, Employee)
            .join(Booking, RouteManagementBooking.booking_id == Booking.booking_id)
            .outerjoin(Employee, Booking.employee_id == Employee.employee_id)
            .filter(RouteManagementBooking.route_id == route_id)
            .all()
        )
        
        if not bookings_data:
            logger.warning(f"[notify.duty_start] No bookings found for route {route_id}")
            return
        
        # Initialize services
        email_service = EmailService()
        sms_service = SMSService()
        push_service = UnifiedNotificationService(db)
        
        success_count = 0
        
        for rb, booking, employee in bookings_data:
            if not employee:
                logger.warning(f"[notify.duty_start] Employee not found for booking {booking.booking_id}")
                continue
            
            employee_name = employee.name or "Employee"
            
            # Email notification (commented out)
            # try:
            #     email_service.send_email(
            #         to_email=employee.email,
            #         subject=f"Driver Started - Route {route.route_code}",
            #         body_text=f"Hi {employee_name},\\n\\n"
            #                   f"Your driver {driver_name} has started duty and is on the way for pickup.\\n"
            #                   f"Route: {route.route_code}\\n"
            #                   f"Estimated pickup time: {rb.estimated_pick_up_time}\\n\\n"
            #                   f"Please be ready at your pickup location.\\n\\n"
            #                   f"Thank you,\\nFleet Manager",
            #         body_html=f"<p>Hi {employee_name},</p>"
            #                  f"<p>Your driver <strong>{driver_name}</strong> has started duty and is on the way for pickup.</p>"
            #                  f"<p><strong>Route:</strong> {route.route_code}<br>"
            #                  f"<strong>Estimated pickup time:</strong> {rb.estimated_pick_up_time}</p>"
            #                  f"<p>Please be ready at your pickup location.</p>"
            #                  f"<p>Thank you,<br>Fleet Manager</p>"
            #     )
            #     logger.info(f"[notify.duty_start] Email sent to {employee.email}")
            # except Exception as e:
            #     logger.error(f"[notify.duty_start] Email failed for {employee.email}: {str(e)}")
            
            # SMS notification (commented out)
            # if employee.phone and sms_service.enabled:
            #     try:
            #         sms_service.send_sms(
            #             to_phone=employee.phone,
            #             message=f"Hi {employee_name}, your driver {driver_name} has started duty. "
            #                    f"Route: {route.route_code}. Est. pickup: {rb.estimated_pick_up_time}. Be ready!"
            #         )
            #         logger.info(f"[notify.duty_start] SMS sent to {employee.phone}")
            #     except Exception as e:
            #         logger.error(f"[notify.duty_start] SMS failed for {employee.phone}: {str(e)}")
            
            # Push notification
            try:
                push_service.send_to_user(
                    user_type="employee",
                    user_id=employee.employee_id,
                    title="Driver Started",
                    body=f"Your driver {driver_name} is on the way. Est. pickup: {rb.estimated_pick_up_time}",
                    data={
                        "type": "duty_started",
                        "route_id": str(route_id),
                        "route_code": route.route_code,
                        "driver_name": driver_name,
                        "estimated_pickup": rb.estimated_pick_up_time or ""
                    },
                    priority="high"
                )
                logger.info(f"[notify.duty_start] Push sent to employee {employee.employee_id}")
            except Exception as e:
                logger.error(f"[notify.duty_start] Push failed for employee {employee.employee_id}: {str(e)}")
            
            success_count += 1
        
        logger.info(f"[notify.duty_start] Sent notifications to {success_count} employees for route {route_id}")
        
    except Exception as e:
        logger.exception(f"[notify.duty_start] Error sending notifications: {str(e)}")


def send_onboard_notification(db: Session, booking_id: int, route_id: int):
    """
    Send notification to employee when they board the vehicle.
    """
    try:
        logger.info(f"[notify.onboard] Sending notification for booking {booking_id}")
        
        # Get booking and employee details
        booking = db.query(Booking).filter(Booking.booking_id == booking_id).first()
        if not booking:
            logger.warning(f"[notify.onboard] Booking {booking_id} not found")
            return
        
        employee = db.query(Employee).filter(Employee.employee_id == booking.employee_id).first()
        if not employee:
            logger.warning(f"[notify.onboard] Employee not found for booking {booking_id}")
            return
        
        route = db.query(RouteManagement).filter(RouteManagement.route_id == route_id).first()
        route_code = route.route_code if route else f"Route {route_id}"
        
        employee_name = employee.name or "Employee"
        
        # Initialize services
        email_service = EmailService()
        sms_service = SMSService()
        push_service = UnifiedNotificationService(db)
        
        # Email notification (commented out)
        # try:
        #     email_service.send_email(
        #         to_email=employee.email,
        #         subject=f"Onboard Confirmed - {route_code}",
        #         body_text=f"Hi {employee_name},\\n\\n"
        #                   f"You have successfully boarded the vehicle.\\n"
        #                   f"Route: {route_code}\\n"
        #                   f"Drop location: {booking.drop_location}\\n\\n"
        #                   f"Have a safe journey!\\n\\n"
        #                   f"Thank you,\\nFleet Manager",
        #         body_html=f"<p>Hi {employee_name},</p>"
        #                  f"<p>You have successfully boarded the vehicle.</p>"
        #                  f"<p><strong>Route:</strong> {route_code}<br>"
        #                  f"<strong>Drop location:</strong> {booking.drop_location}</p>"
        #                  f"<p>Have a safe journey!</p>"
        #                  f"<p>Thank you,<br>Fleet Manager</p>"
        #     )
        #     logger.info(f"[notify.onboard] Email sent to {employee.email}")
        # except Exception as e:
        #     logger.error(f"[notify.onboard] Email failed: {str(e)}")
        
        # SMS notification (commented out)
        # if employee.phone and sms_service.enabled:
        #     try:
        #         sms_service.send_sms(
        #             to_phone=employee.phone,
        #             message=f"Hi {employee_name}, you're now onboard! Route: {route_code}. "
        #                    f"Drop: {booking.drop_location}. Safe journey!"
        #         )
        #         logger.info(f"[notify.onboard] SMS sent to {employee.phone}")
        #     except Exception as e:
        #         logger.error(f"[notify.onboard] SMS failed: {str(e)}")
        
        # Push notification
        try:
            push_service.send_to_user(
                user_type="employee",
                user_id=employee.employee_id,
                title="Onboard Confirmed",
                body=f"You're now onboard! Drop: {booking.drop_location}",
                data={
                    "type": "onboard",
                    "route_id": str(route_id),
                    "booking_id": str(booking_id),
                    "drop_location": booking.drop_location or ""
                },
                priority="high"
            )
            logger.info(f"[notify.onboard] Push sent to employee {employee.employee_id}")
        except Exception as e:
            logger.error(f"[notify.onboard] Push failed: {str(e)}")
        
        logger.info(f"[notify.onboard] Notifications sent for booking {booking_id}")
        
    except Exception as e:
        logger.exception(f"[notify.onboard] Error: {str(e)}")


def send_no_show_notification(db: Session, booking_id: int, reason: str):
    """
    Send notification to employee when marked as no-show.
    """
    try:
        logger.info(f"[notify.no_show] Sending notification for booking {booking_id}")
        
        # Get booking and employee details
        booking = db.query(Booking).filter(Booking.booking_id == booking_id).first()
        if not booking:
            logger.warning(f"[notify.no_show] Booking {booking_id} not found")
            return
        
        employee = db.query(Employee).filter(Employee.employee_id == booking.employee_id).first()
        if not employee:
            logger.warning(f"[notify.no_show] Employee not found for booking {booking_id}")
            return
        
        employee_name = employee.name or "Employee"
        
        # Initialize services
        email_service = EmailService()
        sms_service = SMSService()
        push_service = UnifiedNotificationService(db)
        
        # Email notification (commented out)
        # try:
        #     email_service.send_email(
        #         to_email=employee.email,
        #         subject="No-Show Alert - Booking Cancelled",
        #         body_text=f"Hi {employee_name},\\n\\n"
        #                   f"You were marked as NO-SHOW for your booking on {booking.booking_date}.\\n"
        #                   f"Reason: {reason}\\n"
        #                   f"Pickup location: {booking.pickup_location}\\n\\n"
        #                   f"If this was a mistake, please contact support.\\n\\n"
        #                   f"Thank you,\\nFleet Manager",
        #         body_html=f"<p>Hi {employee_name},</p>"
        #                  f"<p>You were marked as <strong>NO-SHOW</strong> for your booking on {booking.booking_date}.</p>"
        #                  f"<p><strong>Reason:</strong> {reason}<br>"
        #                  f"<strong>Pickup location:</strong> {booking.pickup_location}</p>"
        #                  f"<p>If this was a mistake, please contact support.</p>"
        #                  f"<p>Thank you,<br>Fleet Manager</p>"
        #     )
        #     logger.info(f"[notify.no_show] Email sent to {employee.email}")
        # except Exception as e:
        #     logger.error(f"[notify.no_show] Email failed: {str(e)}")
        
        # SMS notification (commented out)
        # if employee.phone and sms_service.enabled:
        #     try:
        #         sms_service.send_sms(
        #             to_phone=employee.phone,
        #             message=f"Hi {employee_name}, you were marked NO-SHOW for {booking.booking_date}. "
        #                    f"Reason: {reason}. Contact support if this was a mistake."
        #         )
        #         logger.info(f"[notify.no_show] SMS sent to {employee.phone}")
        #     except Exception as e:
        #         logger.error(f"[notify.no_show] SMS failed: {str(e)}")
        
        # Push notification
        try:
            push_service.send_to_user(
                user_type="employee",
                user_id=employee.employee_id,
                title="No-Show Alert",
                body=f"You were marked NO-SHOW. Reason: {reason}",
                data={
                    "type": "no_show",
                    "booking_id": str(booking_id),
                    "reason": reason,
                    "booking_date": str(booking.booking_date)
                },
                priority="high"
            )
            logger.info(f"[notify.no_show] Push sent to employee {employee.employee_id}")
        except Exception as e:
            logger.error(f"[notify.no_show] Push failed: {str(e)}")
        
        logger.info(f"[notify.no_show] Notifications sent for booking {booking_id}")
        
    except Exception as e:
        logger.exception(f"[notify.no_show] Error: {str(e)}")


def send_drop_completion_notification(booking_id: int, route_id: int):
    """
    Send notification to employee when successfully dropped off.
    """
    db = SessionLocal()
    try:
        logger.info(f"[notify.drop] Sending notification for booking {booking_id}")
        
        # Get booking and employee details
        booking = db.query(Booking).filter(Booking.booking_id == booking_id).first()
        if not booking:
            logger.warning(f"[notify.drop] Booking {booking_id} not found")
            return
        
        employee = db.query(Employee).filter(Employee.employee_id == booking.employee_id).first()
        if not employee:
            logger.warning(f"[notify.drop] Employee not found for booking {booking_id}")
            return
        
        route = db.query(RouteManagement).filter(RouteManagement.route_id == route_id).first()
        route_code = route.route_code if route else f"Route {route_id}"
        
        # Get actual drop time
        rb = db.query(RouteManagementBooking).filter(
            RouteManagementBooking.route_id == route_id,
            RouteManagementBooking.booking_id == booking_id
        ).first()
        actual_drop_time = rb.actual_drop_time if rb else "now"
        
        employee_name = employee.name or "Employee"
        
        # Initialize services
        email_service = EmailService()
        sms_service = SMSService()
        push_service = UnifiedNotificationService(db)
        
        # Email notification (commented out)
        # try:
        #     email_service.send_email(
        #         to_email=employee.email,
        #         subject=f"Drop Completed - {route_code}",
        #         body_text=f"Hi {employee_name},\\n\\n"
        #                   f"You have been successfully dropped off.\\n"
        #                   f"Route: {route_code}\\n"
        #                   f"Drop location: {booking.drop_location}\\n"
        #                   f"Drop time: {actual_drop_time}\\n\\n"
        #                   f"Thank you for using our service!\\n\\n"
        #                   f"Best regards,\\nFleet Manager",
        #         body_html=f"<p>Hi {employee_name},</p>"
        #                  f"<p>You have been successfully dropped off.</p>"
        #                  f"<p><strong>Route:</strong> {route_code}<br>"
        #                  f"<strong>Drop location:</strong> {booking.drop_location}<br>"
        #                  f"<strong>Drop time:</strong> {actual_drop_time}</p>"
        #                  f"<p>Thank you for using our service!</p>"
        #                  f"<p>Best regards,<br>Fleet Manager</p>"
        #     )
        #     logger.info(f"[notify.drop] Email sent to {employee.email}")
        # except Exception as e:
        #     logger.error(f"[notify.drop] Email failed: {str(e)}")
        
        # SMS notification (commented out)
        # if employee.phone and sms_service.enabled:
        #     try:
        #         sms_service.send_sms(
        #             to_phone=employee.phone,
        #             message=f"Hi {employee_name}, you've been dropped off successfully at {booking.drop_location}. "
        #                    f"Time: {actual_drop_time}. Thank you!"
        #         )
        #         logger.info(f"[notify.drop] SMS sent to {employee.phone}")
        #     except Exception as e:
        #         logger.error(f"[notify.drop] SMS failed: {str(e)}")
        
        # Push notification
        try:
            push_service.send_to_user(
                user_type="employee",
                user_id=employee.employee_id,
                title="Drop Completed",
                body=f"Successfully dropped at {booking.drop_location}. Time: {actual_drop_time}",
                data={
                    "type": "drop_completed",
                    "route_id": str(route_id),
                    "booking_id": str(booking_id),
                    "drop_location": booking.drop_location or "",
                    "drop_time": actual_drop_time or ""
                },
                priority="normal"
            )
            logger.info(f"[notify.drop] Push sent to employee {employee.employee_id}")
        except Exception as e:
            logger.error(f"[notify.drop] Push failed: {str(e)}")
        
        logger.info(f"[notify.drop] Notifications sent for booking {booking_id}")
        
    except Exception as e:
        logger.exception(f"[notify.drop] Error: {str(e)}")
    finally:
        db.close()


def send_dark_hour_block_notification(
    db: Session,
    tenant_id: str,
    booking_id: int,
    employee_name: str,
) -> None:
    """
    F12 — Female Employee Dark-Hour Boarding Block

    Fires a security push notification to every active admin session for the
    tenant when a female employee's boarding is hard-blocked due to dark-hour
    rules and no boarded escort.

    Queries UserSession for admin users and sends via UnifiedNotificationService.
    Runs as a BackgroundTask — all errors are swallowed to avoid disrupting the
    HTTP error response already raised in start_trip.
    """
    try:
        from app.models.user_session import UserSession

        admin_sessions = (
            db.query(UserSession)
            .filter(
                UserSession.tenant_id == tenant_id,
                UserSession.user_type == "admin",
                UserSession.is_active == True,
                UserSession.fcm_token.isnot(None),
            )
            .all()
        )

        if not admin_sessions:
            logger.info(
                "[notify.dark_hour_block] No active admin sessions for tenant %s",
                tenant_id,
            )
            return

        push_service = UnifiedNotificationService(db)
        title = "Security Alert: Dark-Hour Boarding Blocked"
        body = (
            f"{employee_name} attempted to board without a safety escort "
            f"during restricted hours (booking #{booking_id}). "
            "Boarding was blocked by system policy."
        )
        data_payload = {
            "type": "dark_hour_block",
            "booking_id": str(booking_id),
            "tenant_id": tenant_id,
            "employee_name": employee_name,
        }

        for session in admin_sessions:
            try:
                push_service.send_to_user(
                    user_type="admin",
                    user_id=session.user_id,
                    title=title,
                    body=body,
                    data=data_payload,
                    priority="high",
                )
            except Exception as _e:
                logger.warning(
                    "[notify.dark_hour_block] Push failed for admin %s: %s",
                    session.user_id,
                    _e,
                )

        logger.info(
            "[notify.dark_hour_block] Security notification sent to %d admin(s) "
            "for tenant %s, booking %d",
            len(admin_sessions),
            tenant_id,
            booking_id,
        )

    except Exception as e:
        logger.exception(
            "[notify.dark_hour_block] Unexpected error for booking %d: %s",
            booking_id,
            e,
        )
