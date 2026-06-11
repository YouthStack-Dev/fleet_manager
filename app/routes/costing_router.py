from datetime import date, datetime, time, timedelta
from io import BytesIO
from typing import Any, Dict, List, Optional, Tuple

import openpyxl
from fastapi import APIRouter, Depends, HTTPException, Query, status
from fastapi.responses import StreamingResponse
from openpyxl.styles import Font, PatternFill
from sqlalchemy import func, or_
from sqlalchemy.exc import SQLAlchemyError
from sqlalchemy.orm import Session

from app.database.session import get_db
from app.models.booking import Booking
from app.models.costing import (
    CostCenter,
    GarageConfig,
    RateCard,
    RateCardDistanceSlab,
    RateCardSlot,
    RouteBookingCost,
    RouteCost,
    RouteCostAllocation,
    RouteExpense,
)
from app.models.route_management import RouteManagement
from app.models.vehicle import Vehicle
from app.models.vehicle_type import VehicleType
from app.models.vendor import Vendor
from app.schemas.costing import (
    GarageConfigCreate,
    GarageConfigResponse,
    GarageConfigUpdate,
    RateCardCreate,
    RateCardResponse,
    RateCardSlotCreate,
    RateCardSlotResponse,
    RateCardSlotUpdate,
    RateCardUpdate,
    RouteCostActionRequest,
    RouteCostCalculateRequest,
    RouteExpenseCreate,
    RouteExpenseRejectRequest,
    RouteExpenseResponse,
    RouteExpenseUpdate,
)
from app.services.costing_service import CostingService
from app.utils.response_utils import ResponseWrapper, handle_db_error
from common_utils.auth.permission_checker import PermissionChecker


router = APIRouter(tags=["costing"])


def _value(value: Any) -> Any:
    return value.value if hasattr(value, "value") else value


def resolve_tenant(user_data: dict, tenant_id: Optional[str], *, allow_vendor: bool = False) -> str:
    user_type = user_data.get("user_type")
    token_tenant_id = user_data.get("tenant_id")
    if user_type == "employee":
        tenant_id = token_tenant_id
    elif user_type == "vendor" and allow_vendor:
        tenant_id = token_tenant_id
    elif user_type == "admin":
        tenant_id = token_tenant_id or tenant_id
    else:
        raise HTTPException(status_code=403, detail=ResponseWrapper.error("Forbidden", "FORBIDDEN"))
    if not tenant_id:
        raise HTTPException(status_code=400, detail=ResponseWrapper.error("tenant_id is required", "TENANT_ID_REQUIRED"))
    return tenant_id


def validate_vendor(db: Session, tenant_id: str, vendor_id: Optional[int]) -> None:
    if vendor_id is None:
        return
    if not db.query(Vendor).filter(Vendor.vendor_id == vendor_id, Vendor.tenant_id == tenant_id).first():
        raise HTTPException(status_code=404, detail=ResponseWrapper.error("Vendor not found for this tenant", "VENDOR_NOT_FOUND"))


def validate_vehicle_type(db: Session, tenant_id: str, vehicle_type_id: Optional[int]) -> None:
    if vehicle_type_id is None:
        return
    exists = (
        db.query(VehicleType)
        .join(Vendor, Vendor.vendor_id == VehicleType.vendor_id)
        .filter(VehicleType.vehicle_type_id == vehicle_type_id, Vendor.tenant_id == tenant_id)
        .first()
    )
    if not exists:
        raise HTTPException(status_code=404, detail=ResponseWrapper.error("Vehicle type not found for this tenant", "VEHICLE_TYPE_NOT_FOUND"))


def validate_vehicle(db: Session, tenant_id: str, vehicle_id: Optional[int]) -> None:
    if vehicle_id is None:
        return
    exists = (
        db.query(Vehicle)
        .join(Vendor, Vendor.vendor_id == Vehicle.vendor_id)
        .filter(Vehicle.vehicle_id == vehicle_id, Vendor.tenant_id == tenant_id)
        .first()
    )
    if not exists:
        raise HTTPException(status_code=404, detail=ResponseWrapper.error("Vehicle not found for this tenant", "VEHICLE_NOT_FOUND"))





def route_for_user(db: Session, route_id: int, tenant_id: str, user_data: dict, *, allow_vendor: bool = False) -> RouteManagement:
    route = db.query(RouteManagement).filter(RouteManagement.route_id == route_id, RouteManagement.tenant_id == tenant_id).first()
    if not route:
        raise HTTPException(status_code=404, detail=ResponseWrapper.error("Route not found", "ROUTE_NOT_FOUND"))
    if allow_vendor and user_data.get("user_type") == "vendor":
        token_vendor_id = user_data.get("vendor_id")
        if not token_vendor_id or route.assigned_vendor_id != int(token_vendor_id):
            raise HTTPException(status_code=403, detail=ResponseWrapper.error("Vendor cannot access this route", "VENDOR_FORBIDDEN"))
    return route


def minute_of_day(value: Optional[time]) -> int:
    if not value:
        return 0
    return value.hour * 60 + value.minute


def time_intervals(start: Optional[time], end: Optional[time]) -> List[Tuple[int, int]]:
    if not start or not end:
        return [(0, 1440)]
    start_min = minute_of_day(start)
    end_min = minute_of_day(end)
    if start_min <= end_min:
        return [(start_min, end_min)]
    return [(start_min, 1440), (0, end_min)]


def intervals_overlap(a: List[Tuple[int, int]], b: List[Tuple[int, int]]) -> bool:
    return any(max(a_start, b_start) <= min(a_end, b_end) for a_start, a_end in a for b_start, b_end in b)


def ensure_slot_no_overlap(db: Session, rate_card_id: int, payload: Any, exclude_slot_id: Optional[int] = None) -> None:
    q = db.query(RateCardSlot).filter(
        RateCardSlot.rate_card_id == rate_card_id,
        RateCardSlot.is_active == True,
        RateCardSlot.shift_log_type == _value(payload.shift_log_type),
        RateCardSlot.day_type == _value(payload.day_type),
    )
    if exclude_slot_id:
        q = q.filter(RateCardSlot.slot_id != exclude_slot_id)
    new_intervals = time_intervals(payload.start_time, payload.end_time)
    for slot in q.all():
        if intervals_overlap(new_intervals, time_intervals(slot.start_time, slot.end_time)):
            raise HTTPException(status_code=409, detail=ResponseWrapper.error("Rate slot overlaps an existing active slot", "RATE_SLOT_OVERLAP"))


def create_distance_slabs(db: Session, slot_id: int, distance_slabs: List[Any]) -> None:
    for slab_payload in distance_slabs:
        slab = RateCardDistanceSlab(
            slot_id=slot_id,
            **{key: _value(value) for key, value in slab_payload.model_dump().items()},
        )
        db.add(slab)


@router.post("/costing/rate-cards", status_code=status.HTTP_201_CREATED)
def create_rate_card(
    payload: RateCardCreate,
    db: Session = Depends(get_db),
    user_data=Depends(PermissionChecker(["costing_rate_card.create"], check_tenant=True)),
):
    try:
        tenant_id = resolve_tenant(user_data, payload.tenant_id)
        validate_vendor(db, tenant_id, payload.vendor_id)
        validate_vehicle_type(db, tenant_id, payload.vehicle_type_id)
        card = RateCard(
            tenant_id=tenant_id,
            vendor_id=payload.vendor_id,
            vehicle_type_id=payload.vehicle_type_id,
            name=payload.name,
            currency=payload.currency,
            effective_from=payload.effective_from,
            effective_to=payload.effective_to,
            status=payload.status.value,
        )
        db.add(card)
        db.commit()
        db.refresh(card)
        return ResponseWrapper.success({"rate_card": RateCardResponse.model_validate(card)}, "Rate card created successfully")
    except HTTPException:
        raise
    except SQLAlchemyError as exc:
        db.rollback()
        raise handle_db_error(exc)


@router.get("/costing/rate-cards", status_code=status.HTTP_200_OK)
def list_rate_cards(
    tenant_id: Optional[str] = Query(None),
    vendor_id: Optional[int] = Query(None),
    status_filter: Optional[str] = Query(None, alias="status"),
    db: Session = Depends(get_db),
    user_data=Depends(PermissionChecker(["costing_rate_card.read"], check_tenant=True)),
):
    tenant_id = resolve_tenant(user_data, tenant_id)
    query = db.query(RateCard).filter(RateCard.tenant_id == tenant_id)
    if vendor_id:
        query = query.filter(RateCard.vendor_id == vendor_id)
    if status_filter:
        query = query.filter(RateCard.status == status_filter)
    cards = query.order_by(RateCard.effective_from.desc(), RateCard.rate_card_id.desc()).all()
    return ResponseWrapper.success({"rate_cards": [RateCardResponse.model_validate(card) for card in cards], "total": len(cards)}, "Rate cards fetched successfully")


@router.patch("/costing/rate-cards/{rate_card_id}", status_code=status.HTTP_200_OK)
def update_rate_card(
    rate_card_id: int,
    payload: RateCardUpdate,
    tenant_id: Optional[str] = Query(None),
    db: Session = Depends(get_db),
    user_data=Depends(PermissionChecker(["costing_rate_card.update"], check_tenant=True)),
):
    try:
        tenant_id = resolve_tenant(user_data, tenant_id)
        card = db.query(RateCard).filter(RateCard.rate_card_id == rate_card_id, RateCard.tenant_id == tenant_id).first()
        if not card:
            raise HTTPException(status_code=404, detail=ResponseWrapper.error("Rate card not found", "RATE_CARD_NOT_FOUND"))
        updates = payload.model_dump(exclude_unset=True)
        if "vendor_id" in updates:
            validate_vendor(db, tenant_id, updates["vendor_id"])
        if "vehicle_type_id" in updates:
            validate_vehicle_type(db, tenant_id, updates["vehicle_type_id"])
        for key, value in updates.items():
            setattr(card, key, _value(value))
        db.commit()
        db.refresh(card)
        return ResponseWrapper.success({"rate_card": RateCardResponse.model_validate(card)}, "Rate card updated successfully")
    except HTTPException:
        raise
    except SQLAlchemyError as exc:
        db.rollback()
        raise handle_db_error(exc)


@router.post("/costing/rate-cards/{rate_card_id}/activate", status_code=status.HTTP_200_OK)
def activate_rate_card(
    rate_card_id: int,
    tenant_id: Optional[str] = Query(None),
    db: Session = Depends(get_db),
    user_data=Depends(PermissionChecker(["costing_rate_card.approve"], check_tenant=True)),
):
    tenant_id = resolve_tenant(user_data, tenant_id)
    card = db.query(RateCard).filter(RateCard.rate_card_id == rate_card_id, RateCard.tenant_id == tenant_id).first()
    if not card:
        raise HTTPException(status_code=404, detail=ResponseWrapper.error("Rate card not found", "RATE_CARD_NOT_FOUND"))
    if not card.slots:
        raise HTTPException(status_code=400, detail=ResponseWrapper.error("Rate card requires at least one slot", "RATE_CARD_HAS_NO_SLOTS"))
    card.status = "active"
    db.commit()
    return ResponseWrapper.success({"rate_card": RateCardResponse.model_validate(card)}, "Rate card activated successfully")


@router.post("/costing/rate-cards/{rate_card_id}/slots", status_code=status.HTTP_201_CREATED)
def create_rate_card_slot(
    rate_card_id: int,
    payload: RateCardSlotCreate,
    tenant_id: Optional[str] = Query(None),
    db: Session = Depends(get_db),
    user_data=Depends(PermissionChecker(["costing_rate_card.update"], check_tenant=True)),
):
    try:
        tenant_id = resolve_tenant(user_data, tenant_id)
        card = db.query(RateCard).filter(RateCard.rate_card_id == rate_card_id, RateCard.tenant_id == tenant_id).first()
        if not card:
            raise HTTPException(status_code=404, detail=ResponseWrapper.error("Rate card not found", "RATE_CARD_NOT_FOUND"))
        if payload.is_active:
            ensure_slot_no_overlap(db, rate_card_id, payload)
        slot = RateCardSlot(rate_card_id=rate_card_id, **{key: _value(value) for key, value in payload.model_dump(exclude={"distance_slabs"}).items()})
        db.add(slot)
        db.flush()
        create_distance_slabs(db, slot.slot_id, payload.distance_slabs)
        db.commit()
        db.refresh(slot)
        return ResponseWrapper.success({"slot": RateCardSlotResponse.model_validate(slot)}, "Rate card slot created successfully")
    except HTTPException:
        raise
    except SQLAlchemyError as exc:
        db.rollback()
        raise handle_db_error(exc)


@router.patch("/costing/rate-cards/{rate_card_id}/slots/{slot_id}", status_code=status.HTTP_200_OK)
def update_rate_card_slot(
    rate_card_id: int,
    slot_id: int,
    payload: RateCardSlotUpdate,
    tenant_id: Optional[str] = Query(None),
    db: Session = Depends(get_db),
    user_data=Depends(PermissionChecker(["costing_rate_card.update"], check_tenant=True)),
):
    try:
        tenant_id = resolve_tenant(user_data, tenant_id)
        card = db.query(RateCard).filter(RateCard.rate_card_id == rate_card_id, RateCard.tenant_id == tenant_id).first()
        if not card:
            raise HTTPException(status_code=404, detail=ResponseWrapper.error("Rate card not found", "RATE_CARD_NOT_FOUND"))
        slot = db.query(RateCardSlot).filter(RateCardSlot.slot_id == slot_id, RateCardSlot.rate_card_id == rate_card_id).first()
        if not slot:
            raise HTTPException(status_code=404, detail=ResponseWrapper.error("Rate card slot not found", "RATE_SLOT_NOT_FOUND"))
        merged = RateCardSlotCreate(
            name=payload.name or slot.name,
            shift_log_type=payload.shift_log_type or slot.shift_log_type,
            day_type=payload.day_type or slot.day_type,
            start_time=payload.start_time if payload.start_time is not None else slot.start_time,
            end_time=payload.end_time if payload.end_time is not None else slot.end_time,
            base_amount=payload.base_amount if payload.base_amount is not None else slot.base_amount,
            base_km=payload.base_km if payload.base_km is not None else slot.base_km,
            base_hours=payload.base_hours if payload.base_hours is not None else slot.base_hours,
            extra_km_rate=payload.extra_km_rate if payload.extra_km_rate is not None else slot.extra_km_rate,
            extra_hour_rate=payload.extra_hour_rate if payload.extra_hour_rate is not None else slot.extra_hour_rate,
            waiting_rate_per_hour=payload.waiting_rate_per_hour if payload.waiting_rate_per_hour is not None else slot.waiting_rate_per_hour,
            escort_rate=payload.escort_rate if payload.escort_rate is not None else slot.escort_rate,
            night_allowance=payload.night_allowance if payload.night_allowance is not None else slot.night_allowance,
            tax_percent=payload.tax_percent if payload.tax_percent is not None else slot.tax_percent,
            priority=payload.priority if payload.priority is not None else slot.priority,
            is_active=payload.is_active if payload.is_active is not None else slot.is_active,
        )
        if merged.is_active:
            ensure_slot_no_overlap(db, rate_card_id, merged, exclude_slot_id=slot_id)
        for key, value in payload.model_dump(exclude_unset=True, exclude={"distance_slabs"}).items():
            setattr(slot, key, _value(value))
        if payload.distance_slabs is not None:
            db.query(RateCardDistanceSlab).filter(RateCardDistanceSlab.slot_id == slot_id).delete(synchronize_session=False)
            db.flush()
            create_distance_slabs(db, slot.slot_id, payload.distance_slabs)
        db.commit()
        db.refresh(slot)
        return ResponseWrapper.success({"slot": RateCardSlotResponse.model_validate(slot)}, "Rate card slot updated successfully")
    except HTTPException:
        raise
    except SQLAlchemyError as exc:
        db.rollback()
        raise handle_db_error(exc)


@router.post("/costing/garage-configs", status_code=status.HTTP_201_CREATED)
def create_garage_config(
    payload: GarageConfigCreate,
    db: Session = Depends(get_db),
    user_data=Depends(PermissionChecker(["costing_rate_card.update"], check_tenant=True)),
):
    try:
        tenant_id = resolve_tenant(user_data, payload.tenant_id)
        validate_vendor(db, tenant_id, payload.vendor_id)
        validate_vehicle(db, tenant_id, payload.vehicle_id)
        config = GarageConfig(tenant_id=tenant_id, **{key: _value(value) for key, value in payload.model_dump(exclude={"tenant_id"}).items()})
        db.add(config)
        db.commit()
        db.refresh(config)
        return ResponseWrapper.success({"garage_config": GarageConfigResponse.model_validate(config)}, "Garage config created successfully")
    except HTTPException:
        raise
    except SQLAlchemyError as exc:
        db.rollback()
        raise handle_db_error(exc)


@router.get("/costing/garage-configs", status_code=status.HTTP_200_OK)
def list_garage_configs(
    tenant_id: Optional[str] = Query(None),
    vendor_id: Optional[int] = Query(None),
    db: Session = Depends(get_db),
    user_data=Depends(PermissionChecker(["costing_rate_card.read"], check_tenant=True)),
):
    tenant_id = resolve_tenant(user_data, tenant_id)
    query = db.query(GarageConfig).filter(GarageConfig.tenant_id == tenant_id)
    if vendor_id:
        query = query.filter(GarageConfig.vendor_id == vendor_id)
    configs = query.order_by(GarageConfig.garage_config_id.desc()).all()
    return ResponseWrapper.success({"garage_configs": [GarageConfigResponse.model_validate(config) for config in configs], "total": len(configs)}, "Garage configs fetched successfully")


@router.patch("/costing/garage-configs/{garage_config_id}", status_code=status.HTTP_200_OK)
def update_garage_config(
    garage_config_id: int,
    payload: GarageConfigUpdate,
    tenant_id: Optional[str] = Query(None),
    db: Session = Depends(get_db),
    user_data=Depends(PermissionChecker(["costing_rate_card.update"], check_tenant=True)),
):
    try:
        tenant_id = resolve_tenant(user_data, tenant_id)
        config = db.query(GarageConfig).filter(GarageConfig.garage_config_id == garage_config_id, GarageConfig.tenant_id == tenant_id).first()
        if not config:
            raise HTTPException(status_code=404, detail=ResponseWrapper.error("Garage config not found", "GARAGE_CONFIG_NOT_FOUND"))
        updates = payload.model_dump(exclude_unset=True)
        if "vendor_id" in updates:
            validate_vendor(db, tenant_id, updates["vendor_id"])
        if "vehicle_id" in updates:
            validate_vehicle(db, tenant_id, updates["vehicle_id"])
        for key, value in updates.items():
            setattr(config, key, _value(value))
        db.commit()
        db.refresh(config)
        return ResponseWrapper.success({"garage_config": GarageConfigResponse.model_validate(config)}, "Garage config updated successfully")
    except HTTPException:
        raise
    except SQLAlchemyError as exc:
        db.rollback()
        raise handle_db_error(exc)


@router.post("/routes/{route_id}/costing/calculate", status_code=status.HTTP_200_OK)
def calculate_route_cost(
    route_id: int,
    payload: RouteCostCalculateRequest,
    tenant_id: Optional[str] = Query(None),
    db: Session = Depends(get_db),
    user_data=Depends(PermissionChecker(["route_cost.calculate"], check_tenant=True)),
):
    try:
        tenant_id = resolve_tenant(user_data, tenant_id)
        result = CostingService.calculate_route_cost(
            db,
            route_id,
            tenant_id,
            dry_run=payload.dry_run,
            distance_source=payload.distance_source,
            allocation_basis=payload.allocation_basis,
            manual_trip_km=payload.manual_trip_km,
            manual_trip_hours=payload.manual_trip_hours,
            comment=payload.comment,
        )
        if payload.dry_run:
            db.rollback()
        else:
            db.commit()
        return ResponseWrapper.success({"route_cost": result}, "Route cost calculated successfully")
    except HTTPException:
        db.rollback()
        raise
    except SQLAlchemyError as exc:
        db.rollback()
        raise handle_db_error(exc)


@router.get("/routes/{route_id}/costing", status_code=status.HTTP_200_OK)
def get_route_cost(
    route_id: int,
    tenant_id: Optional[str] = Query(None),
    db: Session = Depends(get_db),
    user_data=Depends(PermissionChecker(["route_cost.read"], check_tenant=True)),
):
    tenant_id = resolve_tenant(user_data, tenant_id, allow_vendor=True)
    route_for_user(db, route_id, tenant_id, user_data, allow_vendor=True)
    route_cost = db.query(RouteCost).filter(RouteCost.route_id == route_id, RouteCost.tenant_id == tenant_id).first()
    if not route_cost:
        raise HTTPException(status_code=404, detail=ResponseWrapper.error("Route cost not found", "ROUTE_COST_NOT_FOUND"))
    return ResponseWrapper.success({"route_cost": CostingService.route_cost_to_dict(route_cost)}, "Route cost fetched successfully")


@router.get("/routes/{route_id}/costing/bookings", status_code=status.HTTP_200_OK)
def get_route_booking_costs(
    route_id: int,
    tenant_id: Optional[str] = Query(None),
    db: Session = Depends(get_db),
    user_data=Depends(PermissionChecker(["route_cost.read"], check_tenant=True)),
):
    tenant_id = resolve_tenant(user_data, tenant_id, allow_vendor=True)
    route_for_user(db, route_id, tenant_id, user_data, allow_vendor=True)
    rows = (
        db.query(RouteBookingCost)
        .join(RouteCost, RouteCost.route_cost_id == RouteBookingCost.route_cost_id)
        .filter(RouteBookingCost.route_id == route_id, RouteBookingCost.tenant_id == tenant_id)
        .order_by(RouteBookingCost.booking_id.asc())
        .all()
    )
    return ResponseWrapper.success(
        {"booking_costs": [CostingService.booking_cost_to_dict(item) for item in rows], "total": len(rows)},
        "Route booking costs fetched successfully",
    )


def route_cost_action(db: Session, route_id: int, tenant_id: str, target_status: str, require_statuses: Optional[List[str]] = None) -> RouteCost:
    route_cost = db.query(RouteCost).filter(RouteCost.route_id == route_id, RouteCost.tenant_id == tenant_id).first()
    if not route_cost:
        raise HTTPException(status_code=404, detail=ResponseWrapper.error("Route cost not found", "ROUTE_COST_NOT_FOUND"))
    if route_cost.status == "finalized" and target_status != "finalized":
        raise HTTPException(status_code=409, detail=ResponseWrapper.error("Finalized route cost is locked", "ROUTE_COST_FINALIZED"))
    if require_statuses and route_cost.status not in require_statuses:
        raise HTTPException(status_code=409, detail=ResponseWrapper.error("Route cost status transition not allowed", "INVALID_ROUTE_COST_STATUS"))
    route_cost.status = target_status
    now = datetime.utcnow()
    if target_status == "approved":
        route_cost.approved_at = now
    if target_status == "finalized":
        route_cost.finalized_at = now
    return route_cost


@router.post("/routes/{route_id}/costing/submit", status_code=status.HTTP_200_OK)
def submit_route_cost(route_id: int, payload: RouteCostActionRequest, tenant_id: Optional[str] = Query(None), db: Session = Depends(get_db), user_data=Depends(PermissionChecker(["route_cost.submit"], check_tenant=True))):
    tenant_id = resolve_tenant(user_data, tenant_id)
    route_cost = route_cost_action(db, route_id, tenant_id, "submitted", ["draft", "rejected"])
    db.commit()
    return ResponseWrapper.success({"route_cost": CostingService.route_cost_to_dict(route_cost)}, "Route cost submitted successfully")


@router.post("/routes/{route_id}/costing/approve", status_code=status.HTTP_200_OK)
def approve_route_cost(route_id: int, payload: RouteCostActionRequest, tenant_id: Optional[str] = Query(None), db: Session = Depends(get_db), user_data=Depends(PermissionChecker(["route_cost.approve"], check_tenant=True))):
    tenant_id = resolve_tenant(user_data, tenant_id)
    route_cost = route_cost_action(db, route_id, tenant_id, "approved", ["draft", "submitted"])
    db.commit()
    return ResponseWrapper.success({"route_cost": CostingService.route_cost_to_dict(route_cost)}, "Route cost approved successfully")


@router.post("/routes/{route_id}/costing/reject", status_code=status.HTTP_200_OK)
def reject_route_cost(route_id: int, payload: RouteCostActionRequest, tenant_id: Optional[str] = Query(None), db: Session = Depends(get_db), user_data=Depends(PermissionChecker(["route_cost.approve"], check_tenant=True))):
    tenant_id = resolve_tenant(user_data, tenant_id)
    route_cost = route_cost_action(db, route_id, tenant_id, "rejected", ["draft", "submitted", "approved"])
    snapshot = route_cost.calculation_snapshot or {}
    snapshot["rejection_comment"] = payload.comment
    route_cost.calculation_snapshot = snapshot
    db.commit()
    return ResponseWrapper.success({"route_cost": CostingService.route_cost_to_dict(route_cost)}, "Route cost rejected successfully")


@router.post("/routes/{route_id}/costing/finalize", status_code=status.HTTP_200_OK)
def finalize_route_cost(route_id: int, payload: RouteCostActionRequest, tenant_id: Optional[str] = Query(None), db: Session = Depends(get_db), user_data=Depends(PermissionChecker(["route_cost.finalize"], check_tenant=True))):
    tenant_id = resolve_tenant(user_data, tenant_id)
    route_cost = route_cost_action(db, route_id, tenant_id, "finalized", ["approved"])
    db.commit()
    return ResponseWrapper.success({"route_cost": CostingService.route_cost_to_dict(route_cost)}, "Route cost finalized successfully")


@router.post("/routes/{route_id}/expenses", status_code=status.HTTP_201_CREATED)
def create_route_expense(route_id: int, payload: RouteExpenseCreate, tenant_id: Optional[str] = Query(None), db: Session = Depends(get_db), user_data=Depends(PermissionChecker(["route_expense.create"], check_tenant=True))):
    try:
        tenant_id = resolve_tenant(user_data, tenant_id, allow_vendor=True)
        route = route_for_user(db, route_id, tenant_id, user_data, allow_vendor=True)
        if not route.assigned_vendor_id:
            raise HTTPException(status_code=400, detail=ResponseWrapper.error("Route has no vendor assigned", "VENDOR_NOT_ASSIGNED"))
        expense = RouteExpense(
            route_id=route_id,
            tenant_id=tenant_id,
            vendor_id=route.assigned_vendor_id,
            expense_type=payload.expense_type.value,
            amount=payload.amount,
            comment=payload.comment,
            attachment_url=payload.attachment_url,
            status="draft",
            created_by_type=user_data.get("user_type"),
            created_by_id=str(user_data.get("user_id")),
        )
        db.add(expense)
        db.commit()
        db.refresh(expense)
        return ResponseWrapper.success({"expense": RouteExpenseResponse.model_validate(expense)}, "Expense created successfully")
    except HTTPException:
        raise
    except SQLAlchemyError as exc:
        db.rollback()
        raise handle_db_error(exc)


@router.get("/routes/{route_id}/expenses", status_code=status.HTTP_200_OK)
def list_route_expenses(route_id: int, tenant_id: Optional[str] = Query(None), db: Session = Depends(get_db), user_data=Depends(PermissionChecker(["route_expense.read"], check_tenant=True))):
    tenant_id = resolve_tenant(user_data, tenant_id, allow_vendor=True)
    route_for_user(db, route_id, tenant_id, user_data, allow_vendor=True)
    expenses = db.query(RouteExpense).filter(RouteExpense.route_id == route_id, RouteExpense.tenant_id == tenant_id).order_by(RouteExpense.created_at.desc()).all()
    return ResponseWrapper.success({"expenses": [RouteExpenseResponse.model_validate(item) for item in expenses], "total": len(expenses)}, "Expenses fetched successfully")


@router.patch("/routes/{route_id}/expenses/{expense_id}", status_code=status.HTTP_200_OK)
def update_route_expense(route_id: int, expense_id: int, payload: RouteExpenseUpdate, tenant_id: Optional[str] = Query(None), db: Session = Depends(get_db), user_data=Depends(PermissionChecker(["route_expense.update"], check_tenant=True))):
    tenant_id = resolve_tenant(user_data, tenant_id, allow_vendor=True)
    route_for_user(db, route_id, tenant_id, user_data, allow_vendor=True)
    expense = db.query(RouteExpense).filter(RouteExpense.expense_id == expense_id, RouteExpense.route_id == route_id, RouteExpense.tenant_id == tenant_id).first()
    if not expense:
        raise HTTPException(status_code=404, detail=ResponseWrapper.error("Expense not found", "EXPENSE_NOT_FOUND"))
    if expense.status != "draft":
        raise HTTPException(status_code=409, detail=ResponseWrapper.error("Only draft expenses can be edited", "EXPENSE_NOT_EDITABLE"))
    for key, value in payload.model_dump(exclude_unset=True).items():
        setattr(expense, key, _value(value))
    db.commit()
    db.refresh(expense)
    return ResponseWrapper.success({"expense": RouteExpenseResponse.model_validate(expense)}, "Expense updated successfully")


@router.post("/routes/{route_id}/expenses/submit", status_code=status.HTTP_200_OK)
def submit_route_expenses(route_id: int, tenant_id: Optional[str] = Query(None), db: Session = Depends(get_db), user_data=Depends(PermissionChecker(["route_expense.submit"], check_tenant=True))):
    tenant_id = resolve_tenant(user_data, tenant_id, allow_vendor=True)
    route_for_user(db, route_id, tenant_id, user_data, allow_vendor=True)
    updated = db.query(RouteExpense).filter(RouteExpense.route_id == route_id, RouteExpense.tenant_id == tenant_id, RouteExpense.status == "draft").update({RouteExpense.status: "pending_approval"})
    db.commit()
    return ResponseWrapper.success({"submitted_count": updated}, "Expenses submitted successfully")


@router.post("/routes/{route_id}/expenses/{expense_id}/approve", status_code=status.HTTP_200_OK)
def approve_route_expense(route_id: int, expense_id: int, tenant_id: Optional[str] = Query(None), db: Session = Depends(get_db), user_data=Depends(PermissionChecker(["route_expense.approve"], check_tenant=True))):
    tenant_id = resolve_tenant(user_data, tenant_id)
    route_for_user(db, route_id, tenant_id, user_data)
    expense = db.query(RouteExpense).filter(RouteExpense.expense_id == expense_id, RouteExpense.route_id == route_id, RouteExpense.tenant_id == tenant_id).first()
    if not expense:
        raise HTTPException(status_code=404, detail=ResponseWrapper.error("Expense not found", "EXPENSE_NOT_FOUND"))
    if expense.status not in {"pending_approval", "draft"}:
        raise HTTPException(status_code=409, detail=ResponseWrapper.error("Expense cannot be approved from current status", "INVALID_EXPENSE_STATUS"))
    expense.status = "approved"
    expense.approved_by_type = user_data.get("user_type")
    expense.approved_by_id = str(user_data.get("user_id"))
    db.commit()
    db.refresh(expense)
    return ResponseWrapper.success({"expense": RouteExpenseResponse.model_validate(expense)}, "Expense approved successfully")


@router.post("/routes/{route_id}/expenses/{expense_id}/reject", status_code=status.HTTP_200_OK)
def reject_route_expense(route_id: int, expense_id: int, payload: RouteExpenseRejectRequest, tenant_id: Optional[str] = Query(None), db: Session = Depends(get_db), user_data=Depends(PermissionChecker(["route_expense.approve"], check_tenant=True))):
    tenant_id = resolve_tenant(user_data, tenant_id)
    route_for_user(db, route_id, tenant_id, user_data)
    expense = db.query(RouteExpense).filter(RouteExpense.expense_id == expense_id, RouteExpense.route_id == route_id, RouteExpense.tenant_id == tenant_id).first()
    if not expense:
        raise HTTPException(status_code=404, detail=ResponseWrapper.error("Expense not found", "EXPENSE_NOT_FOUND"))
    expense.status = "rejected"
    expense.rejection_reason = payload.reason
    expense.approved_by_type = user_data.get("user_type")
    expense.approved_by_id = str(user_data.get("user_id"))
    db.commit()
    db.refresh(expense)
    return ResponseWrapper.success({"expense": RouteExpenseResponse.model_validate(expense)}, "Expense rejected successfully")


def route_cost_report_query(db: Session, tenant_id: str, start_date: date, end_date: date):
    start_dt = datetime.combine(start_date, time.min)
    end_dt = datetime.combine(end_date + timedelta(days=1), time.min)
    return (
        db.query(RouteCost, RouteManagement, Vendor, Vehicle, VehicleType)
        .join(RouteManagement, RouteManagement.route_id == RouteCost.route_id)
        .outerjoin(Vendor, Vendor.vendor_id == RouteCost.vendor_id)
        .outerjoin(Vehicle, Vehicle.vehicle_id == RouteCost.vehicle_id)
        .outerjoin(VehicleType, VehicleType.vehicle_type_id == RouteCost.vehicle_type_id)
        .filter(RouteCost.tenant_id == tenant_id, RouteCost.calculated_at >= start_dt, RouteCost.calculated_at < end_dt)
    )


def booking_cost_report_query(db: Session, tenant_id: str, start_date: date, end_date: date):
    start_dt = datetime.combine(start_date, time.min)
    end_dt = datetime.combine(end_date + timedelta(days=1), time.min)
    return (
        db.query(RouteBookingCost, RouteCost, RouteManagement, Booking, CostCenter)
        .join(RouteCost, RouteCost.route_cost_id == RouteBookingCost.route_cost_id)
        .join(RouteManagement, RouteManagement.route_id == RouteBookingCost.route_id)
        .join(Booking, Booking.booking_id == RouteBookingCost.booking_id)
        .join(CostCenter, CostCenter.cost_center_id == RouteBookingCost.cost_center_id)
        .filter(RouteBookingCost.tenant_id == tenant_id, RouteCost.calculated_at >= start_dt, RouteCost.calculated_at < end_dt)
    )


def report_row(route_cost: RouteCost, route: RouteManagement, vendor: Optional[Vendor], vehicle: Optional[Vehicle], vehicle_type: Optional[VehicleType]) -> Dict[str, Any]:
    return {
        "route_cost_id": route_cost.route_cost_id,
        "route_id": route_cost.route_id,
        "route_code": route.route_code,
        "cost_status": route_cost.status,
        "route_status": route.status.value if hasattr(route.status, "value") else route.status,
        "vendor_id": route_cost.vendor_id,
        "vendor_name": vendor.name if vendor else None,
        "vehicle_id": route_cost.vehicle_id,
        "vehicle_number": vehicle.rc_number if vehicle else None,
        "vehicle_type": vehicle_type.name if vehicle_type else None,
        "distance_source": route_cost.distance_source,
        "trip_km": route_cost.trip_km,
        "trip_hours": route_cost.trip_hours,
        "garage_km": route_cost.garage_km,
        "base_amount": route_cost.base_amount,
        "extra_km_amount": route_cost.extra_km_amount,
        "extra_hour_amount": route_cost.extra_hour_amount,
        "garage_amount": route_cost.garage_amount,
        "expense_amount": route_cost.expense_amount,
        "tax_amount": route_cost.tax_amount,
        "total_amount": route_cost.total_amount,
        "variance_percent": route_cost.variance_percent,
        "calculated_at": route_cost.calculated_at,
    }


def booking_report_row(booking_cost: RouteBookingCost, route_cost: RouteCost, route: RouteManagement, booking: Booking, cost_center: CostCenter) -> Dict[str, Any]:
    return {
        "route_booking_cost_id": booking_cost.route_booking_cost_id,
        "route_cost_id": route_cost.route_cost_id,
        "route_id": booking_cost.route_id,
        "route_code": route.route_code,
        "booking_id": booking_cost.booking_id,
        "employee_id": booking.employee_id,
        "employee_code": booking.employee_code,
        "booking_status": booking.status.value if hasattr(booking.status, "value") else booking.status,
        "cost_center_id": cost_center.cost_center_id,
        "cost_center_code": cost_center.code,
        "cost_center_name": cost_center.name,
        "distance_source": booking_cost.distance_source,
        "route_total_km": booking_cost.route_total_km,
        "route_total_hours": booking_cost.route_total_hours,
        "booking_planned_km": booking_cost.booking_planned_km,
        "booking_actual_km": booking_cost.booking_actual_km,
        "allocation_percent": booking_cost.allocation_percent,
        "allocated_amount": booking_cost.allocated_amount,
        "route_total_amount": route_cost.total_amount,
        "calculated_at": route_cost.calculated_at,
    }


@router.get("/reports/route-costs", status_code=status.HTTP_200_OK)
def route_costs_report(start_date: date, end_date: date, tenant_id: Optional[str] = Query(None), vendor_id: Optional[int] = Query(None), cost_status: Optional[str] = Query(None), cost_center_id: Optional[int] = Query(None), db: Session = Depends(get_db), user_data=Depends(PermissionChecker(["route_cost.report.read"], check_tenant=True))):
    tenant_id = resolve_tenant(user_data, tenant_id, allow_vendor=True)
    query = route_cost_report_query(db, tenant_id, start_date, end_date)
    if user_data.get("user_type") == "vendor":
        query = query.filter(RouteCost.vendor_id == user_data.get("vendor_id"))
    if vendor_id:
        query = query.filter(RouteCost.vendor_id == vendor_id)
    if cost_status:
        query = query.filter(RouteCost.status == cost_status)
    if cost_center_id:
        query = query.join(RouteCostAllocation, RouteCostAllocation.route_cost_id == RouteCost.route_cost_id).filter(RouteCostAllocation.cost_center_id == cost_center_id)
    rows = [report_row(*item) for item in query.order_by(RouteCost.calculated_at.desc()).all()]
    return ResponseWrapper.success({"route_costs": rows, "total": len(rows)}, "Route cost report generated successfully")


@router.get("/reports/cost-centers/summary", status_code=status.HTTP_200_OK)
def cost_center_summary(start_date: date, end_date: date, tenant_id: Optional[str] = Query(None), db: Session = Depends(get_db), user_data=Depends(PermissionChecker(["route_cost.report.read"], check_tenant=True))):
    tenant_id = resolve_tenant(user_data, tenant_id)
    start_dt = datetime.combine(start_date, time.min)
    end_dt = datetime.combine(end_date + timedelta(days=1), time.min)
    rows = (
        db.query(
            CostCenter.cost_center_id,
            CostCenter.code,
            CostCenter.name,
            func.coalesce(func.sum(RouteCostAllocation.allocated_amount), 0).label("total_amount"),
            func.count(RouteCostAllocation.allocation_id).label("allocation_count"),
        )
        .join(RouteCostAllocation, RouteCostAllocation.cost_center_id == CostCenter.cost_center_id)
        .join(RouteCost, RouteCost.route_cost_id == RouteCostAllocation.route_cost_id)
        .filter(RouteCost.tenant_id == tenant_id, RouteCost.calculated_at >= start_dt, RouteCost.calculated_at < end_dt)
        .group_by(CostCenter.cost_center_id, CostCenter.code, CostCenter.name)
        .order_by(CostCenter.code.asc())
        .all()
    )
    data = [
        {"cost_center_id": row.cost_center_id, "code": row.code, "name": row.name, "total_amount": row.total_amount, "allocation_count": row.allocation_count}
        for row in rows
    ]
    return ResponseWrapper.success({"cost_centers": data, "total": len(data)}, "Cost center summary generated successfully")


@router.get("/reports/booking-costs", status_code=status.HTTP_200_OK)
def booking_costs_report(
    start_date: date,
    end_date: date,
    tenant_id: Optional[str] = Query(None),
    route_id: Optional[int] = Query(None),
    booking_id: Optional[int] = Query(None),
    cost_center_id: Optional[int] = Query(None),
    db: Session = Depends(get_db),
    user_data=Depends(PermissionChecker(["route_cost.report.read"], check_tenant=True)),
):
    tenant_id = resolve_tenant(user_data, tenant_id, allow_vendor=True)
    query = booking_cost_report_query(db, tenant_id, start_date, end_date)
    if user_data.get("user_type") == "vendor":
        query = query.filter(RouteCost.vendor_id == user_data.get("vendor_id"))
    if route_id:
        query = query.filter(RouteBookingCost.route_id == route_id)
    if booking_id:
        query = query.filter(RouteBookingCost.booking_id == booking_id)
    if cost_center_id:
        query = query.filter(RouteBookingCost.cost_center_id == cost_center_id)
    rows = [booking_report_row(*item) for item in query.order_by(RouteCost.calculated_at.desc(), RouteBookingCost.booking_id.asc()).all()]
    return ResponseWrapper.success({"booking_costs": rows, "total": len(rows)}, "Booking cost report generated successfully")


@router.get("/reports/route-costs/export", status_code=status.HTTP_200_OK)
def route_costs_export(start_date: date, end_date: date, tenant_id: Optional[str] = Query(None), db: Session = Depends(get_db), user_data=Depends(PermissionChecker(["route_cost.report.read"], check_tenant=True))):
    tenant_id = resolve_tenant(user_data, tenant_id, allow_vendor=True)
    query = route_cost_report_query(db, tenant_id, start_date, end_date)
    if user_data.get("user_type") == "vendor":
        query = query.filter(RouteCost.vendor_id == user_data.get("vendor_id"))
    rows = [report_row(*item) for item in query.order_by(RouteCost.calculated_at.desc()).all()]
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Route Costs"
    headers = [
        "Route Cost ID", "Route ID", "Route Code", "Cost Status", "Route Status", "Vendor", "Vehicle", "Vehicle Type",
        "Distance Source", "Trip KM", "Trip Hours", "Garage KM", "Base Amount", "Extra KM Amount", "Extra Hour Amount",
        "Garage Amount", "Expense Amount", "Tax Amount", "Total Amount", "Variance %", "Calculated At",
    ]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    for row in rows:
        ws.append([
            row["route_cost_id"], row["route_id"], row["route_code"], row["cost_status"], row["route_status"], row["vendor_name"],
            row["vehicle_number"], row["vehicle_type"], row["distance_source"], row["trip_km"], row["trip_hours"], row["garage_km"],
            row["base_amount"], row["extra_km_amount"], row["extra_hour_amount"], row["garage_amount"], row["expense_amount"],
            row["tax_amount"], row["total_amount"], row["variance_percent"], row["calculated_at"],
        ])
    stream = BytesIO()
    wb.save(stream)
    stream.seek(0)
    filename = f"route_costs_{start_date}_{end_date}.xlsx"
    return StreamingResponse(
        stream,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )
