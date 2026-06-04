from fastapi import APIRouter, Depends, HTTPException, Query, status as http_status
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from sqlalchemy import and_, case, or_, func
from typing import Optional, List
from datetime import date, datetime, time, timedelta
from io import BytesIO
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from app.database.session import get_db
from app.models.booking import Booking, BookingStatusEnum
from app.models.route_management import RouteManagement, RouteManagementBooking, RouteManagementStatusEnum
from app.models.route_delay_event import RouteDelayEvent
from app.models.driver import Driver
from app.models.vehicle import Vehicle
from app.models.vendor import Vendor
from app.models.employee import Employee
from app.models.shift import Shift
from app.models.tenant import Tenant
from app.models.escort import Escort
from app.models.vehicle_type import VehicleType
from app.utils.response_utils import ResponseWrapper, handle_db_error
from common_utils.auth.permission_checker import PermissionChecker
from common_utils import get_current_ist_time
from app.core.logging_config import get_logger
from app.utils.cache_manager import cached

logger = get_logger(__name__)

router = APIRouter(prefix="/reports", tags=["reports"])


def validate_date_range(start_date: date, end_date: date):
    """Validate that date range is valid and not too large"""
    if start_date > end_date:
        raise HTTPException(
            status_code=http_status.HTTP_400_BAD_REQUEST,
            detail=ResponseWrapper.error(
                message="start_date cannot be after end_date",
                error_code="INVALID_DATE_RANGE"
            )
        )
    
    # Limit to 90 days to prevent performance issues
    date_diff = (end_date - start_date).days
    if date_diff > 90:
        raise HTTPException(
            status_code=http_status.HTTP_400_BAD_REQUEST,
            detail=ResponseWrapper.error(
                message="Date range cannot exceed 90 days",
                error_code="DATE_RANGE_TOO_LARGE"
            )
        )


def style_excel_report(ws, headers):
    """Apply professional styling to Excel worksheet"""
    # Header styling
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Style header row
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = border
        
        # Auto-adjust column width
        ws.column_dimensions[get_column_letter(col_num)].width = max(len(str(header)) + 5, 15)
    
    # Freeze header row
    ws.freeze_panes = 'A2'
    
    return ws


@router.get("/bookings", status_code=http_status.HTTP_200_OK)
async def list_bookings_report(
    start_date: date = Query(..., description="Start date (YYYY-MM-DD)"),
    end_date: date = Query(..., description="End date (YYYY-MM-DD)"),
    tenant_id: Optional[str] = Query(None, description="Tenant ID (required for admin users)"),
    shift_id: Optional[int] = Query(None, description="Filter by shift ID"),
    booking_status: Optional[List[BookingStatusEnum]] = Query(None, description="Filter by booking status (multi-select)"),
    route_status: Optional[List[RouteManagementStatusEnum]] = Query(None, description="Filter by route status (multi-select)"),
    vendor_id: Optional[int] = Query(None, description="Filter by vendor ID"),
    include_unrouted: Optional[bool] = Query(True, description="Include bookings without routes"),
    page: int = Query(1, ge=1, description="Page number (1-indexed)"),
    page_size: int = Query(100, ge=1, le=500, description="Records per page (max 500)"),
    db: Session = Depends(get_db),
    user_data=Depends(PermissionChecker(["report.read"], check_tenant=True)),
):
    """
    Return bookings report data as JSON for frontend table display.

    Identical filters to GET /bookings/export.  Supports pagination so the
    frontend can render large date ranges in pages without waiting for a full
    Excel build.

    Use GET /bookings/export with the same filters to download the Excel file.

    Response shape
    --------------
    {
      "data": {
        "bookings": [ { ...one row per booking... }, ... ],
        "pagination": { "total": 1234, "page": 1, "page_size": 100, "total_pages": 13 },
        "summary":    { "total_bookings": ..., "routed_bookings": ...,
                        "unrouted_bookings": ..., "status_breakdown": {...} },
        "meta":       { "start_date": ..., "end_date": ..., "generated_at": ... }
      }
    }
    """
    try:
        user_type       = user_data.get("user_type")
        token_tenant_id = user_data.get("tenant_id")
        token_vendor_id = user_data.get("vendor_id")
        user_id         = user_data.get("user_id")

        logger.info(
            "[list_bookings_report] user_id=%s user_type=%s date_range=%s to %s page=%s",
            user_id, user_type, start_date, end_date, page,
        )

        # --- Tenant / vendor resolution (same rules as export) ---
        if user_type == "employee":
            tenant_id = token_tenant_id
        elif user_type == "vendor":
            tenant_id = token_tenant_id
            vendor_id = token_vendor_id
        elif user_type == "admin":
            if not tenant_id:
                raise HTTPException(
                    status_code=http_status.HTTP_400_BAD_REQUEST,
                    detail=ResponseWrapper.error(
                        message="tenant_id is required for admin users",
                        error_code="TENANT_ID_REQUIRED",
                    ),
                )
        else:
            raise HTTPException(
                status_code=http_status.HTTP_403_FORBIDDEN,
                detail=ResponseWrapper.error(
                    message="Insufficient permissions to view reports",
                    error_code="FORBIDDEN",
                ),
            )

        if not tenant_id:
            raise HTTPException(
                status_code=http_status.HTTP_403_FORBIDDEN,
                detail=ResponseWrapper.error(
                    message="Tenant context not available",
                    error_code="TENANT_ID_REQUIRED",
                ),
            )

        validate_date_range(start_date, end_date)

        from app.utils.cache_manager import get_tenant_with_cache, get_shift_with_cache
        tenant = get_tenant_with_cache(db, tenant_id)
        if not tenant:
            raise HTTPException(
                status_code=http_status.HTTP_404_NOT_FOUND,
                detail=ResponseWrapper.error(
                    message=f"Tenant {tenant_id} not found",
                    error_code="TENANT_NOT_FOUND",
                ),
            )

        if shift_id:
            shift = get_shift_with_cache(db, tenant_id, shift_id)
            if not shift:
                raise HTTPException(
                    status_code=http_status.HTTP_404_NOT_FOUND,
                    detail=ResponseWrapper.error(
                        message=f"Shift {shift_id} not found for this tenant",
                        error_code="SHIFT_NOT_FOUND",
                    ),
                )

        # --- Build query (mirrors export_bookings_report exactly) ---
        query = (
            db.query(
                Booking.booking_id,
                Booking.booking_date,
                Booking.status.label("booking_status"),
                Booking.pickup_location,
                Booking.drop_location,
                Booking.pickup_latitude,
                Booking.pickup_longitude,
                Booking.drop_latitude,
                Booking.drop_longitude,
                Booking.reason,
                Booking.employee_id,
                Booking.employee_code,
                Employee.name.label("employee_name"),
                Employee.phone.label("employee_phone"),
                Employee.gender.label("employee_gender"),
                Shift.shift_id,
                Shift.shift_code,
                Shift.shift_time,
                Shift.log_type.label("shift_type"),
                RouteManagement.route_id,
                RouteManagement.route_code,
                RouteManagement.status.label("route_status"),
                RouteManagementBooking.order_id,
                RouteManagementBooking.estimated_pick_up_time,
                RouteManagementBooking.estimated_drop_time,
                RouteManagementBooking.actual_pick_up_time,
                RouteManagementBooking.actual_drop_time,
                RouteManagementBooking.estimated_distance,
                RouteManagement.actual_total_distance,
                Driver.driver_id,
                Driver.name.label("driver_name"),
                Driver.phone.label("driver_phone"),
                Driver.license_number,
                Vehicle.vehicle_id,
                Vehicle.rc_number,
                Vendor.vendor_id,
                Vendor.name.label("vendor_name"),
                Vendor.phone.label("vendor_phone"),
            )
            .outerjoin(Employee, Booking.employee_id == Employee.employee_id)
            .join(Shift, Booking.shift_id == Shift.shift_id)
            .outerjoin(RouteManagementBooking, Booking.booking_id == RouteManagementBooking.booking_id)
            .outerjoin(RouteManagement, RouteManagementBooking.route_id == RouteManagement.route_id)
            .outerjoin(Driver, RouteManagement.assigned_driver_id == Driver.driver_id)
            .outerjoin(Vehicle, RouteManagement.assigned_vehicle_id == Vehicle.vehicle_id)
            .outerjoin(Vendor, RouteManagement.assigned_vendor_id == Vendor.vendor_id)
            .filter(
                Booking.tenant_id == tenant_id,
                Booking.booking_date >= start_date,
                Booking.booking_date <= end_date,
            )
        )

        if shift_id:
            query = query.filter(Booking.shift_id == shift_id)
        if booking_status:
            query = query.filter(Booking.status.in_(booking_status))
        if route_status:
            query = query.filter(RouteManagement.status.in_(route_status))
        if vendor_id:
            query = query.filter(RouteManagement.assigned_vendor_id == vendor_id)
        if not include_unrouted:
            query = query.filter(RouteManagement.route_id.isnot(None))

        query = query.order_by(
            Booking.booking_date.desc(),
            Shift.shift_id,
            RouteManagement.route_id,
            RouteManagementBooking.order_id,
        )

        # --- Pagination ---
        total = query.count()
        offset = (page - 1) * page_size
        results = query.offset(offset).limit(page_size).all()

        # --- Serialise rows ---
        def _fmt_time(t):
            """Convert DB time value to HH:MM format."""

            if t is None:
                return None

            # Your RouteManagementBooking fields are String(10)
            if isinstance(t, str):
                return t[:5] if len(t) >= 5 else t

            # datetime.time
            if isinstance(t, time):
                return t.strftime("%H:%M")

            # timedelta
            if isinstance(t, timedelta):
                total_secs = int(t.total_seconds())
                return f"{total_secs // 3600:02d}:{(total_secs % 3600) // 60:02d}"

            # fallback
            return str(t)

        bookings = []
        for r in results:
            bookings.append({
                "booking_id":               r.booking_id,
                "booking_date":             r.booking_date.strftime("%Y-%m-%d") if r.booking_date else None,
                "booking_status":           r.booking_status.value if r.booking_status else None,
                "employee_id":              r.employee_id,
                "employee_code":            r.employee_code,
                "employee_name":            r.employee_name,
                "employee_phone":           r.employee_phone,
                "employee_gender":          r.employee_gender.value if r.employee_gender and hasattr(r.employee_gender, "value") else str(r.employee_gender) if r.employee_gender else None,
                "shift_id":                 r.shift_id,
                "shift_code":               r.shift_code,
                "shift_time":               _fmt_time(r.shift_time),
                "shift_type":               r.shift_type.value if r.shift_type and hasattr(r.shift_type, "value") else str(r.shift_type) if r.shift_type else None,
                "pickup_location":          r.pickup_location,
                "pickup_latitude":          r.pickup_latitude,
                "pickup_longitude":         r.pickup_longitude,
                "drop_location":            r.drop_location,
                "drop_latitude":            r.drop_latitude,
                "drop_longitude":           r.drop_longitude,
                "route_id":                 r.route_id,
                "route_code":               r.route_code,
                "route_status":             r.route_status.value if r.route_status else None,
                "stop_order":               r.order_id,
                "estimated_pickup_time":    _fmt_time(r.estimated_pick_up_time),
                "estimated_drop_time":      _fmt_time(r.estimated_drop_time),
                "actual_pickup_time":       _fmt_time(r.actual_pick_up_time),
                "actual_drop_time":         _fmt_time(r.actual_drop_time),
                "estimated_distance_km":    float(r.estimated_distance) if r.estimated_distance is not None else None,
                "actual_total_distance_km": float(r.actual_total_distance) if r.actual_total_distance is not None else None,
                "driver_id":                r.driver_id,
                "driver_name":              r.driver_name,
                "driver_phone":             r.driver_phone,
                "driver_license":           r.license_number,
                "vehicle_id":               r.vehicle_id,
                "vehicle_number":           r.rc_number,
                "vendor_id":                r.vendor_id,
                "vendor_name":              r.vendor_name,
                "vendor_phone":             r.vendor_phone,
                "reason":                   r.reason,
            })

        # --- Summary stats across the full date-range result set (not just this page) ---
        # We need counts over ALL matching rows, so fetch id+status+route_id only.
        summary_rows = (
            db.query(
                Booking.status.label("booking_status"),
                RouteManagement.route_id,
            )
            .outerjoin(RouteManagementBooking, Booking.booking_id == RouteManagementBooking.booking_id)
            .outerjoin(RouteManagement, RouteManagementBooking.route_id == RouteManagement.route_id)
            .join(Shift, Booking.shift_id == Shift.shift_id)
            .filter(
                Booking.tenant_id == tenant_id,
                Booking.booking_date >= start_date,
                Booking.booking_date <= end_date,
            )
            .all()
        )
        status_breakdown: dict = {}
        routed = 0
        for r in summary_rows:
            s = r.booking_status.value if r.booking_status else "UNKNOWN"
            status_breakdown[s] = status_breakdown.get(s, 0) + 1
            if r.route_id is not None:
                routed += 1

        import math
        return ResponseWrapper.success(
            data={
                "bookings": bookings,
                "pagination": {
                    "total":       total,
                    "page":        page,
                    "page_size":   page_size,
                    "total_pages": math.ceil(total / page_size) if total else 1,
                },
                "summary": {
                    "total_bookings":    total,
                    "routed_bookings":   routed,
                    "unrouted_bookings": total - routed,
                    "status_breakdown":  status_breakdown,
                },
                "meta": {
                    "start_date":    str(start_date),
                    "end_date":      str(end_date),
                    "tenant_id":     tenant_id,
                    "tenant_name":   tenant.name,
                    "generated_at":  get_current_ist_time().isoformat(),
                },
            }
        )

    except HTTPException:
        raise
    except Exception as e:
        logger.exception("[list_bookings_report] Error: %s", e)
        raise handle_db_error(e)


@router.get("/bookings/export", status_code=http_status.HTTP_200_OK)
async def export_bookings_report(
    start_date: date = Query(..., description="Start date for the report (YYYY-MM-DD)"),
    end_date: date = Query(..., description="End date for the report (YYYY-MM-DD)"),
    tenant_id: Optional[str] = Query(None, description="Tenant ID"),
    shift_id: Optional[int] = Query(None, description="Filter by shift ID"),
    booking_status: Optional[List[BookingStatusEnum]] = Query(None, description="Filter by booking status (can select multiple)"),
    route_status: Optional[List[RouteManagementStatusEnum]] = Query(None, description="Filter by route status (can select multiple)"),
    vendor_id: Optional[int] = Query(None, description="Filter by vendor ID"),
    include_unrouted: Optional[bool] = Query(True, description="Include bookings without routes"),
    db: Session = Depends(get_db),
    user_data=Depends(PermissionChecker(["report.read"], check_tenant=True)),
):
    """
    Generate comprehensive Excel report for bookings with route and assignment details.
    
    Filters:
    - Date range (required, max 90 days)
    - Shift ID
    - Booking status (REQUEST, SCHEDULED, ONGOING, COMPLETED, CANCELLED, NO_SHOW, EXPIRED)
    - Route status (PLANNED, VENDOR_ASSIGNED, DRIVER_ASSIGNED, ONGOING, COMPLETED, CANCELLED)
    - Vendor ID
    - Include/exclude unrouted bookings
    
    Returns: Excel file with detailed booking information
    """
    try:
        user_type = user_data.get("user_type")
        token_tenant_id = user_data.get("tenant_id")
        token_vendor_id = user_data.get("vendor_id")
        user_id = user_data.get("user_id")

        logger.info(
            f"[export_bookings_report] user_id={user_id}, user_type={user_type}, "
            f"date_range={start_date} to {end_date}, shift={shift_id}, status={booking_status}"
        )

        # --- Tenant Resolution ---
        if user_type == "employee":
            tenant_id = token_tenant_id
        elif user_type == "vendor":
            tenant_id = token_tenant_id
            vendor_id = token_vendor_id  # Force vendor to only see their data
        elif user_type == "admin":
            if not tenant_id:
                raise HTTPException(
                    status_code=http_status.HTTP_400_BAD_REQUEST,
                    detail=ResponseWrapper.error(
                        message="tenant_id is required for admin users",
                        error_code="TENANT_ID_REQUIRED"
                    )
                )
        else:
            raise HTTPException(
                status_code=http_status.HTTP_403_FORBIDDEN,
                detail=ResponseWrapper.error(
                    message="Insufficient permissions to generate reports",
                    error_code="FORBIDDEN"
                )
            )

        if not tenant_id:
            raise HTTPException(
                status_code=http_status.HTTP_403_FORBIDDEN,
                detail=ResponseWrapper.error(
                    message="Tenant context not available",
                    error_code="TENANT_ID_REQUIRED"
                )
            )

        # --- Validate Date Range ---
        validate_date_range(start_date, end_date)

        # --- Validate Tenant (use cache) ---
        from app.utils.cache_manager import get_tenant_with_cache, get_shift_with_cache
        tenant = get_tenant_with_cache(db, tenant_id)
        if not tenant:
            raise HTTPException(
                status_code=http_status.HTTP_404_NOT_FOUND,
                detail=ResponseWrapper.error(
                    message=f"Tenant {tenant_id} not found",
                    error_code="TENANT_NOT_FOUND"
                )
            )

        # --- Validate Shift if provided (use cache) ---
        if shift_id:
            shift = get_shift_with_cache(db, tenant_id, shift_id)
            if not shift:
                raise HTTPException(
                    status_code=http_status.HTTP_404_NOT_FOUND,
                    detail=ResponseWrapper.error(
                        message=f"Shift {shift_id} not found for this tenant",
                        error_code="SHIFT_NOT_FOUND"
                    )
                )

        # --- Build Query ---
        query = (
            db.query(
                Booking.booking_id,
                Booking.booking_date,
                Booking.status.label('booking_status'),
                Booking.pickup_location,
                Booking.drop_location,
                Booking.pickup_latitude,
                Booking.pickup_longitude,
                Booking.drop_latitude,
                Booking.drop_longitude,
                Booking.reason,
                Booking.employee_id,
                Booking.employee_code,
                Employee.name.label('employee_name'),
                Employee.phone.label('employee_phone'),
                Employee.gender.label('employee_gender'),
                Shift.shift_id,
                Shift.shift_code,
                Shift.shift_time,
                Shift.log_type.label('shift_type'),
                RouteManagement.route_id,
                RouteManagement.route_code,
                RouteManagement.status.label('route_status'),
                RouteManagementBooking.order_id,
                RouteManagementBooking.estimated_pick_up_time,
                RouteManagementBooking.estimated_drop_time,
                RouteManagementBooking.actual_pick_up_time,
                RouteManagementBooking.actual_drop_time,
                RouteManagementBooking.estimated_distance,
                RouteManagement.actual_total_distance,
                Driver.driver_id,
                Driver.name.label('driver_name'),
                Driver.phone.label('driver_phone'),
                Driver.license_number,
                Vehicle.vehicle_id,
                Vehicle.rc_number,
                VehicleType.name.label('vehicle_type_name'),
                Vendor.vendor_id,
                Vendor.name.label('vendor_name'),
                Vendor.phone.label('vendor_phone'),
                Escort.escort_id,
                Escort.name.label('escort_name'),
                Escort.phone.label('escort_phone'),
                Escort.gender.label('escort_gender')
            )
            .outerjoin(Employee, Booking.employee_id == Employee.employee_id)
            .join(Shift, Booking.shift_id == Shift.shift_id)
            .outerjoin(RouteManagementBooking, Booking.booking_id == RouteManagementBooking.booking_id)
            .outerjoin(RouteManagement, RouteManagementBooking.route_id == RouteManagement.route_id)
            .outerjoin(Driver, RouteManagement.assigned_driver_id == Driver.driver_id)
            .outerjoin(Vehicle, RouteManagement.assigned_vehicle_id == Vehicle.vehicle_id)
            .outerjoin(VehicleType, Vehicle.vehicle_type_id == VehicleType.vehicle_type_id)
            .outerjoin(Vendor, RouteManagement.assigned_vendor_id == Vendor.vendor_id)
            .outerjoin(Escort, RouteManagement.assigned_escort_id == Escort.escort_id)
            .filter(
                Booking.tenant_id == tenant_id,
                Booking.booking_date >= start_date,
                Booking.booking_date <= end_date
            )
        )

        # --- Apply Filters ---
        if shift_id:
            query = query.filter(Booking.shift_id == shift_id)

        if booking_status:
            query = query.filter(Booking.status.in_(booking_status))

        if route_status:
            query = query.filter(RouteManagement.status.in_(route_status))

        if vendor_id:
            query = query.filter(RouteManagement.assigned_vendor_id == vendor_id)

        if not include_unrouted:
            # Only include bookings that have routes
            query = query.filter(RouteManagement.route_id.isnot(None))

        # Order by date, shift, and route
        query = query.order_by(
            Booking.booking_date.desc(),
            Shift.shift_id,
            RouteManagement.route_id,
            RouteManagementBooking.order_id
        )

        # --- Execute Query ---
        results = query.all()

        logger.info(f"[export_bookings_report] Found {len(results)} records")

        # --- Create Excel Workbook ---
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Bookings Report"

        # Define headers
        headers = [
            "Route ID",
            "Booking ID",
            "Booking Date",
            "Booking Status",
            "Employee ID",
            "Employee Code",
            "Employee Name",
            "Employee Phone",
            "Employee Gender",
            "Shift Code",
            "Shift Time",
            "Shift Type",
            "Pickup Location",
            "Drop Location",
            "Route Status",
            "Stop Order",
            "Estimated Pickup Time",
            "Estimated Drop Time",
            "Actual Pickup Time",
            "Actual Drop Time",
            "Estimated Distance (km)",
            "Actual Total Distance (km)",
            "Driver Name",
            "Driver Phone",
            "Vehicle Number",
            "Vehicle Type",
            "Vendor Name",
            "Vendor Phone",
            "Escort Name",
            "Escort Phone",
            "Escort Gender",
            "Reason"
        ]

        # Apply styling
        ws = style_excel_report(ws, headers)

        # --- Group bookings by route and populate data ---
        from itertools import groupby

        # Sort results by route_id to group them
        sorted_results = sorted(results, key=lambda r: (r.route_id if r.route_id else 0))

        row_num = 2

        for _, group_records in groupby(sorted_results, key=lambda r: r.route_id):
            group_list = list(group_records)
            route_record_start = row_num

            for record in group_list:
                # Route first, then booking and employee details
                ws.cell(row=row_num, column=1, value=record.route_id or '')
                ws.cell(row=row_num, column=2, value=record.booking_id)
                ws.cell(row=row_num, column=3, value=record.booking_date.strftime('%Y-%m-%d') if record.booking_date else '')
                ws.cell(row=row_num, column=4, value=record.booking_status.value if record.booking_status else '')
                ws.cell(row=row_num, column=5, value=record.employee_id)
                ws.cell(row=row_num, column=6, value=record.employee_code or '')
                ws.cell(row=row_num, column=7, value=record.employee_name or '')
                ws.cell(row=row_num, column=8, value=record.employee_phone or '')
                ws.cell(row=row_num, column=9, value=record.employee_gender.value if record.employee_gender else '')
                ws.cell(row=row_num, column=10, value=record.shift_code or '')
                ws.cell(row=row_num, column=11, value=str(record.shift_time) if record.shift_time else '')
                ws.cell(row=row_num, column=12, value=record.shift_type.value if record.shift_type else '')
                ws.cell(row=row_num, column=13, value=record.pickup_location or '')
                ws.cell(row=row_num, column=14, value=record.drop_location or '')
                ws.cell(row=row_num, column=15, value=record.route_status.value if record.route_status else '')
                ws.cell(row=row_num, column=16, value=record.order_id if record.order_id is not None else '')
                ws.cell(row=row_num, column=17, value=str(record.estimated_pick_up_time) if record.estimated_pick_up_time else '')
                ws.cell(row=row_num, column=18, value=str(record.estimated_drop_time) if record.estimated_drop_time else '')
                ws.cell(row=row_num, column=19, value=str(record.actual_pick_up_time) if record.actual_pick_up_time else '')
                ws.cell(row=row_num, column=20, value=str(record.actual_drop_time) if record.actual_drop_time else '')
                ws.cell(row=row_num, column=21, value=record.estimated_distance)
                ws.cell(row=row_num, column=22, value=float(record.actual_total_distance) if record.actual_total_distance is not None else '')
                ws.cell(row=row_num, column=23, value=record.driver_name or '')
                ws.cell(row=row_num, column=24, value=record.driver_phone or '')
                ws.cell(row=row_num, column=25, value=record.rc_number or '')
                ws.cell(row=row_num, column=26, value=record.vehicle_type_name or '')
                ws.cell(row=row_num, column=27, value=record.vendor_name or '')
                ws.cell(row=row_num, column=28, value=record.vendor_phone or '')
                ws.cell(row=row_num, column=29, value=record.escort_name or '')
                ws.cell(row=row_num, column=30, value=record.escort_phone or '')
                ws.cell(row=row_num, column=31, value=record.escort_gender or '')
                ws.cell(row=row_num, column=32, value=record.reason or '')

                row_num += 1

            route_record_end = row_num - 1

            # Merge route-level columns for same route (except unrouted records)
            if len(group_list) > 1 and group_list[0].route_id is not None:
                merge_cols = [1, 15, 22, 23, 24, 25, 26, 27, 28, 29, 30, 31]
                for col in merge_cols:
                    ws.merge_cells(
                        f'{get_column_letter(col)}{route_record_start}:{get_column_letter(col)}{route_record_end}'
                    )
                    merged_cell = ws[f'{get_column_letter(col)}{route_record_start}']
                    merged_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

            # Add a clear border at each route end row
            route_end_border = Border(bottom=Side(style='medium'))
            for col in range(1, len(headers) + 1):
                cell = ws.cell(row=route_record_end, column=col)
                cell.border = route_end_border

        # --- Create Summary Sheet ---
        summary_ws = wb.create_sheet(title="Summary")
        
        # Summary statistics
        total_bookings = len(results)
        status_counts = {}
        for record in results:
            status = record.booking_status.value if record.booking_status else 'UNKNOWN'
            status_counts[status] = status_counts.get(status, 0) + 1

        routed_bookings = sum(1 for r in results if r.route_id is not None)
        unrouted_bookings = total_bookings - routed_bookings

        summary_data = [
            ["Report Summary", ""],
            ["Generated On", get_current_ist_time().strftime('%Y-%m-%d %H:%M:%S')],
            ["Generated By", f"User ID: {user_id}"],
            ["Tenant ID", tenant_id],
            ["Tenant Name", tenant.name],
            ["Date Range", f"{start_date} to {end_date}"],
            ["", ""],
            ["Total Bookings", total_bookings],
            ["Routed Bookings", routed_bookings],
            ["Unrouted Bookings", unrouted_bookings],
            ["", ""],
            ["Booking Status Breakdown", "Count"]
        ]

        for status, count in sorted(status_counts.items()):
            summary_data.append([status, count])

        for row_num, row_data in enumerate(summary_data, 1):
            for col_num, value in enumerate(row_data, 1):
                cell = summary_ws.cell(row=row_num, column=col_num, value=value)
                if row_num == 1 or (row_num == 12 and col_num <= 2):
                    cell.font = Font(bold=True, size=12)

        # Adjust column widths
        summary_ws.column_dimensions['A'].width = 30
        summary_ws.column_dimensions['B'].width = 30

        # --- Save to BytesIO ---
        output = BytesIO()
        wb.save(output)
        output.seek(0)

        # --- Generate filename ---
        filename = f"bookings_report_{tenant_id}_{start_date}_to_{end_date}.xlsx"

        if not results:
            logger.info(f"[export_bookings_report] No records found — returning empty report for user {user_id}")
        else:
            logger.info(f"[export_bookings_report] Generated report with {len(results)} records for user {user_id}")

        # --- Return as StreamingResponse ---
        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={
                "Content-Disposition": f'attachment; filename="{filename}"',
                "Content-Type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                "Access-Control-Expose-Headers": "Content-Disposition"
            }
        )

    except HTTPException:
        raise
    except Exception as e:
        logger.exception(f"[export_bookings_report] Error generating report: {e}")
        raise handle_db_error(e)


@router.get("/bookings/analytics", status_code=http_status.HTTP_200_OK)
@cached(ttl_seconds=300, key_prefix="analytics")  # Cache for 5 minutes
async def get_bookings_analytics(
    start_date: date = Query(..., description="Start date for analytics (YYYY-MM-DD)"),
    end_date: date = Query(..., description="End date for analytics (YYYY-MM-DD)"),
    tenant_id: Optional[str] = Query(None, description="Tenant ID"),
    shift_id: Optional[int] = Query(None, description="Filter by shift ID"),
    db: Session = Depends(get_db),
    user_data=Depends(PermissionChecker(["report.read"], check_tenant=True)),
):
    """
    Get analytics summary for bookings within a date range.
    
    Returns JSON with:
    - Total bookings by status
    - Routed vs unrouted breakdown
    - Route status distribution
    - Completion rates
    - Daily breakdown
    """
    try:
        user_type = user_data.get("user_type")
        token_tenant_id = user_data.get("tenant_id")
        user_id = user_data.get("user_id")

        logger.info(
            f"[get_bookings_analytics] user_id={user_id}, date_range={start_date} to {end_date}"
        )

        # --- Tenant Resolution ---
        if user_type == "employee":
            tenant_id = token_tenant_id
        elif user_type == "vendor":
            tenant_id = token_tenant_id
        elif user_type == "admin":
            if not tenant_id:
                raise HTTPException(
                    status_code=http_status.HTTP_400_BAD_REQUEST,
                    detail=ResponseWrapper.error(
                        message="tenant_id is required for admin users",
                        error_code="TENANT_ID_REQUIRED"
                    )
                )
        else:
            raise HTTPException(
                status_code=http_status.HTTP_403_FORBIDDEN,
                detail=ResponseWrapper.error(
                    message="Insufficient permissions",
                    error_code="FORBIDDEN"
                )
            )

        if not tenant_id:
            raise HTTPException(
                status_code=http_status.HTTP_403_FORBIDDEN,
                detail=ResponseWrapper.error(
                    message="Tenant context not available",
                    error_code="TENANT_ID_REQUIRED"
                )
            )

        # --- Validate Date Range ---
        validate_date_range(start_date, end_date)

        # --- Base Query ---
        base_query = db.query(Booking).filter(
            Booking.tenant_id == tenant_id,
            Booking.booking_date >= start_date,
            Booking.booking_date <= end_date
        )

        if shift_id:
            base_query = base_query.filter(Booking.shift_id == shift_id)

        # --- Booking Status Breakdown ---
        status_breakdown = (
            base_query
            .with_entities(
                Booking.status,
                func.count(Booking.booking_id).label('count')
            )
            .group_by(Booking.status)
            .all()
        )

        status_counts = {
            status.value if status else 'UNKNOWN': count
            for status, count in status_breakdown
        }

        # --- Routed vs Unrouted ---
        total_bookings = base_query.count()
        
        routed_count = (
            base_query
            .join(RouteManagementBooking, Booking.booking_id == RouteManagementBooking.booking_id)
            .join(RouteManagement, RouteManagementBooking.route_id == RouteManagement.route_id)
            .filter(RouteManagement.tenant_id == tenant_id)
            .distinct()
            .count()
        )
        
        unrouted_count = total_bookings - routed_count

        # --- Route Status Breakdown (for routed bookings) ---
        route_status_breakdown = (
            db.query(
                RouteManagement.status,
                func.count(func.distinct(Booking.booking_id)).label('booking_count')
            )
            .join(RouteManagementBooking, RouteManagement.route_id == RouteManagementBooking.route_id)
            .join(Booking, RouteManagementBooking.booking_id == Booking.booking_id)
            .filter(
                RouteManagement.tenant_id == tenant_id,
                Booking.booking_date >= start_date,
                Booking.booking_date <= end_date
            )
            .group_by(RouteManagement.status)
            .all()
        )

        route_status_counts = {
            status.value if status else 'UNKNOWN': count
            for status, count in route_status_breakdown
        }

        # --- Daily Breakdown ---
        daily_breakdown = (
            base_query
            .with_entities(
                Booking.booking_date,
                Booking.status,
                func.count(Booking.booking_id).label('count')
            )
            .group_by(Booking.booking_date, Booking.status)
            .order_by(Booking.booking_date)
            .all()
        )

        daily_data = {}
        for booking_date, status, count in daily_breakdown:
            date_str = booking_date.strftime('%Y-%m-%d')
            if date_str not in daily_data:
                daily_data[date_str] = {
                    "booking_status": {},
                    "vendor_assigned": 0,
                    "driver_assigned": 0
                }
            status_str = status.value if status else 'UNKNOWN'
            daily_data[date_str]["booking_status"][status_str] = count

        # --- Daily Vendor and Driver Assignment Breakdown ---
        assignment_daily_breakdown = (
            db.query(
                Booking.booking_date,

                func.count(
                    func.distinct(
                        case(
                            (
                                RouteManagement.assigned_vendor_id.isnot(None),
                                Booking.booking_id,
                            ),
                            else_=None,
                        )
                    )
                ).label("vendor_assigned_count"),

                func.count(
                    func.distinct(
                        case(
                            (
                                RouteManagement.assigned_driver_id.isnot(None),
                                Booking.booking_id,
                            ),
                            else_=None,
                        )
                    )
                ).label("driver_assigned_count"),
            )
            .join(
                RouteManagementBooking,
                Booking.booking_id == RouteManagementBooking.booking_id,
            )
            .join(
                RouteManagement,
                RouteManagementBooking.route_id == RouteManagement.route_id,
            )
            .filter(
                RouteManagement.tenant_id == tenant_id,
                Booking.booking_date >= start_date,
                Booking.booking_date <= end_date,
            )
            .group_by(Booking.booking_date)
            .all()
        )

        for booking_date, vendor_count, driver_count in assignment_daily_breakdown:
            date_str = booking_date.strftime('%Y-%m-%d')
            if date_str not in daily_data:
                daily_data[date_str] = {
                    "booking_status": {},
                    "vendor_assigned": 0,
                    "driver_assigned": 0,
                }
            daily_data[date_str]["vendor_assigned"] = vendor_count or 0
            daily_data[date_str]["driver_assigned"] = driver_count or 0

        # --- Completion Rate ---
        completed_count = status_counts.get('Completed', 0)
        completion_rate = (completed_count / total_bookings * 100) if total_bookings > 0 else 0

        # --- Vendor Assignment Count ---
        vendor_assigned_count = (
            db.query(func.count(func.distinct(Booking.booking_id)))
            .join(RouteManagementBooking, Booking.booking_id == RouteManagementBooking.booking_id)
            .join(RouteManagement, RouteManagementBooking.route_id == RouteManagement.route_id)
            .filter(
                RouteManagement.tenant_id == tenant_id,
                RouteManagement.assigned_vendor_id.isnot(None),
                Booking.booking_date >= start_date,
                Booking.booking_date <= end_date
            )
            .scalar() or 0
        )

        # --- Driver Assignment Count ---
        driver_assigned_count = (
            db.query(func.count(func.distinct(Booking.booking_id)))
            .join(RouteManagementBooking, Booking.booking_id == RouteManagementBooking.booking_id)
            .join(RouteManagement, RouteManagementBooking.route_id == RouteManagement.route_id)
            .filter(
                RouteManagement.tenant_id == tenant_id,
                RouteManagement.assigned_driver_id.isnot(None),
                Booking.booking_date >= start_date,
                Booking.booking_date <= end_date
            )
            .scalar() or 0
        )

        # --- Total Unique Shifts ---
        total_shifts = (
            base_query
            .with_entities(func.count(func.distinct(Booking.shift_id)))
            .scalar() or 0
        )

        # --- Response ---
        analytics = {
            "date_range": {
                "start_date": start_date.strftime('%Y-%m-%d'),
                "end_date": end_date.strftime('%Y-%m-%d')
            },
            "total_bookings": total_bookings,
            "total_shifts": total_shifts,
            "booking_status_breakdown": status_counts,
            "routing_summary": {
                "routed": routed_count,
                "unrouted": unrouted_count,
                "routing_percentage": (routed_count / total_bookings * 100) if total_bookings > 0 else 0
            },
            "assignment_summary": {
                "vendor_assigned": vendor_assigned_count,
                "driver_assigned": driver_assigned_count,
                "vendor_assignment_percentage": (vendor_assigned_count / total_bookings * 100) if total_bookings > 0 else 0,
                "driver_assignment_percentage": (driver_assigned_count / total_bookings * 100) if total_bookings > 0 else 0
            },
            "route_status_breakdown": route_status_counts,
            "completion_rate": round(completion_rate, 2),
            "daily_breakdown": daily_data
        }

        return ResponseWrapper.success(
            data=analytics,
            message="Analytics generated successfully"
        )

    except HTTPException:
        raise
    except Exception as e:
        logger.exception(f"[get_bookings_analytics] Error: {e}")
        raise handle_db_error(e)


# ---------------------------------------------------------------------------
# Feature 6: OTA / OTD Delay Reports
# ---------------------------------------------------------------------------

@router.get("/delays", status_code=http_status.HTTP_200_OK)
async def get_delay_report(
    start_date: date = Query(..., description="Start date (YYYY-MM-DD)"),
    end_date: date = Query(..., description="End date (YYYY-MM-DD)"),
    delay_type: Optional[str] = Query(None, description="Filter: LATE | EARLY | ON_TIME"),
    delay_category: Optional[str] = Query(
        None,
        description="Filter by root-cause category: DRIVER_DELAY | EMPLOYEE_DELAY | TRAFFIC_DELAY | NONE",
    ),
    tenant_id: Optional[str] = Query(None, description="Tenant ID (admin only; employees use token)"),
    db: Session = Depends(get_db),
    ctx=Depends(PermissionChecker(["report.read"], check_tenant=True)),
):
    """
    Return a summary of OTD delay-tagged routes for the given date range.

    Each row includes route metadata and the latest delay tag.
    Results are ordered by route creation date descending.

    Query params
    ------------
    start_date     – inclusive lower bound on route.delay_tagged_at date
    end_date       – inclusive upper bound
    delay_type     – optional filter (LATE | EARLY | ON_TIME)
    delay_category – optional filter (DRIVER_DELAY | EMPLOYEE_DELAY |
                     TRAFFIC_DELAY | NONE)
    tenant_id      – admin-only override; otherwise resolved from token
    """
    try:
        user_data = ctx
        token_tenant_id = user_data.get("tenant_id")

        # Resolve tenant
        resolved_tenant_id = tenant_id or token_tenant_id
        if not resolved_tenant_id:
            raise HTTPException(
                status_code=http_status.HTTP_403_FORBIDDEN,
                detail=ResponseWrapper.error(
                    message="Tenant context not available",
                    error_code="TENANT_ID_MISSING",
                ),
            )

        validate_date_range(start_date, end_date)

        start_dt = datetime.combine(start_date, datetime.min.time())
        end_dt = datetime.combine(end_date, datetime.max.time().replace(microsecond=0))

        query = (
            db.query(RouteManagement)
            .filter(
                RouteManagement.tenant_id == resolved_tenant_id,
                RouteManagement.delay_tagged_at.isnot(None),
                RouteManagement.delay_tagged_at >= start_dt,
                RouteManagement.delay_tagged_at <= end_dt,
            )
        )

        if delay_type:
            query = query.filter(RouteManagement.delay_type == delay_type.upper())

        routes = query.order_by(RouteManagement.delay_tagged_at.desc()).all()

        # Resolve the latest delay_category from route_delay_events when needed
        route_ids = [r.route_id for r in routes]

        # Build a map: route_id → latest delay_category from route_delay_events
        category_map: dict = {}
        if route_ids:
            from sqlalchemy import func as sqlfunc
            latest_events = (
                db.query(
                    RouteDelayEvent.route_id,
                    RouteDelayEvent.delay_category,
                )
                .filter(
                    RouteDelayEvent.route_id.in_(route_ids),
                    RouteDelayEvent.event_kind == "OTD",
                )
                .order_by(
                    RouteDelayEvent.route_id,
                    RouteDelayEvent.tagged_at.desc(),
                )
                .all()
            )
            # Keep only the first (most recent) per route_id
            seen: set = set()
            for ev_route_id, ev_category in latest_events:
                if ev_route_id not in seen:
                    category_map[ev_route_id] = ev_category
                    seen.add(ev_route_id)

        rows = []
        for r in routes:
            cat = category_map.get(r.route_id)
            # Apply delay_category filter if requested
            if delay_category and (cat or "").upper() != delay_category.upper():
                continue
            rows.append({
                "route_id": r.route_id,
                "route_code": r.route_code,
                "shift_id": r.shift_id,
                "status": r.status.value if r.status else None,
                "assigned_driver_id": r.assigned_driver_id,
                "actual_start_time": r.actual_start_time.isoformat() if r.actual_start_time else None,
                "actual_end_time": r.actual_end_time.isoformat() if r.actual_end_time else None,
                "estimated_total_time_min": r.estimated_total_time,
                "delay_type": r.delay_type,
                "delay_minutes": r.delay_minutes,
                "delay_category": cat,
                "delay_tagged_at": r.delay_tagged_at.isoformat() if r.delay_tagged_at else None,
                "ota_grace_minutes": r.ota_grace_minutes,
            })

        # Aggregate summary
        total = len(rows)
        late_count = sum(1 for row in rows if row["delay_type"] == "LATE")
        early_count = sum(1 for row in rows if row["delay_type"] == "EARLY")
        on_time_count = sum(1 for row in rows if row["delay_type"] == "ON_TIME")
        avg_delay = (
            round(sum(row["delay_minutes"] for row in rows if row["delay_minutes"] is not None) / total, 1)
            if total else 0
        )

        # Category breakdown
        driver_delay_count   = sum(1 for row in rows if row["delay_category"] == "DRIVER_DELAY")
        employee_delay_count = sum(1 for row in rows if row["delay_category"] == "EMPLOYEE_DELAY")
        traffic_delay_count  = sum(1 for row in rows if row["delay_category"] == "TRAFFIC_DELAY")
        none_count           = sum(1 for row in rows if row["delay_category"] in ("NONE", None))

        return ResponseWrapper.success(
            data={
                "summary": {
                    "total_routes_tagged": total,
                    "late": late_count,
                    "early": early_count,
                    "on_time": on_time_count,
                    "average_delay_minutes": avg_delay,
                    "by_category": {
                        "DRIVER_DELAY": driver_delay_count,
                        "EMPLOYEE_DELAY": employee_delay_count,
                        "TRAFFIC_DELAY": traffic_delay_count,
                        "NONE": none_count,
                    },
                },
                "routes": rows,
            },
            message="Delay report generated successfully",
        )

    except HTTPException:
        raise
    except Exception as e:
        logger.exception("[get_delay_report] Error: %s", e)
        raise handle_db_error(e)


@router.get("/delays/{route_id}", status_code=http_status.HTTP_200_OK)
async def get_route_delay_detail(
    route_id: int,
    tenant_id: Optional[str] = Query(None, description="Tenant ID (admin only)"),
    db: Session = Depends(get_db),
    ctx=Depends(PermissionChecker(["report.read"], check_tenant=True)),
):
    """
    Return the full delay event history for a specific route.

    Each row in `events` corresponds to one RouteDelayEvent (OTA or OTD).
    Also returns the summary delay fields from the route itself.
    """
    try:
        user_data = ctx
        token_tenant_id = user_data.get("tenant_id")
        resolved_tenant_id = tenant_id or token_tenant_id

        if not resolved_tenant_id:
            raise HTTPException(
                status_code=http_status.HTTP_403_FORBIDDEN,
                detail=ResponseWrapper.error(
                    message="Tenant context not available",
                    error_code="TENANT_ID_MISSING",
                ),
            )

        route = (
            db.query(RouteManagement)
            .filter(
                RouteManagement.route_id == route_id,
                RouteManagement.tenant_id == resolved_tenant_id,
            )
            .first()
        )
        if not route:
            raise HTTPException(
                status_code=http_status.HTTP_404_NOT_FOUND,
                detail=ResponseWrapper.error(
                    message=f"Route {route_id} not found",
                    error_code="ROUTE_NOT_FOUND",
                ),
            )

        events = (
            db.query(RouteDelayEvent)
            .filter(RouteDelayEvent.route_id == route_id)
            .order_by(RouteDelayEvent.tagged_at.asc())
            .all()
        )

        event_list = [
            {
                "id": ev.id,
                "event_kind": ev.event_kind,
                "delay_type": ev.delay_type,
                "delay_minutes": ev.delay_minutes,
                "delay_category": ev.delay_category,
                "notes": ev.notes,
                "tagged_at": ev.tagged_at.isoformat() if ev.tagged_at else None,
            }
            for ev in events
        ]

        return ResponseWrapper.success(
            data={
                "route_id": route.route_id,
                "route_code": route.route_code,
                "shift_id": route.shift_id,
                "status": route.status.value if route.status else None,
                "actual_start_time": route.actual_start_time.isoformat() if route.actual_start_time else None,
                "actual_end_time": route.actual_end_time.isoformat() if route.actual_end_time else None,
                "estimated_total_time_min": route.estimated_total_time,
                "delay_summary": {
                    "delay_type": route.delay_type,
                    "delay_minutes": route.delay_minutes,
                    "delay_tagged_at": route.delay_tagged_at.isoformat() if route.delay_tagged_at else None,
                    "ota_grace_minutes": route.ota_grace_minutes,
                },
                "events": event_list,
            },
            message="Route delay detail fetched successfully",
        )

    except HTTPException:
        raise
    except Exception as e:
        logger.exception("[get_route_delay_detail] Error: %s", e)
        raise handle_db_error(e)


# ---------------------------------------------------------------------------
# GET /reports/driver-duty-hours
# Feature 1 — Driver Duty Hours & Rest-Time Enforcement
# ---------------------------------------------------------------------------

@router.get("/driver-duty-hours", status_code=http_status.HTTP_200_OK)
def get_driver_duty_hours_report(
    start_date: date = Query(..., description="Start date (inclusive)"),
    end_date: date = Query(..., description="End date (inclusive)"),
    driver_id: Optional[int] = Query(None, description="Filter by driver ID"),
    db: Session = Depends(get_db),
    user_data=Depends(PermissionChecker(["report.read"], check_tenant=True)),
):
    """
    Report: hours driven per driver within the date range, with rest-violation flags.

    For every completed route in the window, the endpoint sums duty minutes
    per driver per day.  A violation is flagged when the driver's longest
    continuous rest gap within any 24-hour window is shorter than
    ``required_rest_minutes`` (derived from ``tenant_configs.driver_max_duty_minutes``).

    Response shape
    --------------
    {
        "data": {
            "drivers": [
                {
                    "driver_id":              int,
                    "driver_name":            str,
                    "total_duty_minutes":     int,
                    "total_routes":           int,
                    "rest_violations":        int,   // # trips where rest was insufficient
                    "routes": [
                        {
                            "route_id":       int,
                            "route_code":     str | null,
                            "actual_start":   str | null,
                            "actual_end":     str | null,
                            "duty_minutes":   int,
                            "rest_ok":        bool,
                            "rest_gap_minutes": int
                        }, ...
                    ]
                }, ...
            ],
            "summary": {
                "total_drivers":   int,
                "total_routes":    int,
                "total_violations": int,
                "driver_max_duty_minutes": int
            }
        }
    }
    """
    try:
        validate_date_range(start_date, end_date)

        tenant_id: str = user_data.get("tenant_id")

        # Load tenant config for duty limits
        from app.models.tenant_config import TenantConfig
        from app.services.driver_duty_hours_service import check_rest

        tenant_cfg = db.query(TenantConfig).filter(
            TenantConfig.tenant_id == tenant_id
        ).first()
        max_duty = tenant_cfg.driver_max_duty_minutes if tenant_cfg else 600

        # Convert dates to datetimes for comparison
        window_start_dt = datetime.combine(start_date, datetime.min.time())
        window_end_dt   = datetime.combine(end_date,   datetime.max.time())

        # Query completed routes in the date window
        routes_q = (
            db.query(RouteManagement)
            .filter(
                RouteManagement.tenant_id == tenant_id,
                RouteManagement.status == RouteManagementStatusEnum.COMPLETED,
                RouteManagement.actual_start_time.isnot(None),
                RouteManagement.actual_start_time >= window_start_dt,
                RouteManagement.actual_start_time <= window_end_dt,
                RouteManagement.assigned_driver_id.isnot(None),
            )
        )
        if driver_id is not None:
            routes_q = routes_q.filter(RouteManagement.assigned_driver_id == driver_id)

        routes = routes_q.order_by(RouteManagement.actual_start_time.asc()).all()

        # Collect unique driver IDs and batch-fetch their records
        driver_ids = list({r.assigned_driver_id for r in routes})
        drivers = (
            db.query(Driver)
            .filter(Driver.driver_id.in_(driver_ids))
            .all()
        ) if driver_ids else []
        driver_map: dict[int, Driver] = {d.driver_id: d for d in drivers}

        # Group routes by driver
        from collections import defaultdict
        by_driver: dict[int, list] = defaultdict(list)
        for route in routes:
            by_driver[route.assigned_driver_id].append(route)

        driver_rows = []
        total_violations_global = 0

        for did, driver_routes in by_driver.items():
            driver_obj = driver_map.get(did)
            driver_name = driver_obj.name if driver_obj else f"Driver #{did}"

            total_duty = 0
            violations = 0
            route_items = []

            for route in driver_routes:
                start_dt = route.actual_start_time
                end_dt   = route.actual_end_time

                duty_min = 0
                if start_dt and end_dt:
                    duty_min = max(0, int((end_dt - start_dt).total_seconds() / 60))
                total_duty += duty_min

                # Per-trip rest check: was rest sufficient before THIS trip?
                rest_result = check_rest(
                    driver_id=did,
                    proposed_start_dt=start_dt,
                    db=db,
                    max_duty_minutes=max_duty,
                )
                if not rest_result["ok"]:
                    violations += 1

                route_items.append({
                    "route_id":         route.route_id,
                    "route_code":       route.route_code,
                    "actual_start":     start_dt.isoformat() if start_dt else None,
                    "actual_end":       end_dt.isoformat()   if end_dt   else None,
                    "duty_minutes":     duty_min,
                    "rest_ok":          rest_result["ok"],
                    "rest_gap_minutes": rest_result["rest_gap_minutes"],
                })

            total_violations_global += violations
            driver_rows.append({
                "driver_id":          did,
                "driver_name":        driver_name,
                "total_duty_minutes": total_duty,
                "total_routes":       len(driver_routes),
                "rest_violations":    violations,
                "routes":             route_items,
            })

        # Sort by driver name for consistent output
        driver_rows.sort(key=lambda x: x["driver_name"])

        return ResponseWrapper.success(
            data={
                "drivers": driver_rows,
                "summary": {
                    "total_drivers":           len(driver_rows),
                    "total_routes":            len(routes),
                    "total_violations":        total_violations_global,
                    "driver_max_duty_minutes": max_duty,
                },
            },
            message="Driver duty hours report fetched successfully",
        )

    except HTTPException:
        raise
    except Exception as e:
        logger.exception("[get_driver_duty_hours_report] Error: %s", e)
        raise handle_db_error(e)
