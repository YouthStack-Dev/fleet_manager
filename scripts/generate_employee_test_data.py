"""
Script to generate sample Excel files for testing bulk employee upload.
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from datetime import datetime


def create_sample_template():
    """Create a sample Excel template with headers and example data"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Employees"
    
    # Headers
    headers = ['name', 'email', 'phone', 'employee_code', 'team_id', 
               'address', 'latitude', 'longitude', 'password']
    
    # Style headers
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Add instructions row (row 2)
    instructions = [
        'Required',
        'Required (unique)',
        'Required (unique, 10+ digits)',
        'Optional (unique)',
        'Required (must exist)',
        'Optional',
        'Optional (-90 to 90)',
        'Optional (-180 to 180)',
        'Optional (min 6 chars)'
    ]
    
    instruction_fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
    instruction_font = Font(italic=True, size=9)
    
    for col_num, instruction in enumerate(instructions, 1):
        cell = ws.cell(row=2, column=col_num)
        cell.value = instruction
        cell.fill = instruction_fill
        cell.font = instruction_font
        cell.alignment = Alignment(horizontal="left", vertical="center")
    
    # Add sample data (rows 3-7)
    sample_data = [
        ['John Doe', 'john.doe@company.com', '+1-234-567-8901', 'EMP001', '1', 
         '123 Main St, City, State', '37.7749', '-122.4194', 'Welcome@123'],
        ['Jane Smith', 'jane.smith@company.com', '+1-234-567-8902', 'EMP002', '1',
         '456 Oak Ave, City, State', '34.0522', '-118.2437', 'Secure@456'],
        ['Bob Johnson', 'bob.johnson@company.com', '+1-234-567-8903', 'EMP003', '2',
         '789 Pine Rd, City, State', '40.7128', '-74.0060', 'Strong#789'],
        ['Alice Williams', 'alice.williams@company.com', '+1-234-567-8904', 'EMP004', '2',
         '321 Elm Blvd, City, State', '41.8781', '-87.6298', 'Password@101'],
        ['Charlie Brown', 'charlie.brown@company.com', '+1-234-567-8905', 'EMP005', '1',
         '654 Maple Dr, City, State', '29.7604', '-95.3698', 'MyPass@202'],
    ]
    
    for row_num, row_data in enumerate(sample_data, 3):
        for col_num, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_num, column=col_num)
            cell.value = value
            cell.alignment = Alignment(horizontal="left", vertical="center")
    
    # Adjust column widths
    column_widths = [20, 30, 20, 15, 10, 30, 12, 12, 15]
    for col_num, width in enumerate(column_widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_num)].width = width
    
    # Save template
    filename = f"employee_bulk_upload_template_{datetime.now().strftime('%Y%m%d')}.xlsx"
    wb.save(filename)
    print(f"✅ Sample template created: {filename}")
    return filename


def create_test_data_valid():
    """Create test Excel file with valid data"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Valid Employees"
    
    headers = ['name', 'email', 'phone', 'employee_code', 'team_id', 
               'address', 'latitude', 'longitude', 'password']
    ws.append(headers)
    
    # 10 valid employees
    for i in range(1, 11):
        ws.append([
            f'Test Employee {i}',
            f'test.emp{i}@company.com',
            f'+1-555-010{i:02d}',
            f'TEST{i:03d}',
            '1',
            f'{i*100} Test Street, Test City',
            str(37.7749 + i * 0.01),
            str(-122.4194 + i * 0.01),
            'Test@123'
        ])
    
    filename = "test_data_valid.xlsx"
    wb.save(filename)
    print(f"✅ Valid test data created: {filename}")
    return filename


def create_test_data_invalid():
    """Create test Excel file with various validation errors"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Invalid Employees"
    
    headers = ['name', 'email', 'phone', 'employee_code', 'team_id', 
               'address', 'latitude', 'longitude', 'password']
    ws.append(headers)
    
    # Various error cases
    test_cases = [
        # Missing name
        ['', 'error1@test.com', '+1234567890', 'ERR001', '1', 'Address', '37.7', '-122.4', 'Pass@123'],
        
        # Invalid email
        ['Error User 2', 'invalid-email', '+1234567891', 'ERR002', '1', 'Address', '37.7', '-122.4', 'Pass@123'],
        
        # Invalid phone (too short)
        ['Error User 3', 'error3@test.com', '123', 'ERR003', '1', 'Address', '37.7', '-122.4', 'Pass@123'],
        
        # Missing team_id
        ['Error User 4', 'error4@test.com', '+1234567893', 'ERR004', '', 'Address', '37.7', '-122.4', 'Pass@123'],
        
        # Invalid latitude
        ['Error User 5', 'error5@test.com', '+1234567894', 'ERR005', '1', 'Address', '200', '-122.4', 'Pass@123'],
        
        # Invalid longitude
        ['Error User 6', 'error6@test.com', '+1234567895', 'ERR006', '1', 'Address', '37.7', 'invalid', 'Pass@123'],
        
        # Short password
        ['Error User 7', 'error7@test.com', '+1234567896', 'ERR007', '1', 'Address', '37.7', '-122.4', '123'],
        
        # Duplicate email (both rows have same email)
        ['Error User 8A', 'duplicate@test.com', '+1234567897', 'ERR008A', '1', 'Address', '37.7', '-122.4', 'Pass@123'],
        ['Error User 8B', 'duplicate@test.com', '+1234567898', 'ERR008B', '1', 'Address', '37.7', '-122.4', 'Pass@123'],
    ]
    
    for row_data in test_cases:
        ws.append(row_data)
    
    filename = "test_data_invalid.xlsx"
    wb.save(filename)
    print(f"✅ Invalid test data created: {filename}")
    return filename


def create_test_data_mixed():
    """Create test Excel file with mix of valid and invalid data"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Mixed Employees"
    
    headers = ['name', 'email', 'phone', 'employee_code', 'team_id', 
               'address', 'latitude', 'longitude', 'password']
    ws.append(headers)
    
    # 5 valid + 3 invalid
    data = [
        # Valid
        ['Valid User 1', 'valid1@test.com', '+1234567801', 'VAL001', '1', 'Address 1', '37.7', '-122.4', 'Pass@123'],
        ['Valid User 2', 'valid2@test.com', '+1234567802', 'VAL002', '1', 'Address 2', '37.8', '-122.5', 'Pass@123'],
        
        # Invalid email
        ['Invalid User 1', 'not-an-email', '+1234567803', 'INV001', '1', 'Address 3', '37.9', '-122.6', 'Pass@123'],
        
        # Valid
        ['Valid User 3', 'valid3@test.com', '+1234567804', 'VAL003', '1', 'Address 4', '38.0', '-122.7', 'Pass@123'],
        
        # Invalid team
        ['Invalid User 2', 'invalid2@test.com', '+1234567805', 'INV002', '99999', 'Address 5', '38.1', '-122.8', 'Pass@123'],
        
        # Valid
        ['Valid User 4', 'valid4@test.com', '+1234567806', 'VAL004', '1', 'Address 6', '38.2', '-122.9', 'Pass@123'],
        ['Valid User 5', 'valid5@test.com', '+1234567807', 'VAL005', '1', 'Address 7', '38.3', '-123.0', 'Pass@123'],
        
        # Missing required field
        ['Invalid User 3', 'invalid3@test.com', '', 'INV003', '1', 'Address 8', '38.4', '-123.1', 'Pass@123'],
    ]
    
    for row_data in data:
        ws.append(row_data)
    
    filename = "test_data_mixed.xlsx"
    wb.save(filename)
    print(f"✅ Mixed test data created: {filename}")
    return filename


def create_test_data_large():
    """Create test Excel file with large dataset (near limit)"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Large Dataset"
    
    headers = ['name', 'email', 'phone', 'employee_code', 'team_id', 
               'address', 'latitude', 'longitude', 'password']
    ws.append(headers)
    
    # Create 450 employees (below 500 limit)
    for i in range(1, 451):
        ws.append([
            f'Employee {i:04d}',
            f'employee{i:04d}@company.com',
            f'+1-555-{i:04d}',
            f'EMP{i:04d}',
            '1',
            f'{i} Employee Street, Test City',
            str(37.7749 + (i * 0.001)),
            str(-122.4194 + (i * 0.001)),
            'Secure@123'
        ])
    
    filename = "test_data_large.xlsx"
    wb.save(filename)
    print(f"✅ Large test data created: {filename} (450 employees)")
    return filename


def create_empty_template():
    """Create an empty template for users to fill"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Employees"
    
    headers = ['name', 'email', 'phone', 'employee_code', 'team_id', 
               'address', 'latitude', 'longitude', 'password']
    
    # Style headers
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Adjust column widths
    column_widths = [20, 30, 20, 15, 10, 30, 12, 12, 15]
    for col_num, width in enumerate(column_widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_num)].width = width
    
    filename = "employee_upload_template_empty.xlsx"
    wb.save(filename)
    print(f"✅ Empty template created: {filename}")
    return filename


if __name__ == "__main__":
    print("\n🔧 Generating Excel test files for bulk employee upload...\n")
    
    # Create all test files
    create_sample_template()
    create_empty_template()
    create_test_data_valid()
    create_test_data_invalid()
    create_test_data_mixed()
    create_test_data_large()
    
    print("\n✅ All test files generated successfully!")
    print("\n📋 Files created:")
    print("   1. employee_bulk_upload_template_YYYYMMDD.xlsx - Sample template with examples")
    print("   2. employee_upload_template_empty.xlsx - Empty template to fill")
    print("   3. test_data_valid.xlsx - 10 valid employees")
    print("   4. test_data_invalid.xlsx - Various validation errors")
    print("   5. test_data_mixed.xlsx - Mix of valid and invalid data")
    print("   6. test_data_large.xlsx - 450 employees (near limit)")
    print("\n📝 Usage:")
    print("   Use these files to test the /api/v1/employees/bulk-upload endpoint")
    print("   curl -X POST http://localhost:8000/api/v1/employees/bulk-upload \\")
    print("        -H 'Authorization: Bearer YOUR_TOKEN' \\")
    print("        -F 'file=@test_data_valid.xlsx'")
    print()
